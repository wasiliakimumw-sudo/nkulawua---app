from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseRedirect, JsonResponse
from django.contrib.auth import login, logout, authenticate
import os
from decimal import Decimal
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q, F
from django.db import models, transaction
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.utils import timezone
import uuid
from datetime import timedelta, date
from django import forms
from .models import GalleryImage, Service, LandingPageSettings
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import (
    Account, Beneficiary, Vendor, Invoice, InvoiceItem, Expense, ExpenseItem,
    JournalEntry, JournalEntryLine, Payment, Budget, BudgetLine, ActivityLog,
    UserProfile, OpeningBalance, YearEndRollover, Scheme, Village, VillagePopulation,
    BoardOfTrustees, GeneralAssemblyMember, Employee, EmployeeSalary, Report,
    BeneficiaryHistory, BeneficiaryStatusLog, LoginSession, UserMessage,
    CommunicationLog, UserCall, DataMigrationLog, SystemVersion, SystemUpdateLog,
    BalanceHistory, TaxRate, DeletedRecord
)
from .forms import (
    AccountForm, BeneficiaryForm, VendorForm, InvoiceForm, InvoiceItemForm,
    ExpenseForm, ExpenseItemForm, PaymentForm, JournalEntryForm, JournalEntryLineForm, BudgetForm, BudgetLineFormSet, UserForm, UserProfileForm, EditUserForm,
    SchemeForm, VillageForm, VillagePopulationForm,
    BoardOfTrusteesForm, GeneralAssemblyMemberForm, EmployeeForm, ReportForm
)
from .ai_service import GroqAIService
import json


def save_deleted_record(obj, user):
    model_name = obj.__class__.__name__
    data = {}
    for field in obj._meta.fields:
        value = getattr(obj, field.name, None)
        if hasattr(value, 'isoformat'):
            value = value.isoformat()
        elif hasattr(value, 'pk'):
            value = value.pk
        data[field.name] = value
    
    DeletedRecord.objects.create(
        model_name=model_name,
        object_id=str(obj.pk),
        data=json.dumps(data, default=str),
        deleted_by=user
    )


def landing_page(request):
    if request.user.is_authenticated:
        return redirect("dashboard")
    gallery_images = GalleryImage.objects.filter(is_active=True)[:12]
    services = Service.objects.filter(is_active=True)
    landing_settings = LandingPageSettings.objects.filter(is_active=True).first()
    if not landing_settings:
        landing_settings = LandingPageSettings.objects.create()
    return render(request, "accounting_app/landing.html", {
        "gallery_images": gallery_images,
        "services": services,
        "landing_settings": landing_settings
    })


@login_required
def superuser_view(request):
    if not (request.user.is_superuser or (hasattr(request.user, 'userprofile') and (
        request.user.userprofile.role in ['admin', 'manager'] or
        request.user.userprofile.has_settings_access or
        request.user.userprofile.has_user_management_access
    ))):
        messages.error(request, "Access denied. Admin access required.")
        return redirect("dashboard")
    return render(request, "accounting_app/superuser_view.html")


def generate_invoice_number():
    today = timezone.now()
    prefix = f"INV-{today.strftime('%Y%m%d')}"
    last_invoice = Invoice.objects.filter(invoice_number__startswith=prefix).order_by('-invoice_number').first()
    if last_invoice:
        last_num = int(last_invoice.invoice_number.split('-')[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    return f"{prefix}-{new_num:04d}"


def generate_expense_number():
    today = timezone.now()
    prefix = f"EXP-{today.strftime('%Y%m%d')}"
    last_expense = Expense.objects.filter(expense_number__startswith=prefix).order_by('-expense_number').first()
    if last_expense:
        last_num = int(last_expense.expense_number.split('-')[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    return f"{prefix}-{new_num:04d}"


def login_view(request):
    if request.user.is_authenticated:
        return redirect("dashboard")
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            
            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            ip = x_forwarded_for.split(',')[0] if x_forwarded_for else request.META.get('REMOTE_ADDR')
            user_agent = request.META.get('HTTP_USER_AGENT', '')
            device_info = LoginSession.parse_user_agent(user_agent)
            
            LoginSession.objects.create(
                user=user,
                session_key=request.session.session_key,
                ip_address=ip,
                user_agent=user_agent,
                device_info=device_info
            )

            ActivityLog.objects.create(
                user=user,
                action="Login",
                model_name="Auth",
                description=f"User {user.username} logged in",
                ip_address=ip,
            )
            
            messages.success(request, f"Welcome back, {user.username}!")
            next_url = request.GET.get('next') or request.POST.get('next')
            if next_url:
                return redirect(next_url)
            if hasattr(user, 'userprofile'):
                p = user.userprofile
                if p.role in ('admin', 'manager'):
                    return redirect("superuser_view")
                return redirect("dashboard")
            return redirect("dashboard")
        else:
            messages.error(request, "Invalid username or password")
    return render(request, "accounting_app/login.html")


def logout_view(request):
    session_key = request.session.session_key
    user = request.user if request.user.is_authenticated else None
    username = user.username if user else "Unknown"
    if session_key:
        LoginSession.objects.filter(session_key=session_key, is_active=True).update(
            logout_time=timezone.now(),
            is_active=False
        )
    if user:
        ActivityLog.objects.create(
            user=user,
            action="Logout",
            model_name="Auth",
            description=f"User {username} logged out",
            ip_address=request.META.get('REMOTE_ADDR'),
        )
    logout(request)
    messages.success(request, "You have been logged out")
    return redirect("login")


@login_required
def dashboard(request):
    today = timezone.now().date()
    month_start = today.replace(day=1)
    year_start = today.replace(month=1, day=1)

    total_revenue = Payment.objects.filter(payment_date__gte=year_start).aggregate(
        total=Sum("amount"))["total"] or 0
    total_expenses = Expense.objects.filter(expense_date__gte=year_start).aggregate(
        total=Sum("amount"))["total"] or 0
    total_invoiced = Invoice.objects.filter(issue_date__gte=year_start).aggregate(
        total=Sum("total_amount"))["total"] or 0
    total_invoiced_all = Invoice.objects.all().aggregate(
        total=Sum("total_amount"))["total"] or 0
    total_revenue_all = Payment.objects.all().aggregate(
        total=Sum("amount"))["total"] or 0
    outstanding_invoices = Invoice.objects.filter(
        status__in=["sent", "viewed", "partial", "overdue"]
    ).aggregate(total=Sum("total_amount"))["total"] or 0
    overdue_invoices = Invoice.objects.filter(status="overdue").count()
    
    clients = Beneficiary.objects.exclude(credit_limit=0)
    
    overdue_accounts = Beneficiary.objects.filter(
        credit_limit__gt=0,
        total_outstanding__gt=F('credit_limit')
    )

    total_clients = Beneficiary.objects.count()
    private_beneficiaries_count = Beneficiary.objects.filter(beneficiary_type__in=['private', 'town']).count()
    communal_beneficiaries_count = Beneficiary.objects.filter(beneficiary_type__iexact='communal').count()
    active_clients = Beneficiary.objects.filter(is_active=True).count()
    inactive_clients = Beneficiary.objects.filter(is_active=False).count()
    total_households = Beneficiary.objects.aggregate(total=Sum("household_count"))["total"] or 0
    total_private_households = Beneficiary.objects.filter(beneficiary_type__in=['private', 'town']).aggregate(total=Sum("household_count"))["total"] or 0
    total_communal_households = Beneficiary.objects.filter(beneficiary_type__iexact='communal').aggregate(total=Sum("household_count"))["total"] or 0
    
    scheme_data = []
    for scheme_code, scheme_name in Beneficiary.SCHEME_CHOICES:
        priv_clients = Beneficiary.objects.filter(scheme=scheme_code, beneficiary_type__in=['private', 'town'])
        comm_clients = Beneficiary.objects.filter(scheme=scheme_code, beneficiary_type__iexact='communal')
        
        scheme_data.append({
            "code": scheme_code,
            "name": scheme_name,
            "total_clients": priv_clients.count() + comm_clients.count(),
            "private_clients": priv_clients.count(),
            "communal_clients": comm_clients.count(),
            "private_households": priv_clients.aggregate(total=Sum("household_count"))["total"] or 0,
            "communal_households": comm_clients.aggregate(total=Sum("household_count"))["total"] or 0,
            "total_households": (priv_clients.aggregate(total=Sum("household_count"))["total"] or 0) + (comm_clients.aggregate(total=Sum("household_count"))["total"] or 0),
        })

    scheme_total_invoiced = Decimal('0.00')
    scheme_total_collected = Decimal('0.00')
    scheme_total_balance = Decimal('0.00')
    
    for scheme_code, scheme_name in Beneficiary.SCHEME_CHOICES:
        beneficiaries = Beneficiary.objects.filter(scheme=scheme_code, is_active=True)
        beneficiary_ids = beneficiaries.values_list('id', flat=True)
        
        scheme_invoiced = Invoice.objects.filter(
            beneficiary_id__in=beneficiary_ids,
            issue_date__gte=year_start,
            issue_date__lte=today
        ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        
        total_collected = Payment.objects.filter(
            beneficiary_id__in=beneficiary_ids,
            payment_date__gte=year_start,
            payment_date__lte=today
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        scheme_total_invoiced += scheme_invoiced
        scheme_total_collected += total_collected
        scheme_total_balance += (scheme_invoiced - total_collected)

    total_invoices = Invoice.objects.count()
    total_expenses_count = Expense.objects.count()
    total_vendors = Vendor.objects.count()
    total_payments = Payment.objects.count()
    total_accounts = Account.objects.count()

    recent_invoices = Invoice.objects.all()[:5]
    recent_expenses = Expense.objects.all()[:5]
    pending_payments = Payment.objects.filter(payment_date=today)[:5]
    
    current_year = today.year
    beneficiaries = Beneficiary.objects.filter(is_active=True)
    
    total_opening_balance = Decimal('0.00')
    total_current_outstanding = Decimal('0.00')
    clients_with_opening = 0
    
    for beneficiary in beneficiaries:
        opening = beneficiary.opening_balances.filter(fiscal_year=current_year).first()
        if opening:
            total_opening_balance += opening.amount
            clients_with_opening += 1
        total_current_outstanding += beneficiary.total_outstanding
    
    total_balance_with_opening = total_opening_balance + total_current_outstanding
    
    last_rollover = YearEndRollover.objects.first()

    monthly_revenue = []
    current_month = today.month
    current_year = today.year
    
    for i in range(12):
        month = current_month - i
        year = current_year
        while month <= 0:
            month += 12
            year -= 1
        
        if month == 12:
            month_start = date(year, 12, 1)
            month_end = date(year + 1, 1, 1)
        else:
            month_start = date(year, month, 1)
            if month == 12:
                month_end = date(year + 1, 1, 1)
            else:
                month_end = date(year, month + 1, 1)
        
        revenue = Invoice.objects.filter(
            status="paid",
            issue_date__gte=month_start,
            issue_date__lt=month_end
        ).aggregate(total=Sum("total_amount"))["total"] or 0
        monthly_revenue.append(float(revenue))

    context = {
        "total_revenue": total_revenue_all,
        "total_expenses": total_expenses,
        "total_invoiced": total_invoiced_all,
        "net_profit": total_revenue - total_expenses,
        "outstanding": total_invoiced_all - total_revenue_all,
        "outstanding_invoices": outstanding_invoices,
        "overdue_invoices": overdue_invoices,
        "total_beneficiaries": total_clients,
        "private_beneficiaries": private_beneficiaries_count,
        "communal_beneficiaries": communal_beneficiaries_count,
        "total_private_households": total_private_households,
        "total_communal_households": total_communal_households,
        "active_clients": active_clients,
        "inactive_clients": inactive_clients,
        "total_households": total_households,
        "scheme_total_invoiced": scheme_total_invoiced,
        "scheme_total_collected": scheme_total_collected,
        "scheme_total_balance": scheme_total_balance,
        "total_invoices": total_invoices,
        "total_expenses_count": total_expenses_count,
        "total_vendors": total_vendors,
        "total_payments": total_payments,
        "total_accounts": total_accounts,
        "recent_invoices": recent_invoices,
        "recent_expenses": recent_expenses,
        "pending_payments": pending_payments,
        "monthly_revenue": monthly_revenue,
        "overdue_accounts": overdue_accounts,
        "scheme_data": scheme_data,
        "current_fiscal_year": current_year,
        "total_opening_balance": total_opening_balance,
        "total_current_outstanding": total_current_outstanding,
        "total_balance_with_opening": total_balance_with_opening,
        "clients_with_opening": clients_with_opening,
        "last_rollover": last_rollover,
        "pending_invoices": Invoice.objects.filter(status__in=['sent', 'viewed', 'partial']).count(),
        "total_collections": total_revenue_all,
        "total_overdue": outstanding_invoices,
        "available_balance": total_revenue_all - total_expenses,
    }
    return render(request, "accounting_app/dashboard.html", context)


@login_required
def account_list(request):
    accounts = Account.objects.all()
    return render(request, "accounting_app/account_list.html", {"accounts": accounts})


@login_required
def account_create(request):
    if request.method == "POST":
        form = AccountForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Account created successfully")
            return redirect("account_list")
    else:
        form = AccountForm()
    return render(request, "accounting_app/account_form.html", {"form": form, "action": "Create"})


@login_required
def account_edit(request, pk):
    account = get_object_or_404(Account, pk=pk)
    if request.method == "POST":
        form = AccountForm(request.POST, instance=account)
        if form.is_valid():
            form.save()
            messages.success(request, "Account updated successfully")
            return redirect("account_list")
    else:
        form = AccountForm(instance=account)
    return render(request, "accounting_app/account_form.html", {"form": form, "action": "Update"})


@login_required
def account_delete(request, pk):
    account = get_object_or_404(Account, pk=pk)
    if request.method == "POST":
        if account.children.exists():
            messages.error(request, f"Cannot delete '{account.name}' because it has sub-accounts. Please delete or reassign the sub-accounts first.")
            return redirect("account_list")
        
        if account.expenses.exists():
            messages.error(request, f"Cannot delete '{account.name}' because it has linked expenses. Please reassign or remove the linked expenses first.")
            return redirect("account_list")
        
        save_deleted_record(account, request.user)
        account.delete()
        messages.success(request, "Account deleted successfully")
        return redirect("account_list")
    
    return render(request, "accounting_app/account_delete.html", {"account": account})


@login_required
def overdue_accounts(request):
    overdue_accounts = Beneficiary.objects.filter(
        credit_limit__gt=0,
        total_outstanding__gt=F('credit_limit')
    )
    return render(request, "accounting_app/overdue_accounts.html", {
        "overdue_accounts": overdue_accounts,
    })


@login_required
def beneficiary_list(request):
    beneficiaries = Beneficiary.objects.all().order_by('name')
    
    beneficiary_type = request.GET.get("beneficiary_type")
    scheme = request.GET.get("scheme")
    is_active = request.GET.get("is_active")
    search = request.GET.get("search")
    
    if beneficiary_type:
        beneficiaries = beneficiaries.filter(beneficiary_type=beneficiary_type)
    if scheme:
        beneficiaries = beneficiaries.filter(scheme=scheme)
    if is_active:
        beneficiaries = beneficiaries.filter(is_active=(is_active == "true"))
    if search:
        beneficiaries = beneficiaries.filter(
            Q(name__icontains=search) |
            Q(phone__icontains=search) |
            Q(village__icontains=search)
        )
    
    return render(request, "accounting_app/beneficiary_list.html", {
        "beneficiaries": beneficiaries,
        "beneficiary_type_filter": beneficiary_type or "",
        "scheme_filter": scheme or "",
        "is_active_filter": is_active or "",
        "search_filter": search or "",
        "symbol": request.user.userprofile.get_currency_symbol if hasattr(request.user, 'userprofile') else "K",
    })


@login_required
def beneficiary_autocomplete(request):
    """AJAX endpoint for beneficiary autocomplete search."""
    from django.http import JsonResponse

    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=403)

    query = request.GET.get('q', '').strip()

    if len(query) < 2:
        return JsonResponse({'results': []})

    beneficiaries = Beneficiary.objects.filter(
        Q(name__icontains=query) |
        Q(phone__icontains=query) |
        Q(id__icontains=query) |
        Q(village__icontains=query) |
        Q(scheme__icontains=query)
    ).values('id', 'name', 'phone', 'village', 'scheme', 'household_count', 'total_outstanding')[:20]

    results = []
    for b in beneficiaries:
        results.append({
            'id': b['id'],
            'name': b['name'],
            'phone': b['phone'] or '',
            'village': b['village'] or '',
            'scheme': b['scheme'] or '',
            'household_count': b['household_count'] or 0,
            'total_outstanding': float(b['total_outstanding'] or 0),
            'label': f"{b['name']} - {b['village'] or 'No Village'} ({b['phone'] or 'No Phone'})"
        })

    return JsonResponse({'results': results})


def search_suggestions(request):
    from django.http import JsonResponse
    
    query = request.GET.get('q', '').strip()
    search_type = request.GET.get('type', 'beneficiary')
    
    if len(query) < 2:
        return JsonResponse({'suggestions': []})
    
    suggestions = []
    
    if search_type == 'beneficiary':
        results = Beneficiary.objects.filter(
            Q(name__icontains=query) |
            Q(phone__icontains=query) |
            Q(village__icontains=query)
        ).values('id', 'name')[:10]
        suggestions = [{'id': r['id'], 'label': r['name']} for r in results]
    
    elif search_type == 'user':
        results = User.objects.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        ).values('id', 'username', 'first_name', 'last_name', 'email')[:10]
        suggestions = [{'id': r['id'], 'label': f"{r['first_name']} {r['last_name']}".strip() or r['username']} for r in results]
    
    return JsonResponse({'suggestions': suggestions})


@login_required
def beneficiary_create(request):
    if request.method == "POST":
        form = BeneficiaryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Beneficiary created successfully")
            return redirect("beneficiary_list")
    else:
        form = BeneficiaryForm()
    return render(request, "accounting_app/beneficiary_form.html", {"form": form, "action": "Create"})


@login_required
def beneficiary_edit(request, pk):
    beneficiary = get_object_or_404(Beneficiary, pk=pk)
    # Get old status from database to ensure accuracy
    old_status = Beneficiary.objects.get(pk=pk).is_active
    
    if request.method == "POST":
        form = BeneficiaryForm(request.POST, request.FILES, instance=beneficiary)
        if form.is_valid():
            # Determine new status - checkbox unchecked means not in POST
            new_status = 'is_active' in request.POST
            
            # Save the beneficiary
            beneficiary = form.save(commit=False)
            beneficiary.is_active = new_status
            beneficiary.created_by = request.user
            beneficiary.save()
            
            # Always create a log entry for status changes
            if old_status != new_status:
                BeneficiaryStatusLog.objects.create(
                    beneficiary=beneficiary,
                    user=request.user if request.user.is_authenticated else None,
                    status='activated' if new_status else 'deactivated'
                )
                messages.success(request, f"Beneficiary {beneficiary.name} {'activated' if new_status else 'deactivated'} successfully!")
            else:
                messages.success(request, "Beneficiary updated successfully!")
            
            return redirect("beneficiary_list")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = BeneficiaryForm(instance=beneficiary)
    return render(request, "accounting_app/beneficiary_form.html", {"form": form, "action": "Update"})


def beneficiary_delete(request, pk):
    beneficiary = get_object_or_404(Beneficiary, pk=pk)
    if request.method == "POST":
        save_deleted_record(beneficiary, request.user)
        beneficiary.delete()
        messages.success(request, "Beneficiary deleted successfully")
        return redirect("beneficiary_list")
    return render(request, "accounting_app/beneficiary_delete.html", {"client": client})


@login_required
def beneficiary_toggle_status(request, pk):
    beneficiary = get_object_or_404(Beneficiary, pk=pk)
    # Toggle the status
    beneficiary.is_active = not beneficiary.is_active
    beneficiary.created_by = request.user
    beneficiary.save()
    # Log the status change with user info
    BeneficiaryStatusLog.objects.create(
        beneficiary=beneficiary,
        user=request.user if request.user.is_authenticated else None,
        status='activated' if beneficiary.is_active else 'deactivated'
    )
    status = "activated" if beneficiary.is_active else "deactivated"
    messages.success(request, f"Client {beneficiary.name} has been {status}")
    return redirect("beneficiary_list")


@login_required
def bulk_beneficiary_create(request):
    if request.method == "POST":
        beneficiary_names = request.POST.get('beneficiary_names', '').strip()
        beneficiary_type = request.POST.get('beneficiary_type', '').strip()
        scheme = request.POST.get('scheme', '').strip()
        village = request.POST.get('village', '').strip()
        country = request.POST.get('country', '').strip()
        household_count = int(request.POST.get('household_count', 0) or 0)
        credit_limit = float(request.POST.get('credit_limit', 0) or 0)
        
        errors = []
        
        if not beneficiary_names:
            errors.append("Please enter at least one beneficiary name")
        if not beneficiary_type:
            errors.append("Beneficiary Type is required")
        if not village:
            errors.append("Village is required")
        if not scheme:
            errors.append("Scheme is required")
        
        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, "accounting_app/bulk_beneficiary_create.html")
        
        names_list = [name.strip() for name in beneficiary_names.split('\n') if name.strip()]
        
        created_count = 0
        error_list = []
        
        for name in names_list:
            try:
                existing = Beneficiary.objects.filter(name=name).first()
                if existing:
                    existing.beneficiary_type = beneficiary_type
                    existing.scheme = scheme
                    existing.village = village
                    if country:
                        existing.country = country
                    existing.household_count = household_count
                    existing.credit_limit = credit_limit
                    existing.created_by = request.user
                    existing.save()
                else:
                    b = Beneficiary(
                        name=name,
                        beneficiary_type=beneficiary_type,
                        scheme=scheme,
                        village=village,
                        country=country,
                        household_count=household_count,
                        credit_limit=credit_limit,
                        is_active=True,
                    )
                    b.created_by = request.user
                    b.save()
                created_count += 1
            except Exception as e:
                error_list.append(f"{name}: {str(e)}")
        
        if created_count > 0:
            messages.success(request, f"Successfully created/updated {created_count} clients")
        if error_list:
            messages.error(request, f"Errors: {'; '.join(error_list[:5])}")
            if len(error_list) > 5:
                messages.error(request, f"... and {len(error_list) - 5} more errors")
        
        return redirect("beneficiary_list")
    
    return render(request, "accounting_app/bulk_beneficiary_create.html")


@login_required
def bulk_beneficiary_import(request):
    if request.method == "POST" and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(excel_file)
            ws = wb.active
            
            headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
            
            required_fields = ['name']
            if not all(field in headers for field in required_fields):
                messages.error(request, "The template must have a 'name' column")
                return redirect("bulk_beneficiary_import")
            
            created_count = 0
            updated_count = 0
            error_count = 0
            errors = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(cell for cell in row):
                    continue
                
                row_data = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        row_data[header] = row[i]
                
                try:
                    name = str(row_data.get('name', '')).strip()
                    if not name:
                        errors.append(f"Row {row_idx}: Name is required")
                        error_count += 1
                        continue
                    
                    beneficiary_type = str(row_data.get('beneficiary_type', 'private')).strip() or 'private'
                    phone = str(row_data.get('phone', '')).strip()
                    village = str(row_data.get('village', '')).strip()
                    scheme = str(row_data.get('scheme', '')).strip()
                    country = str(row_data.get('country', '')).strip()
                    tax_id = str(row_data.get('tax_id', '')).strip()
                    
                    h_count = row_data.get('household_count', 0) or 0
                    try:
                        household_count = int(float(h_count))
                    except (ValueError, TypeError):
                        household_count = 0
                    
                    c_limit = row_data.get('credit_limit', 0) or 0
                    try:
                        credit_limit = float(c_limit)
                    except (ValueError, TypeError):
                        credit_limit = 0
                    
                    p_terms = row_data.get('payment_terms', 30) or 30
                    try:
                        payment_terms = int(float(p_terms))
                    except (ValueError, TypeError):
                        payment_terms = 30
                    
                    is_active_str = row_data.get('is_active', True)
                    if isinstance(is_active_str, str):
                        is_active = is_active_str.lower().strip() in ['true', 'yes', '1', 'active']
                    else:
                        is_active = bool(is_active_str)
                    
                    existing_client = Beneficiary.objects.filter(name=name).first()
                    
                    if existing_client:
                        existing_client.beneficiary_type = beneficiary_type
                        existing_client.phone = phone
                        existing_client.village = village
                        existing_client.scheme = scheme
                        existing_client.country = country
                        existing_client.tax_id = tax_id
                        existing_client.household_count = household_count
                        existing_client.credit_limit = credit_limit
                        existing_client.payment_terms = payment_terms
                        existing_client.is_active = is_active
                        existing_client.created_by = request.user
                        existing_client.save()
                        updated_count += 1
                    else:
                        b = Beneficiary(
                            name=name,
                            beneficiary_type=beneficiary_type,
                            phone=phone,
                            village=village,
                            scheme=scheme,
                            country=country,
                            tax_id=tax_id,
                            household_count=household_count,
                            credit_limit=credit_limit,
                            payment_terms=payment_terms,
                            is_active=is_active,
                        )
                        b.created_by = request.user
                        b.save()
                        created_count += 1
                        
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
                    error_count += 1
            
            wb.close()
            
            messages.success(request, f"Import completed: {created_count} created, {updated_count} updated, {error_count} errors")
            if errors:
                for error in errors[:10]:
                    messages.warning(request, error)
                if len(errors) > 10:
                    messages.warning(request, f"... and {len(errors) - 10} more errors")
                    
        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
        
        return redirect("beneficiary_list")
    
    return render(request, "accounting_app/bulk_beneficiary_import.html")


@login_required
def download_beneficiary_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"
    
    headers = ['name', 'beneficiary_type', 'phone', 'village', 'scheme', 'country', 'tax_id', 'household_count', 'credit_limit', 'payment_terms', 'is_active']
    ws.append(headers)
    
    ws.append(['Acme Corp', 'private', '+265991234567', 'Lilongwe', 'Mangale', 'Malawi', 'TAX123', '50', '100000', '30', 'True'])
    ws.append(['Village Water Committee', 'communal', '+265991234568', 'Mzuzu', 'Nkala', 'Malawi', '', '200', '0', '30', 'True'])
    ws.append(['John Doe', 'private', '+265991234569', 'Blantyre', 'Dodza', 'Malawi', '', '25', '50000', '30', 'True'])
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="client_import_template.xlsx"'
    wb.save(response)
    return response


@login_required
def bulk_beneficiary_delete(request):
    if request.method == "POST":
        beneficiary_ids = request.POST.get('beneficiary_ids', '')
        
        if not beneficiary_ids:
            messages.error(request, "No beneficiaries selected")
            return redirect("beneficiary_list")
        
        try:
            ids = [int(x) for x in beneficiary_ids.split(',') if x.strip()]
            beneficiaries = Beneficiary.objects.filter(pk__in=ids)
            count = beneficiaries.count()
            beneficiaries.delete()
            messages.success(request, f"Successfully deleted {count} beneficiary(ies)")
        except Exception as e:
            messages.error(request, f"Error deleting beneficiaries: {str(e)}")
    
    return redirect("beneficiary_list")


@login_required
def bulk_beneficiary_edit(request):
    today = timezone.now().date()
    beneficiaries = Beneficiary.objects.filter(is_active=True).order_by('scheme', 'name')
    schemes = Beneficiary.SCHEME_CHOICES
    types = Beneficiary.BENEFICIARY_TYPE_CHOICES

    base_ctx = {
        "beneficiaries": beneficiaries,
        "schemes": schemes,
        "types": types,
        "today": today,
        "symbol": '',
    }
    if hasattr(request.user, 'userprofile'):
        base_ctx['symbol'] = request.user.userprofile.get_currency_symbol()

    if request.method == "POST":
        selected_ids = request.POST.getlist('selected_beneficiaries')

        if not selected_ids:
            messages.error(request, "Please select at least one beneficiary")
            return render(request, "accounting_app/bulk_beneficiary_edit.html", base_ctx)

        scheme = request.POST.get('scheme', '')
        beneficiary_type = request.POST.get('beneficiary_type', '')
        village = request.POST.get('village', '')
        household_count = request.POST.get('household_count')
        credit_limit = request.POST.get('credit_limit')
        payment_terms = request.POST.get('payment_terms')
        is_active = request.POST.get('is_active')

        updated = 0
        errors = []

        try:
            with transaction.atomic():
                for bid in selected_ids:
                    try:
                        b = Beneficiary.objects.get(pk=bid)
                        changed = False

                        if scheme and scheme != b.scheme:
                            b.scheme = scheme
                            changed = True
                        if beneficiary_type and beneficiary_type != b.beneficiary_type:
                            b.beneficiary_type = beneficiary_type
                            changed = True
                        if village and village != b.village:
                            b.village = village
                            changed = True
                        if household_count and int(household_count) != b.household_count:
                            b.household_count = int(household_count)
                            changed = True
                        if credit_limit and float(credit_limit) != b.credit_limit:
                            b.credit_limit = float(credit_limit)
                            changed = True
                        if payment_terms and int(payment_terms) != b.payment_terms:
                            b.payment_terms = int(payment_terms)
                            changed = True
                        if is_active:
                            new_active = is_active == 'true'
                            if new_active != b.is_active:
                                b.is_active = new_active
                                changed = True

                        if changed:
                            b.save()
                            updated += 1
                    except Beneficiary.DoesNotExist:
                        errors.append(f"Beneficiary #{bid} not found")
                    except Exception as e:
                        b_name = getattr(b, 'name', f"Beneficiary #{bid}")
                        errors.append(f"{b_name}: {str(e)}")
        except Exception as e:
            errors.append(f"Database error: {str(e)}")

        if updated:
            messages.success(request, f"Successfully updated {updated} beneficiary(ies)")
        if errors:
            messages.warning(request, f"Failed to update {len(errors)} beneficiary(ies): {'; '.join(errors)}")

        return redirect("beneficiary_list")

    return render(request, "accounting_app/bulk_beneficiary_edit.html", base_ctx)


@login_required
def beneficiary_pdf_report(request, pk):
    beneficiary = get_object_or_404(Beneficiary, pk=pk)
    invoices = beneficiary.invoices.all().order_by('-issue_date')
    payments = beneficiary.payments.all().order_by('-payment_date')
    user_profile = request.user.userprofile
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Client_Report_{beneficiary.name}_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    elements = []
    
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Center', alignment=TA_CENTER))
    styles.add(ParagraphStyle(name='Right', alignment=TA_RIGHT))
    styles.add(ParagraphStyle(name='Title_Custom', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=18, spaceAfter=20))
    styles.add(ParagraphStyle(name='SubTitle', parent=styles['Normal'], alignment=TA_CENTER, fontSize=10, textColor=colors.grey))
    
    company_name = user_profile.company_name or "Nkulawua"
    currency_symbol = user_profile.get_currency_symbol()
    
    elements.append(Spacer(1, 10))
    
    if user_profile.logo:
        try:
            logo = Image(user_profile.logo.path, width=2*cm, height=2*cm)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 10))
        except:
            pass
    
    elements.append(Paragraph(company_name, styles['Title_Custom']))
    elements.append(Paragraph(f"Client Report: {beneficiary.name}", styles['Heading2']))
    elements.append(Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}", styles['SubTitle']))
    elements.append(Spacer(1, 20))
    
    client_info = [
        ["BENEFICIARY INFORMATION", ""],
        ["Name:", beneficiary.name],
        ["Type:", beneficiary.get_beneficiary_type_display()],
        ["Phone:", beneficiary.phone or "-"],
        ["Village:", beneficiary.village or "-"],
        ["Scheme:", beneficiary.scheme or "-"],
        ["Country:", beneficiary.country or "-"],
        ["Tax ID:", beneficiary.tax_id or "-"],
        ["Households:", str(beneficiary.household_count)],
        ["Credit Limit:", f"{currency_symbol}{beneficiary.credit_limit:,.2f}"],
        ["Payment Terms:", f"{beneficiary.payment_terms} days"],
        ["Status:", "Active" if beneficiary.is_active else "Inactive"],
        ["Created:", beneficiary.created_at.strftime('%Y-%m-%d')],
    ]
    
    info_table = Table(client_info, colWidths=[5*cm, 10*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('SPAN', (0, 0), (-1, 0)),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    fin_info = [
        ["FINANCIAL SUMMARY", "", "", ""],
        ["Total Bill:", f"{currency_symbol}{beneficiary.total_bill:,.2f}", "Total Paid:", f"{currency_symbol}{beneficiary.total_paid:,.2f}"],
        ["Outstanding Balance:", f"{currency_symbol}{beneficiary.total_outstanding:,.2f}", "", ""],
    ]
    fin_table = Table(fin_info, colWidths=[4*cm, 4*cm, 4*cm, 4*cm])
    fin_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('SPAN', (0, 0), (-1, 0)),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#e74c3c')),
        ('TEXTCOLOR', (0, 3), (-1, 3), colors.whitesmoke),
        ('SPAN', (1, 3), (-1, 3)),
        ('ALIGN', (0, 3), (-1, 3), 'CENTER'),
    ]))
    elements.append(fin_table)
    elements.append(Spacer(1, 25))
    
    if invoices:
        elements.append(Paragraph("INVOICE HISTORY", styles['Heading3']))
        elements.append(Spacer(1, 10))
        inv_data = [["Invoice #", "Date", "Due Date", "Status", "Amount"]]
        for inv in invoices:
            inv_data.append([
                inv.invoice_number,
                inv.issue_date.strftime('%Y-%m-%d'),
                inv.due_date.strftime('%Y-%m-%d'),
                inv.get_status_display(),
                f"{currency_symbol}{inv.total_amount:,.2f}"
            ])
        inv_table = Table(inv_data, colWidths=[3*cm, 3*cm, 3*cm, 3*cm, 4*cm])
        inv_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('ALIGN', (3, 0), (3, -1), 'CENTER'),
            ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
        ]))
        elements.append(inv_table)
        elements.append(Spacer(1, 20))
    
    if payments:
        elements.append(Paragraph("PAYMENT HISTORY", styles['Heading3']))
        elements.append(Spacer(1, 10))
        pay_data = [["Date", "Invoice", "Amount", "Method", "Reference"]]
        for pay in payments:
            inv_num = pay.invoice.invoice_number if pay.invoice else "N/A"
            pay_data.append([
                pay.payment_date.strftime('%Y-%m-%d'),
                inv_num,
                f"{currency_symbol}{pay.amount:,.2f}",
                pay.get_payment_method_display(),
                pay.reference or "-"
            ])
        pay_table = Table(pay_data, colWidths=[3*cm, 3*cm, 3*cm, 3*cm, 4*cm])
        pay_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9b59b6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ]))
        elements.append(pay_table)
    
    doc.build(elements)
    return response


@login_required
def vendor_list(request):
    vendors = Vendor.objects.all()
    return render(request, "accounting_app/vendor_list.html", {"vendors": vendors})


@login_required
def vendor_create(request):
    if request.method == "POST":
        form = VendorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Vendor created successfully")
            return redirect("vendor_list")
    else:
        form = VendorForm()
    return render(request, "accounting_app/vendor_form.html", {"form": form, "action": "Create"})


@login_required
def vendor_edit(request, pk):
    vendor = get_object_or_404(Vendor, pk=pk)
    if request.method == "POST":
        form = VendorForm(request.POST, instance=vendor)
        if form.is_valid():
            form.save()
            messages.success(request, "Vendor updated successfully")
            return redirect("vendor_list")
    else:
        form = VendorForm(instance=vendor)
    return render(request, "accounting_app/vendor_form.html", {"form": form, "action": "Update"})


@login_required
def invoice_list(request):
    try:
        from django.db.models import Max, Min
        invoices = Invoice.objects.all().order_by('-created_at')
        status_filter = request.GET.get("status")
        if status_filter:
            invoices = invoices.filter(status=status_filter)
        
        # Annotate bulk group info
        invoice_list = []
        processed_groups = set()
        has_bulk_invoices = False
        first_bulk_group = None
        
        for inv in invoices:
            try:
                is_bulk = inv.is_bulk if hasattr(inv, 'is_bulk') else False
                bulk_group_id = inv.bulk_group_id if hasattr(inv, 'bulk_group_id') else None
            except Exception:
                is_bulk = False
                bulk_group_id = None
            
            # Calculate balance_before: sum of amount_due from all other unpaid invoices for same beneficiary
            try:
                other_invoices = Invoice.objects.filter(
                    beneficiary=inv.beneficiary,
                    status__in=['sent', 'viewed', 'partial', 'overdue']
                ).exclude(pk=inv.pk)
                inv.balance_before = sum(inv.amount_due for inv in other_invoices)
                inv.total_balance = inv.total_amount + inv.balance_before
            except:
                inv.balance_before = 0
                inv.total_balance = inv.total_amount
            
            if is_bulk and bulk_group_id:
                if bulk_group_id not in processed_groups:
                    inv.bulk_first = True
                    try:
                        inv.bulk_count = Invoice.objects.filter(bulk_group_id=bulk_group_id).count()
                    except:
                        inv.bulk_count = 0
                    processed_groups.add(bulk_group_id)
                    if not has_bulk_invoices:
                        has_bulk_invoices = True
                        first_bulk_group = bulk_group_id
                else:
                    inv.bulk_first = False
            else:
                inv.bulk_first = True
            
            invoice_list.append(inv)
        
        # Calculate summary statistics
        total_amount = sum(inv.total_amount for inv in invoice_list)
        
        return render(request, "accounting_app/invoice_list.html", {
            "invoices": invoice_list, 
            "status_filter": status_filter,
            "has_bulk_invoices": has_bulk_invoices,
            "first_bulk_group": first_bulk_group,
            "total_amount": total_amount,
        })
    except Exception as e:
        # Fallback - just show basic invoice list
        invoices = Invoice.objects.all().order_by('-created_at')
        return render(request, "accounting_app/invoice_list.html", {
            "invoices": invoices, 
            "status_filter": request.GET.get("status"),
            "has_bulk_invoices": False,
            "first_bulk_group": None,
            "total_amount": 0,
        })


@login_required
def invoice_create(request):
    today = timezone.now().date()
    invoice_number = generate_invoice_number()
    clients = Beneficiary.objects.filter(is_active=True).order_by('name')
    
    if request.method == "POST":
        try:
            beneficiary_id = request.POST.get('client')
            household_count = int(request.POST.get('household_count', 0))
            cost_per_unit = float(request.POST.get('cost_per_unit', 0))
            issue_date = request.POST.get('issue_date')
            due_date = request.POST.get('due_date')
            tax_rate = float(request.POST.get('tax_rate', 0))
            discount = float(request.POST.get('discount', 0))
            notes = request.POST.get('notes', '')
            terms = request.POST.get('terms', '')
            
            if not beneficiary_id:
                messages.error(request, "Please select a beneficiary")
                return render(request, "accounting_app/invoice_form.html", {
                    "action": "Create", 
                    "clients": clients, 
                    "invoice_number": invoice_number,
                    "today": today,
                    "status_choices": Invoice.STATUS_CHOICES,
                    "user": request.user
                })
            
            beneficiary = get_object_or_404(Beneficiary, pk=beneficiary_id)
            
            subtotal = household_count * cost_per_unit
            tax_amount = subtotal * (tax_rate / 100)
            total_amount = subtotal + tax_amount - discount
            
            invoice = Invoice.objects.create(
                invoice_number=invoice_number,
                beneficiary=beneficiary,
                issue_date=issue_date,
                due_date=due_date,
                household_count=household_count,
                cost_per_unit=cost_per_unit,
                tax_rate=tax_rate,
                tax_amount=tax_amount,
                discount=discount,
                total_amount=total_amount,
                notes=notes,
                terms=terms,
                created_by=request.user
            )

            ActivityLog.objects.create(
                user=request.user,
                action="Sales",
                model_name="Invoice",
                object_id=invoice.pk,
                description=f"Created invoice {invoice_number} for {beneficiary.name} - {total_amount}",
                ip_address=request.META.get('REMOTE_ADDR'),
            )
            
            beneficiary.recalculate_totals()
            
            messages.success(request, f"Invoice {invoice_number} created successfully")
            return redirect("invoice_list")
        except Exception as e:
            messages.error(request, f"Error creating invoice: {str(e)}")
    
    return render(request, "accounting_app/invoice_form.html", {
        "action": "Create", 
        "clients": clients, 
        "invoice_number": invoice_number,
        "today": today,
        "status_choices": Invoice.STATUS_CHOICES,
        "user": request.user
    })


@login_required
def bulk_invoice_create(request):
    today = timezone.now().date()
    beneficiaries = Beneficiary.objects.filter(is_active=True).order_by('scheme', 'name')
    
    schemes = Beneficiary.SCHEME_CHOICES
    
    base_ctx = {
        "beneficiaries": beneficiaries,
        "schemes": schemes,
        "today": today,
        "symbol": '',
    }
    if hasattr(request.user, 'userprofile'):
        base_ctx['symbol'] = request.user.userprofile.get_currency_symbol()
    
    if request.method == "POST":
        amount = float(request.POST.get('amount', 0))
        raw_issue_date = request.POST.get('issue_date')
        raw_due_date = request.POST.get('due_date')
        issue_date = date.fromisoformat(raw_issue_date) if raw_issue_date else None
        due_date = date.fromisoformat(raw_due_date) if raw_due_date else None
        notes = request.POST.get('notes', '')
        terms = request.POST.get('terms', '')
        
        selected_ids = request.POST.getlist('selected_beneficiaries')
        
        post_ctx = {
            **base_ctx,
            "amount": amount,
            "notes": notes,
            "terms": terms,
        }
        
        if not selected_ids:
            messages.error(request, "Please select at least one beneficiary")
            return render(request, "accounting_app/bulk_invoice_form.html", post_ctx)
        
        if amount <= 0:
            messages.error(request, "Please enter a valid amount greater than 0")
            return render(request, "accounting_app/bulk_invoice_form.html", post_ctx)
        
        if not issue_date or not due_date:
            messages.error(request, "Please select issue date and due date")
            return render(request, "accounting_app/bulk_invoice_form.html", post_ctx)
        
        bulk_group_id = f"BULK-{uuid.uuid4().hex[:12]}"
        
        created_invoices = []
        errors = []
        
        try:
            with transaction.atomic():
                for beneficiary_id in selected_ids:
                    try:
                        client = Beneficiary.objects.get(pk=beneficiary_id)
                        invoice_number = generate_invoice_number()
                        
                        invoice = Invoice.objects.create(
                            invoice_number=invoice_number,
                            beneficiary=client,
                            issue_date=issue_date,
                            due_date=due_date,
                            household_count=client.household_count or 0,
                            cost_per_unit=amount,
                            tax_rate=0,
                            tax_amount=0,
                            discount=0,
                            total_amount=amount,
                            notes=notes,
                            terms=terms,
                            created_by=request.user,
                            is_bulk=True,
                            bulk_group_id=bulk_group_id
                        )
                        created_invoices.append(invoice)
                    except Beneficiary.DoesNotExist:
                        errors.append(f"Beneficiary #{beneficiary_id} not found")
                    except Exception as e:
                        client_name = getattr(client, 'name', f"Beneficiary #{beneficiary_id}")
                        errors.append(f"{client_name}: {str(e)}")
                
                for inv in created_invoices:
                    inv.beneficiary.recalculate_totals()
        except Exception as e:
            errors.append(f"Database error: {str(e)}")
        
        if created_invoices:
            messages.success(request, f"Successfully created {len(created_invoices)} invoices")
        if errors:
            messages.warning(request, f"Failed to create {len(errors)} invoices: {'; '.join(errors)}")
        
        return redirect("invoice_list")
    
    return render(request, "accounting_app/bulk_invoice_form.html", base_ctx)


@login_required
def bulk_invoice_edit(request, bulk_group_id):
    invoices = Invoice.objects.filter(bulk_group_id=bulk_group_id, is_bulk=True)
    if not invoices.exists():
        messages.error(request, "Bulk invoices not found")
        return redirect("invoice_list")
    
    first_invoice = invoices.first()
    today = timezone.now().date()
    clients = Beneficiary.objects.filter(is_active=True).order_by('scheme', 'name')
    schemes = Beneficiary.SCHEME_CHOICES
    
    if request.method == "POST":
        household_count = int(request.POST.get('household_count', 0))
        cost_per_unit = float(request.POST.get('cost_per_unit', 0))
        raw_issue_date = request.POST.get('issue_date')
        raw_due_date = request.POST.get('due_date')
        issue_date = date.fromisoformat(raw_issue_date) if raw_issue_date else None
        due_date = date.fromisoformat(raw_due_date) if raw_due_date else None
        tax_rate = float(request.POST.get('tax_rate', 0))
        discount = float(request.POST.get('discount', 0))
        notes = request.POST.get('notes', '')
        terms = request.POST.get('terms', '')
        
        if not issue_date or not due_date:
            messages.error(request, "Please select issue date and due date")
            return render(request, "accounting_app/bulk_invoice_edit.html", {
                "invoices": invoices,
                "first_invoice": first_invoice,
                "bulk_group_id": bulk_group_id,
                "schemes": schemes,
                "today": today,
            })
        
        subtotal = household_count * cost_per_unit
        tax_amount = subtotal * (tax_rate / 100)
        total_amount = subtotal + tax_amount - discount
        
        # Update all invoices in the bulk group
        for invoice in invoices:
            invoice.household_count = household_count
            invoice.cost_per_unit = cost_per_unit
            invoice.issue_date = issue_date
            invoice.due_date = due_date
            invoice.tax_rate = tax_rate
            invoice.discount = discount
            invoice.notes = notes
            invoice.terms = terms
            invoice.tax_amount = tax_amount
            invoice.total_amount = total_amount
            invoice.save()
            invoice.beneficiary.recalculate_totals()
        
        messages.success(request, f"Successfully updated {invoices.count()} bulk invoices")
        return redirect("invoice_list")
    
    return render(request, "accounting_app/bulk_invoice_edit.html", {
        "invoices": invoices,
        "first_invoice": first_invoice,
        "bulk_group_id": bulk_group_id,
        "schemes": schemes,
        "today": today,
    })


@login_required
def bulk_invoice_delete(request, bulk_group_id):
    invoices = Invoice.objects.filter(bulk_group_id=bulk_group_id, is_bulk=True)
    if not invoices.exists():
        messages.error(request, "Bulk invoices not found")
        return redirect("invoice_list")
    
    if request.method == "POST":
        count = invoices.count()
        for invoice in invoices:
            invoice.beneficiary.recalculate_totals()
        invoices.delete()
        messages.success(request, f"Successfully deleted {count} bulk invoices")
        return redirect("invoice_list")
    
    return render(request, "accounting_app/bulk_invoice_delete.html", {
        "invoices": invoices,
        "bulk_group_id": bulk_group_id,
    })


@login_required
def invoice_edit(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    clients = Beneficiary.objects.filter(is_active=True).order_by('name')
    today = timezone.now().date()
    
    if request.method == "POST":
        if "add_item" in request.POST:
            desc = request.POST.get("description")
            qty = request.POST.get("quantity")
            price = request.POST.get("unit_price")
            InvoiceItem.objects.create(
                invoice=invoice, description=desc, quantity=qty, unit_price=price
            )
            return redirect("invoice_edit", pk=pk)
        if "change_status" in request.POST:
            new_status = request.POST.get("status")
            if new_status in [choice[0] for choice in Invoice.STATUS_CHOICES]:
                invoice.status = new_status
                invoice.save()
                messages.success(request, f"Invoice status changed to {invoice.get_status_display()}")
            return redirect("invoice_edit", pk=pk)
        
        # Update invoice fields directly
        try:
            invoice.household_count = int(request.POST.get('household_count', invoice.household_count) or 0)
        except:
            invoice.household_count = 0
            
        try:
            invoice.cost_per_unit = float(request.POST.get('cost_per_unit', invoice.cost_per_unit) or 0)
        except:
            invoice.cost_per_unit = 0
            
        issue_date = request.POST.get('issue_date')
        if issue_date:
            invoice.issue_date = issue_date
            
        due_date = request.POST.get('due_date')
        if due_date:
            invoice.due_date = due_date
            
        try:
            invoice.tax_rate = float(request.POST.get('tax_rate', invoice.tax_rate) or 0)
        except:
            invoice.tax_rate = 0
            
        try:
            invoice.discount = float(request.POST.get('discount', invoice.discount) or 0)
        except:
            invoice.discount = 0
            
        invoice.notes = request.POST.get('notes', '')
        invoice.terms = request.POST.get('terms', '')
        invoice.status = request.POST.get('status', invoice.status)
        
        # Recalculate totals
        subtotal = invoice.household_count * invoice.cost_per_unit
        tax_amount = subtotal * (invoice.tax_rate / 100) if invoice.tax_rate else 0
        invoice.tax_amount = tax_amount
        invoice.total_amount = subtotal + tax_amount - invoice.discount
        
        invoice.save()
        
        if invoice.beneficiary:
            invoice.beneficiary.recalculate_totals()
        
        messages.success(request, "Invoice updated successfully")
        return redirect("invoice_list")
    else:
        form = InvoiceForm(instance=invoice)
    
    return render(request, "accounting_app/invoice_form.html", {
        "form": form, 
        "invoice": invoice, 
        "action": "Update",
        "clients": clients,
        "status_choices": Invoice.STATUS_CHOICES,
        "today": today,
        "user": request.user
    })


@login_required
def invoice_delete_item(request, pk, item_pk):
    item = get_object_or_404(InvoiceItem, pk=item_pk)
    invoice_pk = item.invoice.pk
    item.delete()
    messages.success(request, "Item removed")
    return redirect("invoice_edit", pk=invoice_pk)


@login_required
def invoice_delete(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    if request.method == "POST":
        invoice.status = "cancelled"
        invoice.save()
        ActivityLog.objects.create(
            user=request.user,
            action="Voided",
            model_name="Invoice",
            object_id=invoice.pk,
            description=f"Voided invoice {invoice.invoice_number} for {invoice.beneficiary.name}",
            ip_address=request.META.get('REMOTE_ADDR'),
        )
        save_deleted_record(invoice, request.user)
        invoice.delete()
        messages.success(request, "Invoice deleted")
        return redirect("invoice_list")
    return render(request, "accounting_app/invoice_delete.html", {"invoice": invoice})


@login_required
@login_required
def expense_list(request):
    today = timezone.now().date()
    year_start = today.replace(month=1, day=1)
    
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    category_filter = request.GET.get('category')
    
    expenses = Expense.objects.filter(items__isnull=False).distinct()
    
    if from_date:
        from_date = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
    else:
        from_date = year_start
    
    if to_date:
        to_date = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
    else:
        to_date = today
    
    expenses = expenses.filter(expense_date__gte=from_date, expense_date__lte=to_date)
    
    if category_filter:
        expenses = expenses.filter(items__category=category_filter).distinct()
    
    total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or 0
    
    return render(request, "accounting_app/expense_list.html", {
        "expenses": expenses, 
        "category_filter": category_filter,
        "from_date": from_date,
        "to_date": to_date,
        "total_expenses": total_expenses
    })


@login_required
def expense_create(request):
    expense_number = generate_expense_number()
    vendors = Vendor.objects.all()
    accounts = Account.objects.filter(is_active=True)
    today = timezone.now()
    
    if request.method == "POST":
        form = ExpenseForm(request.POST, request.FILES)
        items_data = []
        total_amount = Decimal('0.00')
        
        categories = request.POST.getlist('category')
        descriptions = request.POST.getlist('description_item')
        quantities = request.POST.getlist('quantity')
        unit_prices = request.POST.getlist('unit_price')
        
        for i in range(len(categories)):
            if categories[i] and descriptions[i]:
                quantity = Decimal(quantities[i]) if quantities[i] else Decimal('1')
                unit_price = Decimal(unit_prices[i]) if unit_prices[i] else Decimal('0.00')
                amount = quantity * unit_price
                total_amount += amount
                items_data.append({
                    'category': categories[i],
                    'description': descriptions[i],
                    'quantity': quantity,
                    'unit_price': unit_price,
                    'amount': amount
                })
        
        if not items_data:
            messages.error(request, "Please add at least one expense item with category, description, and valid amount")
            return render(request, "accounting_app/expense_form.html", {
                "form": form, 
                "action": "Create", 
                "expense_number": expense_number,
                "vendors": vendors,
                "accounts": accounts,
                "today": today
            })
        
        if form.is_valid():
            expense = form.save(commit=False)
            expense.expense_number = expense_number
            expense.amount = total_amount
            expense.created_by = request.user
            expense.save()
            
            for item_data in items_data:
                ExpenseItem.objects.create(
                    expense=expense,
                    category=item_data['category'],
                    description=item_data['description'],
                    quantity=item_data['quantity'],
                    unit_price=item_data['unit_price'],
                    amount=item_data['amount']
                )
            
            if expense.account:
                expense.account.balance -= expense.amount
                expense.account.save()
            
            messages.success(request, f"Expense {expense.expense_number} created successfully with {len(items_data)} items")
            return redirect("expense_list")
        else:
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            messages.error(request, f"Error creating expense: {'; '.join(error_messages)}")
    else:
        form = ExpenseForm()
    
    return render(request, "accounting_app/expense_form.html", {
        "form": form, 
        "action": "Create", 
        "expense_number": expense_number,
        "vendors": vendors,
        "accounts": accounts,
        "today": today
    })


@login_required
def expense_edit(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    old_amount = expense.amount
    old_account = expense.account
    vendors = Vendor.objects.all()
    accounts = Account.objects.filter(is_active=True)
    
    if request.method == "POST":
        form = ExpenseForm(request.POST, request.FILES, instance=expense)
        items_data = []
        total_amount = Decimal('0.00')
        
        categories = request.POST.getlist('category')
        descriptions = request.POST.getlist('description_item')
        quantities = request.POST.getlist('quantity')
        unit_prices = request.POST.getlist('unit_price')
        amounts = request.POST.getlist('amount')
        item_ids = request.POST.getlist('item_id')
        
        for i in range(len(categories)):
            if categories[i] and descriptions[i]:
                quantity = Decimal(quantities[i]) if quantities[i] else Decimal('1')
                unit_price = Decimal(unit_prices[i]) if unit_prices[i] else Decimal('0.00')
                amount = quantity * unit_price
                total_amount += amount
                items_data.append({
                    'category': categories[i],
                    'description': descriptions[i],
                    'quantity': quantity,
                    'unit_price': unit_price,
                    'amount': amount
                })
        
        if not items_data:
            messages.error(request, "Please add at least one expense item with category, description, and valid amount")
            return render(request, "accounting_app/expense_form.html", {
                "form": form, 
                "action": "Update", 
                "expense_number": expense.expense_number,
                "vendors": vendors,
                "accounts": accounts,
                "today": timezone.now(),
                "expense_items": expense.items.all()
            })
        
        if form.is_valid():
            expense = form.save()
            expense.items.all().delete()
            
            for item_data in items_data:
                ExpenseItem.objects.create(
                    expense=expense,
                    category=item_data['category'],
                    description=item_data['description'],
                    quantity=item_data['quantity'],
                    unit_price=item_data['unit_price'],
                    amount=item_data['amount']
                )
            
            expense.amount = total_amount
            expense.save()
            
            if expense.account:
                if old_account:
                    old_account.balance += old_amount
                    old_account.save()
                expense.account.balance -= expense.amount
                expense.account.save()
            
            messages.success(request, "Expense updated successfully")
            return redirect("expense_list")
        else:
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            messages.error(request, f"Error updating expense: {'; '.join(error_messages)}")
    else:
        form = ExpenseForm(instance=expense)
    
    return render(request, "accounting_app/expense_form.html", {
        "form": form, 
        "action": "Update", 
        "expense_number": expense.expense_number,
        "vendors": vendors,
        "accounts": accounts,
        "today": timezone.now(),
        "expense_items": expense.items.all()
    })


@login_required
def expense_delete(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == "POST":
        save_deleted_record(expense, request.user)
        if expense.account:
            expense.account.balance += expense.amount
            expense.account.save()
        expense.delete()
        messages.success(request, "Expense deleted successfully")
        return redirect("expense_list")
    return render(request, "accounting_app/expense_delete.html", {"expense": expense})


@login_required
def scheme_reports(request):
    today = timezone.now().date()
    year_start = today.replace(month=1, day=1)
    
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if from_date:
        from_date = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
    else:
        from_date = year_start
    
    if to_date:
        to_date = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
    else:
        to_date = today
    
    schemes = Beneficiary.SCHEME_CHOICES
    scheme_data = []
    
    grand_total_invoiced = Decimal('0.00')
    grand_total_collected = Decimal('0.00')
    grand_total_balance = Decimal('0.00')
    grand_total_households = 0
    
    for scheme_code, scheme_name in schemes:
        beneficiaries = Beneficiary.objects.filter(scheme=scheme_code, is_active=True)
        total_households = sum(beneficiary.household_count for beneficiary in beneficiaries)
        
        beneficiary_ids = beneficiaries.values_list('id', flat=True)
        
        total_invoiced = Invoice.objects.filter(
            beneficiary_id__in=beneficiary_ids,
            issue_date__gte=from_date,
            issue_date__lte=to_date
        ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        
        total_collected = Payment.objects.filter(
            beneficiary_id__in=beneficiary_ids,
            payment_date__gte=from_date,
            payment_date__lte=to_date
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        total_balance = total_invoiced - total_collected
        
        scheme_data.append({
            'code': scheme_code,
            'name': scheme_name,
            'beneficiaries': beneficiaries.count(),
            'households': total_households,
            'total_invoiced': total_invoiced,
            'total_collected': total_collected,
            'balance': total_balance,
            'collection_rate': (total_collected / total_invoiced * 100) if total_invoiced > 0 else 0,
        })
        
        grand_total_invoiced += total_invoiced
        grand_total_collected += total_collected
        grand_total_balance += total_balance
        grand_total_households += total_households

    
    collection_rate = (grand_total_collected / grand_total_invoiced * 100) if grand_total_invoiced > 0 else 0
    
    context = {
        "from_date": from_date,
        "to_date": to_date,
        "scheme_data": scheme_data,
        "grand_total_invoiced": grand_total_invoiced,
        "grand_total_collected": grand_total_collected,
        "grand_total_balance": grand_total_balance,
        "grand_total_households": grand_total_households,
        "collection_rate": collection_rate,
    }
    return render(request, "accounting_app/scheme_reports.html", context)


@login_required
def export_scheme_report_pdf(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if from_date:
        from_date = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
    else:
        from_date = timezone.now().date().replace(month=1, day=1)
    
    if to_date:
        to_date = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
    else:
        to_date = timezone.now().date()
    
    schemes = Beneficiary.SCHEME_CHOICES
    scheme_data = []
    
    grand_total_invoiced = Decimal('0.00')
    grand_total_collected = Decimal('0.00')
    grand_total_balance = Decimal('0.00')
    grand_total_households = 0
    
    for scheme_code, scheme_name in schemes:
        beneficiaries = Beneficiary.objects.filter(scheme=scheme_code, is_active=True)
        total_households = sum(beneficiary.household_count for beneficiary in beneficiaries)
        
        beneficiary_ids = beneficiaries.values_list('id', flat=True)
        
        total_invoiced = Invoice.objects.filter(
            beneficiary_id__in=beneficiary_ids,
            issue_date__gte=from_date,
            issue_date__lte=to_date
        ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        
        total_collected = Payment.objects.filter(
            beneficiary_id__in=beneficiary_ids,
            payment_date__gte=from_date,
            payment_date__lte=to_date
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        total_balance = total_invoiced - total_collected
        
        scheme_data.append({
            'name': scheme_name,
            'beneficiaries': beneficiaries.count(),
            'households': total_households,
            'total_invoiced': total_invoiced,
            'total_collected': total_collected,
            'balance': total_balance,
        })
        
        grand_total_invoiced += total_invoiced
        grand_total_collected += total_collected
        grand_total_balance += total_balance
        grand_total_households += total_households
    
    user_profile = request.user.userprofile
    currency_symbol = user_profile.get_currency_symbol()
    company_name = user_profile.company_name or "Nkulawua"
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Scheme_Performance_Report_{from_date}_{to_date}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    elements = []
    
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Title_Custom', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=18, spaceAfter=5, textColor=colors.HexColor('#2c3e50')))
    
    if user_profile.logo:
        try:
            logo = Image(user_profile.logo.path, width=2*cm, height=2*cm)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 10))
        except:
            pass
    
    company_data = [[Paragraph(company_name, styles['Title_Custom'])]]
    company_data.append([Paragraph("Scheme Performance Report", styles['Title_Custom'])])
    
    company_table = Table(company_data, colWidths=[16*cm])
    company_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(company_table)
    elements.append(Spacer(1, 10))
    
    report_info = [
        ["Period:", f"From {from_date} to {to_date}"],
        ["Generated:", timezone.now().strftime('%Y-%m-%d %H:%M')],
    ]
    
    info_table = Table(report_info, colWidths=[3*cm, 13*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    scheme_data_rows = [["Scheme", "Clients", "Households", "Total Invoiced", "Total Collected", "Balance"]]
    
    for scheme in scheme_data:
        scheme_data_rows.append([
            scheme['name'],
            str(scheme['clients']),
            str(scheme['households']),
            f"{currency_symbol}{scheme['total_invoiced']:,.2f}",
            f"{currency_symbol}{scheme['total_collected']:,.2f}",
            f"{currency_symbol}{scheme['balance']:,.2f}",
        ])
    
    scheme_data_rows.append(["GRAND TOTAL", "-", str(grand_total_households), f"{currency_symbol}{grand_total_invoiced:,.2f}", f"{currency_symbol}{grand_total_collected:,.2f}", f"{currency_symbol}{grand_total_balance:,.2f}"])
    
    table = Table(scheme_data_rows, colWidths=[3*cm, 2*cm, 2.5*cm, 3*cm, 3*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#27ae60')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
    ]))
    elements.append(table)
    
    doc.build(elements)
    return response


@login_required
def export_scheme_report_excel(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if from_date:
        from_date = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
    else:
        from_date = timezone.now().date().replace(month=1, day=1)
    
    if to_date:
        to_date = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
    else:
        to_date = timezone.now().date()
    
    schemes = Beneficiary.SCHEME_CHOICES
    scheme_data = []
    
    grand_total_invoiced = Decimal('0.00')
    grand_total_collected = Decimal('0.00')
    grand_total_balance = Decimal('0.00')
    grand_total_households = 0
    
    for scheme_code, scheme_name in schemes:
        beneficiaries = Beneficiary.objects.filter(scheme=scheme_code, is_active=True)
        total_households = sum(beneficiary.household_count for beneficiary in beneficiaries)
        
        beneficiary_ids = beneficiaries.values_list('id', flat=True)
        
        total_invoiced = Invoice.objects.filter(
            beneficiary_id__in=beneficiary_ids,
            issue_date__gte=from_date,
            issue_date__lte=to_date
        ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        
        total_collected = Payment.objects.filter(
            beneficiary_id__in=beneficiary_ids,
            payment_date__gte=from_date,
            payment_date__lte=to_date
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        total_balance = total_invoiced - total_collected
        
        scheme_data.append({
            'name': scheme_name,
            'beneficiaries': beneficiaries.count(),
            'households': total_households,
            'total_invoiced': total_invoiced,
            'total_collected': total_collected,
            'balance': total_balance,
        })
        
        grand_total_invoiced += total_invoiced
        grand_total_collected += total_collected
        grand_total_balance += total_balance
        grand_total_households += total_households
    
    user_profile = request.user.userprofile
    currency_symbol = user_profile.get_currency_symbol()
    company_name = user_profile.company_name or "Nkulawua"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Scheme Performance"
    
    ws.merge_cells('A1:F1')
    ws['A1'] = company_name
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = "SCHEME PERFORMANCE REPORT"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws['A4'] = "Period:"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = f"From {from_date} to {to_date}"
    
    ws['A5'] = "Generated:"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = timezone.now().strftime('%Y-%m-%d %H:%M')
    
    ws.append([])
    
    headers = ["Scheme", "Clients", "Households", "Total Invoiced", "Total Collected", "Balance"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col in range(1, 7):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for scheme in scheme_data:
        ws.append([
            scheme['name'],
            scheme['clients'],
            scheme['households'],
            f"{currency_symbol}{scheme['total_invoiced']:,.2f}",
            f"{currency_symbol}{scheme['total_collected']:,.2f}",
            f"{currency_symbol}{scheme['balance']:,.2f}",
        ])
    
    ws.append(["GRAND TOTAL", "-", grand_total_households, f"{currency_symbol}{grand_total_invoiced:,.2f}", f"{currency_symbol}{grand_total_collected:,.2f}", f"{currency_symbol}{grand_total_balance:,.2f}"])
    
    green_fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
    last_row = ws.max_row
    for col in range(1, 7):
        cell = ws.cell(row=last_row, column=col)
        cell.fill = green_fill
        cell.font = Font(bold=True, color="FFFFFF", size=11)
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Scheme_Performance_Report_{from_date}_{to_date}.xlsx"'
    wb.save(response)
    return response


@login_required
def send_scheme_report_email(request):
    if request.method == "POST":
        recipient_email = request.POST.get('email')
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')
        
        if not recipient_email:
            messages.error(request, "Please enter an email address")
            return redirect("scheme_reports")
        
        if from_date:
            from_date = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
        else:
            from_date = timezone.now().date().replace(month=1, day=1)
        
        if to_date:
            to_date = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
        else:
            to_date = timezone.now().date()
        
        schemes = Beneficiary.SCHEME_CHOICES
        scheme_data = []
        
        grand_total_invoiced = Decimal('0.00')
        grand_total_collected = Decimal('0.00')
        grand_total_balance = Decimal('0.00')
        
        for scheme_code, scheme_name in schemes:
            beneficiaries = Beneficiary.objects.filter(scheme=scheme_code, is_active=True)
            beneficiary_ids = beneficiaries.values_list('id', flat=True)
            
            total_invoiced = Invoice.objects.filter(
                beneficiary_id__in=beneficiary_ids,
                issue_date__gte=from_date,
                issue_date__lte=to_date
            ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
            
            total_collected = Payment.objects.filter(
                beneficiary_id__in=beneficiary_ids,
                payment_date__gte=from_date,
                payment_date__lte=to_date
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            
            total_balance = total_invoiced - total_collected
            
            scheme_data.append({
                'name': scheme_name,
                'beneficiaries': beneficiaries.count(),
                'total_invoiced': total_invoiced,
                'total_collected': total_collected,
                'balance': total_balance,
            })
            
            grand_total_invoiced += total_invoiced
            grand_total_collected += total_collected
            grand_total_balance += total_balance
        
        user_profile = request.user.userprofile
        currency_symbol = user_profile.get_currency_symbol()
        company_name = user_profile.company_name or "Nkulawua"
        
        subject = f"{company_name} - Scheme Performance Report ({from_date} to {to_date})"
        
        message_lines = [
            f"Scheme Performance Report from {company_name}",
            f"Period: {from_date} to {to_date}",
            "",
            "SCHEME PERFORMANCE:",
            "=" * 80,
            f"{'Scheme':<15} {'Clients':>10} {'Invoiced':>15} {'Collected':>15} {'Balance':>15}",
            "-" * 80,
        ]
        
        for scheme in scheme_data:
            message_lines.append(f"{scheme['name']:<15} {scheme['clients']:>10} {currency_symbol}{scheme['total_invoiced']:>13,.2f} {currency_symbol}{scheme['total_collected']:>13,.2f} {currency_symbol}{scheme['balance']:>13,.2f}")
        
        message_lines.extend([
            "=" * 80,
            f"{'GRAND TOTAL':<15} {'':<10} {currency_symbol}{grand_total_invoiced:>13,.2f} {currency_symbol}{grand_total_collected:>13,.2f} {currency_symbol}{grand_total_balance:>13,.2f}",
            "=" * 80,
            f"\nGenerated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}",
            f"\nBest regards,\n{company_name}"
        ])
        
        message = "\n".join(message_lines)
        
        try:
            send_mail(
                subject,
                message,
                user_profile.company_name + " <noreply@" + request.get_host() + ">",
                [recipient_email],
                fail_silently=False,
            )
            messages.success(request, f"Scheme report sent successfully to {recipient_email}")
        except Exception as e:
            messages.error(request, f"Failed to send email: {str(e)}")
        
        return redirect("scheme_reports")
    
    return redirect("scheme_reports")


@login_required
def share_scheme_report(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    method = request.GET.get('method', '')
    phone = request.GET.get('phone', '')
    
    if from_date:
        from_date = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
    else:
        from_date = timezone.now().date().replace(month=1, day=1)
    
    if to_date:
        to_date = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
    else:
        to_date = timezone.now().date()
    
    schemes = Beneficiary.SCHEME_CHOICES
    scheme_data = []
    
    grand_total_invoiced = Decimal('0.00')
    grand_total_collected = Decimal('0.00')
    grand_total_balance = Decimal('0.00')
    
    for scheme_code, scheme_name in schemes:
        beneficiaries = Beneficiary.objects.filter(scheme=scheme_code, is_active=True)
        beneficiary_ids = beneficiaries.values_list('id', flat=True)
        
        total_invoiced = Invoice.objects.filter(
            beneficiary_id__in=beneficiary_ids,
            issue_date__gte=from_date,
            issue_date__lte=to_date
        ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        
        total_collected = Payment.objects.filter(
            beneficiary_id__in=beneficiary_ids,
            payment_date__gte=from_date,
            payment_date__lte=to_date
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        total_balance = total_invoiced - total_collected
        
        scheme_data.append({
            'name': scheme_name,
            'beneficiaries': beneficiaries.count(),
            'total_invoiced': total_invoiced,
            'total_collected': total_collected,
            'balance': total_balance,
        })
        
        grand_total_invoiced += total_invoiced
        grand_total_collected += total_collected
        grand_total_balance += total_balance
    
    user_profile = request.user.userprofile
    currency_symbol = user_profile.get_currency_symbol()
    company_name = user_profile.company_name or "Nkulawua"
    
    report_message = f"""
{company_name}
SCHEME PERFORMANCE REPORT
Period: {from_date} to {to_date}

SUMMARY:
--------
"""
    
    for scheme in scheme_data:
        report_message += f"{scheme['name']}: Invoiced={currency_symbol}{scheme['total_invoiced']:,.2f}, Collected={currency_symbol}{scheme['total_collected']:,.2f}, Balance={currency_symbol}{scheme['balance']:,.2f}\n"
    
    report_message += f"""
GRAND TOTAL:
Invoiced: {currency_symbol}{grand_total_invoiced:,.2f}
Collected: {currency_symbol}{grand_total_collected:,.2f}
Balance: {currency_symbol}{grand_total_balance:,.2f}

Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}
"""
    
    if method == 'whatsapp':
        from urllib.parse import quote
        phone = phone.strip().replace(' ', '').replace('+', '')
        if not phone.startswith('265') and not phone.startswith('0'):
            phone = '265' + phone.lstrip('0')
        encoded_message = quote(report_message)
        whatsapp_url = f"https://wa.me/{phone}?text={encoded_message}"
        return HttpResponseRedirect(whatsapp_url)
    
    elif method == 'sms':
        profile = user_profile
        if profile.sms_provider == "none":
            messages.info(request, f"SMS to {phone}: {report_message[:160]}... (Demo Mode)")
        else:
            from .sms_service import send_sms_to_client
            result = send_sms_to_client(profile, phone, report_message[:160])
            if result["status"] == "success":
                messages.success(request, f"SMS sent successfully to {phone}")
            else:
                messages.error(request, f"Failed to send SMS: {result['message']}")
        return redirect("scheme_reports")
    
    return redirect("scheme_reports")


@login_required
def expense_reports(request):
    today = timezone.now().date()
    year_start = today.replace(month=1, day=1)
    
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    category_filter = request.GET.get('category')
    
    expenses = Expense.objects.filter(items__isnull=False).distinct()
    
    if from_date:
        from_date = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
    else:
        from_date = year_start
    
    if to_date:
        to_date = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
    else:
        to_date = today
    
    expenses = expenses.filter(expense_date__gte=from_date, expense_date__lte=to_date)
    
    if category_filter:
        expenses = expenses.filter(items__category=category_filter).distinct()
    
    expenses_with_items = []
    category_totals = {}
    
    for expense in expenses:
        items = expense.items.all()
        if category_filter:
            items = items.filter(category=category_filter)
        
        for item in items:
            category_name = item.get_category_display()
            if category_name not in category_totals:
                category_totals[category_name] = {
                    'category': item.category,
                    'name': category_name,
                    'total': Decimal('0.00'),
                    'items': []
                }
            category_totals[category_name]['total'] += item.amount
            category_totals[category_name]['items'].append({
                'item': item,
                'expense': expense
            })
    
    total_expenses = sum(cat['total'] for cat in category_totals.values()) if category_totals else Decimal('0.00')
    
    context = {
        "from_date": from_date,
        "to_date": to_date,
        "category_filter": category_filter,
        "total_expenses": total_expenses,
        "category_totals": category_totals,
        "category_list": list(category_totals.values()) if category_totals else [],
    }
    return render(request, "accounting_app/expense_reports.html", context)


@login_required
def export_expenses_pdf(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    category_filter = request.GET.get('category')
    
    if from_date:
        from_date = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
    else:
        from_date = timezone.now().date().replace(month=1, day=1)
    
    if to_date:
        to_date = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
    else:
        to_date = timezone.now().date()
    
    expenses = Expense.objects.filter(
        expense_date__gte=from_date, 
        expense_date__lte=to_date,
        items__isnull=False
    ).distinct()
    
    if category_filter:
        expenses = expenses.filter(items__category=category_filter).distinct()
    
    total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or 0
    
    user_profile = request.user.userprofile
    currency_symbol = user_profile.get_currency_symbol()
    company_name = user_profile.company_name or "Nkulawua"
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Expense_Report_{from_date}_{to_date}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    elements = []
    
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Center', alignment=TA_CENTER))
    styles.add(ParagraphStyle(name='Title_Custom', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=18, spaceAfter=5, textColor=colors.HexColor('#2c3e50')))
    styles.add(ParagraphStyle(name='SubTitle', parent=styles['Normal'], alignment=TA_CENTER, fontSize=11, textColor=colors.grey))
    
    company_data = []
    if user_profile.logo:
        try:
            logo_path = user_profile.logo.path
            if os.path.exists(logo_path):
                logo = Image(logo_path, width=3*cm, height=2*cm)
                company_data.append([logo])
        except:
            pass
    
    company_data.append([Paragraph(company_name, styles['Title_Custom'])])
    company_data.append([Paragraph("Expense Report", styles['SubTitle'])])
    
    company_table = Table(company_data, colWidths=[16*cm])
    company_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(company_table)
    elements.append(Spacer(1, 10))
    
    report_info = [
        ["REPORT DETAILS", "", ""],
        ["Period:", f"From {from_date} to {to_date}", ""],
        ["Category Filter:", f"{category_filter or 'All Categories'}", ""],
        ["Generated:", timezone.now().strftime('%Y-%m-%d %H:%M'), ""],
    ]
    
    info_table = Table(report_info, colWidths=[4*cm, 8*cm, 4*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('SPAN', (0, 0), (-1, 0)),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    expense_data = [["Date", "Ref #", "Category", "Description", "Amount"]]
    
    for expense in expenses:
        for item in expense.items.all():
            expense_data.append([
                expense.expense_date.strftime('%Y-%m-%d'),
                expense.expense_number,
                item.get_category_display(),
                item.description[:40],
                f"{currency_symbol}{item.amount:,.2f}"
            ])
    
    table = Table(expense_data, colWidths=[2.5*cm, 2.5*cm, 3*cm, 5*cm, 3*cm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 15))
    
    total_table = Table([
        ["TOTAL EXPENSES:", f"{currency_symbol}{total_expenses:,.2f}"]
    ], colWidths=[10*cm, 6*cm])
    total_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#e8f5e9')),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#27ae60')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(total_table)
    
    doc.build(elements)
    return response


@login_required
def export_expenses_excel(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    category_filter = request.GET.get('category')
    
    if from_date:
        from_date = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
    else:
        from_date = timezone.now().date().replace(month=1, day=1)
    
    if to_date:
        to_date = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
    else:
        to_date = timezone.now().date()
    
    expenses = Expense.objects.filter(
        expense_date__gte=from_date, 
        expense_date__lte=to_date,
        items__isnull=False
    ).distinct()
    
    if category_filter:
        expenses = expenses.filter(items__category=category_filter).distinct()
    
    total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or 0
    
    user_profile = request.user.userprofile
    currency_symbol = user_profile.get_currency_symbol()
    company_name = user_profile.company_name or "Nkulawua"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Report"
    
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    red_fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
    green_fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    bold_font = Font(bold=True, size=10)
    
    ws.merge_cells('A1:E1')
    ws['A1'] = company_name
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:E2')
    ws['A2'] = "EXPENSE REPORT"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws['A4'] = "Period:"
    ws['A4'].font = bold_font
    ws['B4'] = f"From {from_date} to {to_date}"
    
    ws['A5'] = "Category Filter:"
    ws['A5'].font = bold_font
    ws['B5'] = category_filter or "All Categories"
    
    ws['A6'] = "Generated:"
    ws['A6'].font = bold_font
    ws['B6'] = timezone.now().strftime('%Y-%m-%d %H:%M')
    
    ws.append([])
    
    headers = ["Date", "Ref #", "Category", "Description", "Amount"]
    ws.append(headers)
    for col in range(1, 6):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = red_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for expense in expenses:
        for item in expense.items.all():
            ws.append([
                expense.expense_date.strftime('%Y-%m-%d'),
                expense.expense_number,
                item.get_category_display(),
                item.description,
                f"{currency_symbol}{item.amount:,.2f}"
            ])
    
    ws.append([])
    ws.append(["", "", "", "TOTAL:", f"{currency_symbol}{total_expenses:,.2f}"])
    last_row = ws.max_row
    for col in range(1, 6):
        cell = ws.cell(row=last_row, column=col)
        cell.font = Font(bold=True, size=11)
        if col == 5:
            cell.fill = green_fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Expense_Report_{from_date}_{to_date}.xlsx"'
    wb.save(response)
    return response


@login_required
def send_expense_report_email(request):
    if request.method == "POST":
        recipient_email = request.POST.get('email')
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')
        category_filter = request.POST.get('category')
        
        if not recipient_email:
            messages.error(request, "Please enter an email address")
            return redirect("expense_reports")
        
        if from_date:
            from_date = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
        else:
            from_date = timezone.now().date().replace(month=1, day=1)
        
        if to_date:
            to_date = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
        else:
            to_date = timezone.now().date()
        
        expenses = Expense.objects.filter(
            expense_date__gte=from_date, 
            expense_date__lte=to_date,
            items__isnull=False
        ).distinct()
        
        if category_filter:
            expenses = expenses.filter(items__category=category_filter).distinct()
        
        total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or 0
        
        user_profile = request.user.userprofile
        currency_symbol = user_profile.get_currency_symbol()
        company_name = user_profile.company_name or "Nkulawua"
        
        subject = f"{company_name} - Expense Report ({from_date} to {to_date})"
        
        message_lines = [
            f"Expense Report from {company_name}",
            f"Period: {from_date} to {to_date}",
            f"Category Filter: {category_filter or 'All Categories'}",
            "",
            "EXPENSE DETAILS:",
            "=" * 60,
        ]
        
        for expense in expenses:
            message_lines.append(f"\nDate: {expense.expense_date} | Ref: {expense.expense_number}")
            for item in expense.items.all():
                message_lines.append(f"  - {item.get_category_display()}: {item.description[:50]} - {currency_symbol}{item.amount:,.2f}")
            message_lines.append(f"  Subtotal: {currency_symbol}{expense.amount:,.2f}")
        
        message_lines.extend([
            "",
            "=" * 60,
            f"TOTAL EXPENSES: {currency_symbol}{total_expenses:,.2f}",
            "=" * 60,
            f"\nGenerated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}",
            f"\nBest regards,\n{company_name}"
        ])
        
        message = "\n".join(message_lines)
        
        try:
            send_mail(
                subject,
                message,
                user_profile.company_name + " <noreply@" + request.get_host() + ">",
                [recipient_email],
                fail_silently=False,
            )
            messages.success(request, f"Expense report sent successfully to {recipient_email}")
        except Exception as e:
            messages.error(request, f"Failed to send email: {str(e)}")
        
        return redirect("expense_reports")
    
    return redirect("expense_reports")


@login_required
def share_expense_report(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    category_filter = request.GET.get('category')
    method = request.GET.get('method', '')
    phone = request.GET.get('phone', '')
    
    if from_date:
        from_date = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
    else:
        from_date = timezone.now().date().replace(month=1, day=1)
    
    if to_date:
        to_date = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
    else:
        to_date = timezone.now().date()
    
    expenses = Expense.objects.filter(
        expense_date__gte=from_date, 
        expense_date__lte=to_date,
        items__isnull=False
    ).distinct()
    
    if category_filter:
        expenses = expenses.filter(items__category=category_filter).distinct()
    
    total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or 0
    
    user_profile = request.user.userprofile
    currency_symbol = user_profile.get_currency_symbol()
    company_name = user_profile.company_name or "Nkulawua"
    
    report_message = f"""
{company_name}
EXPENSE REPORT
Period: {from_date} to {to_date}
Category: {category_filter or 'All'}

SUMMARY:
--------
Total Expenses: {currency_symbol}{total_expenses:,.2f}
Number of Transactions: {expenses.count()}

Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}
"""
    
    if method == 'whatsapp':
        from urllib.parse import quote
        phone = phone.strip().replace(' ', '').replace('+', '')
        if not phone.startswith('265') and not phone.startswith('0'):
            phone = '265' + phone.lstrip('0')
        encoded_message = quote(report_message)
        whatsapp_url = f"https://wa.me/{phone}?text={encoded_message}"
        return HttpResponseRedirect(whatsapp_url)
    
    elif method == 'sms':
        profile = user_profile
        if profile.sms_provider == "none":
            messages.info(request, f"SMS to {phone}: {report_message[:160]}... (Demo Mode)")
        else:
            from .sms_service import send_sms_to_client
            result = send_sms_to_client(profile, phone, report_message[:160])
            if result["status"] == "success":
                messages.success(request, f"SMS sent successfully to {phone}")
            else:
                messages.error(request, f"Failed to send SMS: {result['message']}")
        return redirect("expense_reports")
    
    return redirect("expense_reports")


@login_required
def bulk_payment_create(request):
    today = timezone.now().date()
    beneficiaries = Beneficiary.objects.filter(is_active=True).order_by('scheme', 'name')
    schemes = Beneficiary.SCHEME_CHOICES
    accounts = Account.objects.filter(is_active=True, account_type='revenue').order_by('code')

    base_ctx = {
        "beneficiaries": beneficiaries,
        "schemes": schemes,
        "accounts": accounts,
        "today": today,
        "symbol": '',
    }
    if hasattr(request.user, 'userprofile'):
        base_ctx['symbol'] = request.user.userprofile.get_currency_symbol()

    if request.method == "POST":
        amount = float(request.POST.get('amount', 0))
        raw_payment_date = request.POST.get('payment_date')
        payment_date = date.fromisoformat(raw_payment_date) if raw_payment_date else None
        account_id = request.POST.get('account')
        payment_method = request.POST.get('payment_method', 'bank_transfer')
        reference = request.POST.get('reference', '')
        notes = request.POST.get('notes', '')

        selected_ids = request.POST.getlist('selected_beneficiaries')

        post_ctx = {
            **base_ctx,
            "amount": amount,
            "payment_date": payment_date,
            "account_id": account_id,
            "payment_method": payment_method,
            "reference": reference,
            "notes": notes,
        }

        if not selected_ids:
            messages.error(request, "Please select at least one beneficiary")
            return render(request, "accounting_app/bulk_payment_form.html", post_ctx)

        if amount <= 0:
            messages.error(request, "Please enter a valid amount greater than 0")
            return render(request, "accounting_app/bulk_payment_form.html", post_ctx)

        if not payment_date:
            messages.error(request, "Please select a payment date")
            return render(request, "accounting_app/bulk_payment_form.html", post_ctx)

        account = None
        if account_id:
            try:
                account = Account.objects.get(pk=account_id)
            except Account.DoesNotExist:
                pass

        bulk_group_id = f"BULK-PAY-{uuid.uuid4().hex[:12]}"

        created_payments = []
        errors = []

        try:
            with transaction.atomic():
                for beneficiary_id in selected_ids:
                    try:
                        client = Beneficiary.objects.get(pk=beneficiary_id)

                        payment = Payment.objects.create(
                            beneficiary=client,
                            invoice=None,
                            amount=amount,
                            account=account,
                            payment_date=payment_date,
                            payment_method=payment_method,
                            reference=reference,
                            notes=notes,
                            created_by=request.user,
                            is_bulk=True,
                            bulk_group_id=bulk_group_id,
                        )
                        created_payments.append(payment)
                    except Beneficiary.DoesNotExist:
                        errors.append(f"Beneficiary #{beneficiary_id} not found")
                    except Exception as e:
                        client_name = getattr(client, 'name', f"Beneficiary #{beneficiary_id}")
                        errors.append(f"{client_name}: {str(e)}")
                
                for p in created_payments:
                    p.beneficiary.recalculate_totals()
        except Exception as e:
            errors.append(f"Database error: {str(e)}")
        
        if created_payments:
            messages.success(request, f"Successfully recorded {len(created_payments)} payments")
        if errors:
            messages.warning(request, f"Failed to record {len(errors)} payments: {'; '.join(errors)}")
        
        return redirect("payment_list")

    return render(request, "accounting_app/bulk_payment_form.html", base_ctx)


@login_required
def payment_list(request):
    payments = Payment.objects.all().order_by('-payment_date')
    accounts = Account.objects.filter(is_active=True, account_type='revenue').order_by('code')
    return render(request, "accounting_app/payment_list.html", {
        "payments": payments,
        "accounts": accounts,
    })


@login_required
def payment_create(request):
    from django.utils import timezone
    
    pre_selected_invoice_id = request.GET.get('invoice')
    pre_selected_invoice = None
    pre_selected_beneficiary = None
    
    if pre_selected_invoice_id:
        try:
            pre_selected_invoice = Invoice.objects.get(pk=pre_selected_invoice_id)
            pre_selected_beneficiary = pre_selected_invoice.beneficiary
        except Invoice.DoesNotExist:
            pass
    
    if request.method == "POST":
        form = PaymentForm(request.POST)
        if form.is_valid():
            beneficiary = form.cleaned_data['beneficiary']
            payment_date = form.cleaned_data['payment_date']
            amount = form.cleaned_data['amount']
            account = form.cleaned_data['account']
            payment_method = form.cleaned_data['payment_method']
            reference = form.cleaned_data['reference']
            notes = form.cleaned_data['notes']

            invoice_id = request.POST.get('invoice')
            invoice = None
            selected_invoice_for_payment = None

            if invoice_id:
                try:
                    invoice = Invoice.objects.get(pk=invoice_id, beneficiary=beneficiary)
                    selected_invoice_for_payment = invoice
                except Invoice.DoesNotExist:
                    pass

            # If no specific invoice selected, allocate payment to oldest outstanding invoices
            if not selected_invoice_for_payment:
                outstanding_invoices = Invoice.objects.filter(
                    beneficiary=beneficiary,
                    status__in=['sent', 'viewed', 'partial', 'overdue']
                ).order_by('due_date', 'issue_date')

                remaining_payment = amount
                for inv in outstanding_invoices:
                    if remaining_payment <= 0:
                        break
                    inv_due = inv.amount_due()
                    if inv_due > 0:
                        # Create payment linked to this invoice
                        payment_amount = min(remaining_payment, inv_due)
                        Payment.objects.create(
                            beneficiary=beneficiary,
                            invoice=inv,
                            amount=payment_amount,
                            account=account,
                            payment_date=payment_date,
                            payment_method=payment_method,
                            reference=reference,
                            notes=notes,
                            created_by=request.user
                        )
                        remaining_payment -= payment_amount

                # If there's still remaining payment, create one without invoice link
                if remaining_payment > 0:
                    Payment.objects.create(
                        beneficiary=beneficiary,
                        invoice=None,
                        amount=remaining_payment,
                        account=account,
                        payment_date=payment_date,
                        payment_method=payment_method,
                        reference=reference,
                        notes=notes,
                        created_by=request.user
                    )
                beneficiary.recalculate_totals()
                messages.success(request, f"Payment of {amount} recorded successfully for {beneficiary.name}")
                return redirect("invoice_list")
            else:
                # Specific invoice selected, create single payment
                payment = Payment.objects.create(
                    beneficiary=beneficiary,
                    invoice=selected_invoice_for_payment,
                    amount=amount,
                    account=account,
                    payment_date=payment_date,
                    payment_method=payment_method,
                    reference=reference,
                    notes=notes,
                    created_by=request.user
                )

            if account:
                account.balance += amount
                account.save()

            beneficiary.recalculate_totals()

            action_label = "Refund" if amount < 0 else "Payment"
            ActivityLog.objects.create(
                user=request.user,
                action=action_label,
                model_name="Payment",
                object_id=getattr(payment, 'pk', None),
                description=f"{action_label} of {amount} for {beneficiary.name}",
                ip_address=request.META.get('REMOTE_ADDR'),
            )

            messages.success(request, f"Payment of {amount} recorded successfully for {beneficiary.name}")
            return redirect("invoice_list")
    else:
        form = PaymentForm()
        if pre_selected_beneficiary:
            form.initial['beneficiary'] = pre_selected_beneficiary
    
    today = timezone.now().date()
    beneficiaries = Beneficiary.objects.filter(is_active=True).order_by('name')
    accounts = Account.objects.filter(is_active=True, account_type='revenue').order_by('code')
    payments = Payment.objects.all().order_by('-payment_date')[:50]
    
    return render(request, "accounting_app/payment_form.html", {
        "form": form,
        "beneficiaries": beneficiaries,
        "accounts": accounts,
        "payments": payments,
        "today": today,
        "action": "Create",
        "pre_selected_beneficiary_id": pre_selected_beneficiary.id if pre_selected_beneficiary else None,
        "pre_selected_invoice_id": pre_selected_invoice.id if pre_selected_invoice else None,
        "pre_selected_invoice_amount_due": pre_selected_invoice.amount_due() if pre_selected_invoice else None,
        "pre_selected_invoice_number": pre_selected_invoice.invoice_number if pre_selected_invoice else None,
    })


@login_required
def bulk_payment_import(request):
    if request.method == "POST" and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(excel_file)
            ws = wb.active
            
            headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
            
            required_fields = ['client_name', 'payment_date', 'amount']
            if not all(field in headers for field in required_fields):
                messages.error(request, "The template must have 'client_name', 'payment_date', and 'amount' columns")
                return redirect("bulk_payment_import")
            
            created_count = 0
            error_count = 0
            errors = []
            
            accounts = Account.objects.filter(is_active=True, account_type='revenue')
            default_account = accounts.first()
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(cell for cell in row):
                    continue
                
                row_data = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        row_data[header] = row[i]
                
                try:
                    client_name = str(row_data.get('client_name', '')).strip()
                    if not client_name:
                        errors.append(f"Row {row_idx}: Beneficiary name is required")
                        error_count += 1
                        continue
                    
                    client = Beneficiary.objects.filter(name__iexact=client_name).first()
                    if not client:
                        errors.append(f"Row {row_idx}: Client '{client_name}' not found")
                        error_count += 1
                        continue
                    
                    payment_date = row_data.get('payment_date')
                    if payment_date:
                        from datetime import datetime
                        if isinstance(payment_date, str):
                            try:
                                payment_date = datetime.strptime(payment_date, '%Y-%m-%d').date()
                            except:
                                payment_date = timezone.now().date()
                        elif hasattr(payment_date, 'date'):
                            payment_date = payment_date.date()
                    else:
                        payment_date = timezone.now().date()
                    
                    amount = row_data.get('amount', 0) or 0
                    try:
                        amount = float(amount)
                    except:
                        amount = 0
                    
                    if amount <= 0:
                        errors.append(f"Row {row_idx}: Invalid amount")
                        error_count += 1
                        continue
                    
                    payment_method = str(row_data.get('payment_method', 'cash')).strip().lower()
                    reference = str(row_data.get('reference', '')).strip()
                    notes = str(row_data.get('notes', '')).strip()
                    account_name = str(row_data.get('account', '')).strip()
                    
                    account = None
                    if account_name:
                        account = Account.objects.filter(name__icontains=account_name).first()
                    if not account:
                        account = default_account
                    
                    Payment.objects.create(
                        beneficiary=beneficiary,
                        amount=amount,
                        account=account,
                        payment_date=payment_date,
                        payment_method=payment_method,
                        reference=reference,
                        notes=notes,
                        created_by=request.user
                    )
                    
                    beneficiary.recalculate_totals()
                    created_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
                    error_count += 1
            
            wb.close()
            
            messages.success(request, f"Import completed: {created_count} payments created, {error_count} errors")
            if errors:
                for error in errors[:10]:
                    messages.warning(request, error)
                if len(errors) > 10:
                    messages.warning(request, f"... and {len(errors) - 10} more errors")
                    
        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
        
        return redirect("payment_list")
    
    accounts = Account.objects.filter(is_active=True, account_type='revenue').order_by('code')
    return render(request, "accounting_app/bulk_payment_import.html", {"accounts": accounts})


@login_required
def bulk_payment_delete(request, bulk_group_id):
    payments = Payment.objects.filter(bulk_group_id=bulk_group_id, is_bulk=True)
    if not payments.exists():
        messages.error(request, "Bulk payments not found")
        return redirect("payment_list")
    
    if request.method == "POST":
        count = payments.count()
        for payment in payments:
            payment.beneficiary.recalculate_totals()
        payments.delete()
        messages.success(request, f"Successfully deleted {count} bulk payments")
        return redirect("payment_list")
    
    return render(request, "accounting_app/bulk_payment_delete.html", {
        "payments": payments,
        "bulk_group_id": bulk_group_id,
    })


@login_required
def notifications_view(request):
    from accounting_app.models import Report
    if request.method == "POST":
        notification_id = request.POST.get('notification_id')
        if notification_id:
            try:
                report = Report.objects.get(pk=notification_id)
                report.delete()
                messages.success(request, "Notification deleted")
            except Report.DoesNotExist:
                messages.error(request, "Notification not found")
        return redirect('notifications')
    
    notifications = Report.objects.all().order_by('-created_at')[:50]
    return render(request, "accounting_app/notifications.html", {"notifications": notifications})


@login_required
def download_payment_template(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    
    headers = ['client_name', 'payment_date', 'amount', 'payment_method', 'reference', 'account', 'notes']
    ws.append(headers)
    
    sample_data = [
        ['John Doe', '2026-01-15', '5000', 'cash', 'REF-001', 'Water Charges', 'January payment'],
        ['Jane Smith', '2026-01-15', '3000', 'bank_transfer', 'REF-002', 'Water Charges', ''],
        ['Village Committee', '2026-01-20', '10000', 'mobile_money', 'REF-003', 'Water Charges', 'Quarterly payment'],
    ]
    
    for row in sample_data:
        ws.append(row)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="payment_import_template.xlsx"'
    wb.save(response)
    return response


@login_required
def journal_list(request):
    entries = JournalEntry.objects.all()
    return render(request, "accounting_app/journal_list.html", {"entries": entries})


@login_required
def journal_create(request):
    if request.method == "POST":
        form = JournalEntryForm(request.POST)
        if form.is_valid():
            entry = form.save(commit=False)
            entry.created_by = request.user
            entry.save()
            ActivityLog.objects.create(
                user=request.user,
                action="Adjustment",
                model_name="JournalEntry",
                object_id=entry.pk,
                description=f"Created journal entry {entry.entry_number}: {entry.description}",
                ip_address=request.META.get('REMOTE_ADDR'),
            )
            messages.success(request, "Journal entry created")
            return redirect("journal_edit", pk=entry.pk)
    else:
        form = JournalEntryForm()
    return render(request, "accounting_app/journal_form.html", {"form": form, "action": "Create"})


@login_required
def journal_edit(request, pk):
    entry = get_object_or_404(JournalEntry, pk=pk)
    if request.method == "POST":
        if "add_line" in request.POST:
            account_id = request.POST.get("account")
            debit = request.POST.get("debit") or 0
            credit = request.POST.get("credit") or 0
            memo = request.POST.get("memo", "")
            JournalEntryLine.objects.create(
                journal_entry=entry, account_id=account_id,
                debit=debit, credit=credit, memo=memo
            )
            return redirect("journal_edit", pk=pk)
        form = JournalEntryForm(request.POST, instance=entry)
        if form.is_valid():
            form.save()
            messages.success(request, "Journal entry updated")
            return redirect("journal_list")
    else:
        form = JournalEntryForm(instance=entry)
    return render(request, "accounting_app/journal_form.html", {"form": form, "entry": entry, "action": "Update"})


@login_required
def journal_delete_line(request, pk, line_pk):
    line = get_object_or_404(JournalEntryLine, pk=line_pk)
    entry_pk = line.journal_entry.pk
    line.delete()
    messages.success(request, "Journal entry line removed successfully")
    return redirect("journal_edit", pk=entry_pk)


@login_required
def reports(request):
    today = timezone.now().date()
    year_start = today.replace(month=1, day=1)
    
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if from_date:
        from_date = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
    else:
        from_date = year_start
    
    if to_date:
        to_date = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
    else:
        to_date = today

    invoices = Invoice.objects.filter(issue_date__gte=from_date, issue_date__lte=to_date)
    payments = Payment.objects.filter(payment_date__gte=from_date, payment_date__lte=to_date)
    expenses = Expense.objects.filter(expense_date__gte=from_date, expense_date__lte=to_date)

    total_invoiced = invoices.aggregate(total=Sum('total_amount'))['total'] or 0
    total_payments = payments.aggregate(total=Sum('amount'))['total'] or 0
    total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or 0
    balance_to_collect = total_invoiced - total_payments
    net_income = total_payments - total_expenses

    beneficiaries_with_balance_count = Beneficiary.objects.filter(is_active=True, total_outstanding__gt=0).count()

    saved_reports = Report.objects.filter(report_type='financial')[:10]
    report_form = ReportForm()

    context = {
        "from_date": from_date,
        "to_date": to_date,
        "total_invoiced": total_invoiced,
        "total_payments": total_payments,
        "total_expenses": total_expenses,
        "balance_to_collect": balance_to_collect,
        "net_income": net_income,
        "net_income_abs": abs(net_income) if net_income < 0 else net_income,
        "saved_reports": saved_reports,
        "report_form": report_form,
        "beneficiaries_with_balance_count": beneficiaries_with_balance_count,
    }
    return render(request, "accounting_app/reports.html", context)


@login_required
def beneficiaries_with_balances(request):
    scheme = request.GET.get('scheme')
    
    beneficiaries = Beneficiary.objects.filter(is_active=True, total_outstanding__gt=0)
    
    if scheme:
        beneficiaries = beneficiaries.filter(scheme=scheme)
    
    beneficiaries_with_balance = []
    grand_total_invoiced = Decimal('0.00')
    grand_total_paid = Decimal('0.00')
    grand_total_balance = Decimal('0.00')
    
    for b in beneficiaries:
        beneficiaries_with_balance.append({
            'id': b.id,
            'name': b.name,
            'scheme': b.scheme,
            'village': b.village,
            'phone': b.phone,
            'total_invoiced': b.total_bill,
            'total_paid': b.total_paid,
            'balance': b.total_outstanding,
        })
        grand_total_invoiced += b.total_bill
        grand_total_paid += b.total_paid
        grand_total_balance += b.total_outstanding
    
    beneficiaries_with_balance.sort(key=lambda x: x['balance'], reverse=True)
    
    schemes = Beneficiary.SCHEME_CHOICES
    
    context = {
        "scheme": scheme,
        "schemes": schemes,
        "beneficiaries_with_balance": beneficiaries_with_balance,
        "grand_total_invoiced": grand_total_invoiced,
        "grand_total_paid": grand_total_paid,
        "grand_total_balance": grand_total_balance,
        "total_count": len(beneficiaries_with_balance),
    }
    return render(request, "accounting_app/beneficiaries_with_balances.html", context)


def get_report_data(from_date, to_date):
    today = timezone.now().date()
    year_start = today.replace(month=1, day=1)
    
    if not from_date:
        from_date = year_start
    if not to_date:
        to_date = today
    
    invoices = Invoice.objects.filter(issue_date__gte=from_date, issue_date__lte=to_date)
    payments = Payment.objects.filter(payment_date__gte=from_date, payment_date__lte=to_date)
    expenses = Expense.objects.filter(expense_date__gte=from_date, expense_date__lte=to_date)
    
    total_invoiced = invoices.aggregate(total=Sum('total_amount'))['total'] or 0
    total_payments = payments.aggregate(total=Sum('amount'))['total'] or 0
    total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or 0
    balance_to_collect = total_invoiced - total_payments
    net_income = total_payments - total_expenses
    
    return {
        "from_date": from_date,
        "to_date": to_date,
        "total_invoiced": total_invoiced,
        "total_payments": total_payments,
        "total_expenses": total_expenses,
        "balance_to_collect": balance_to_collect,
        "net_income": net_income,
        "net_income_abs": abs(net_income) if net_income < 0 else net_income,
    }


@login_required
def save_report(request):
    if request.method == 'POST':
        form = ReportForm(request.POST)
        if form.is_valid():
            from_date = form.cleaned_data['from_date']
            to_date = form.cleaned_data['to_date']
            report_type = form.cleaned_data['report_type']
            
            data = get_report_data(from_date, to_date)
            
            report = Report.objects.create(
                name=form.cleaned_data['name'],
                report_type=report_type,
                from_date=from_date,
                to_date=to_date,
                total_invoiced=data['total_invoiced'],
                total_payments=data['total_payments'],
                total_expenses=data['total_expenses'],
                balance_to_collect=data['balance_to_collect'],
                net_income=data['net_income'],
                created_by=request.user
            )
            messages.success(request, f"Report '{report.name}' saved successfully!")
        else:
            messages.error(request, "Failed to save report. Please check the form.")
    return redirect('reports')


@login_required
def export_report_pdf(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if from_date:
        from_date = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
    if to_date:
        to_date = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
    
    data = get_report_data(from_date, to_date)
    
    user_profile = request.user.userprofile
    currency_symbol = user_profile.get_currency_symbol()
    company_name = user_profile.company_name or "Nkulawua"
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Financial_Report_{data["from_date"]}_{data["to_date"]}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    elements = []
    
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Center', alignment=TA_CENTER))
    styles.add(ParagraphStyle(name='Right', alignment=TA_RIGHT))
    styles.add(ParagraphStyle(name='Title_Custom', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=20, spaceAfter=5, textColor=colors.HexColor('#2c3e50')))
    styles.add(ParagraphStyle(name='SubTitle', parent=styles['Normal'], alignment=TA_CENTER, fontSize=11, textColor=colors.grey))
    styles.add(ParagraphStyle(name='CompanyInfo', parent=styles['Normal'], alignment=TA_CENTER, fontSize=9, textColor=colors.grey))
    
    company_data = []
    if user_profile.logo:
        try:
            logo_path = user_profile.logo.path
            if os.path.exists(logo_path):
                logo = Image(logo_path, width=3*cm, height=2*cm)
                company_data.append([logo])
        except:
            pass
    
    company_data.append([Paragraph(company_name, styles['Title_Custom'])])
    if user_profile.address:
        company_data.append([Paragraph(user_profile.address, styles['CompanyInfo'])])
    if user_profile.phone:
        company_data.append([Paragraph(f"Phone: {user_profile.phone}", styles['CompanyInfo'])])
    if user_profile.tax_id:
        company_data.append([Paragraph(f"Tax ID: {user_profile.tax_id}", styles['CompanyInfo'])])
    
    company_table = Table(company_data, colWidths=[16*cm])
    company_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(company_table)
    elements.append(Spacer(1, 10))
    
    report_info = [
        ["REPORT DETAILS", "", ""],
        ["Report Type:", "Financial Report", ""],
        ["Period:", f"From {data['from_date']} to {data['to_date']}", ""],
        ["Generated:", timezone.now().strftime('%Y-%m-%d %H:%M'), ""],
    ]
    
    info_table = Table(report_info, colWidths=[4*cm, 8*cm, 4*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('SPAN', (0, 0), (-1, 0)),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    summary_data = [
        ["Description", "Amount"],
        ["Total Invoiced", f"{currency_symbol}{data['total_invoiced']:,.2f}"],
        ["Total Revenue (Payments Received)", f"{currency_symbol}{data['total_payments']:,.2f}"],
        ["Total Expenses", f"{currency_symbol}{data['total_expenses']:,.2f}"],
        ["Balance to be Collected", f"{currency_symbol}{data['balance_to_collect']:,.2f}"],
        ["Net Income / (Loss)", f"{currency_symbol}{data['net_income']:,.2f}"],
    ]
    
    table = Table(summary_data, colWidths=[10*cm, 6*cm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f5e9')),
    ]))
    elements.append(table)
    
    elements.append(Spacer(1, 30))
    
    net_color = colors.green if data['net_income'] >= 0 else colors.red
    net_text = f"PROFIT: {currency_symbol}{data['net_income']:,.2f}" if data['net_income'] >= 0 else f"LOSS: {currency_symbol}{abs(data['net_income']):,.2f}"
    
    net_para = Paragraph(net_text, ParagraphStyle(
        name='NetIncome',
        parent=styles['Heading3'],
        alignment=TA_CENTER,
        textColor=net_color
    ))
    elements.append(net_para)
    
    doc.build(elements)
    return response


@login_required
def export_report_excel(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if from_date:
        from_date = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
    if to_date:
        to_date = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
    
    data = get_report_data(from_date, to_date)
    
    user_profile = request.user.userprofile
    currency_symbol = user_profile.get_currency_symbol()
    company_name = user_profile.company_name or "Nkulawua"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"
    
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    blue_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    data_font = Font(size=10)
    bold_font = Font(bold=True, size=10)
    
    ws.merge_cells('A1:D1')
    ws['A1'] = company_name
    ws['A1'].font = Font(bold=True, size=18)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    if user_profile.address:
        ws.merge_cells('A2:D2')
        ws['A2'] = user_profile.address
        ws['A2'].font = Font(size=10, italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        row_start = 3
    else:
        row_start = 2
    
    if user_profile.phone or user_profile.tax_id:
        contact_info = []
        if user_profile.phone:
            contact_info.append(f"Phone: {user_profile.phone}")
        if user_profile.tax_id:
            contact_info.append(f"Tax ID: {user_profile.tax_id}")
        ws.merge_cells(f'A{row_start}:D{row_start}')
        ws[f'A{row_start}'] = " | ".join(contact_info)
        ws[f'A{row_start}'].font = Font(size=9)
        ws[f'A{row_start}'].alignment = Alignment(horizontal='center')
        row_start += 1
    
    ws.append([])
    row_start += 1
    
    ws.merge_cells(f'A{row_start}:D{row_start}')
    ws[f'A{row_start}'] = "REPORT DETAILS"
    ws[f'A{row_start}'].font = header_font
    ws[f'A{row_start}'].fill = blue_fill
    ws[f'A{row_start}'].alignment = Alignment(horizontal='center')
    row_start += 1
    
    ws[f'A{row_start}'] = "Report Type:"
    ws[f'A{row_start}'].font = bold_font
    ws.merge_cells(f'B{row_start}:D{row_start}')
    ws[f'B{row_start}'] = "Financial Report"
    row_start += 1
    
    ws[f'A{row_start}'] = "Period:"
    ws[f'A{row_start}'].font = bold_font
    ws.merge_cells(f'B{row_start}:D{row_start}')
    ws[f'B{row_start}'] = f"From {data['from_date']} to {data['to_date']}"
    row_start += 1
    
    ws[f'A{row_start}'] = "Generated:"
    ws[f'A{row_start}'].font = bold_font
    ws.merge_cells(f'B{row_start}:D{row_start}')
    ws[f'B{row_start}'] = timezone.now().strftime('%Y-%m-%d %H:%M')
    row_start += 2
    
    row_start += 1
    
    ws.append(["Total Invoiced", f"{currency_symbol}{data['total_invoiced']:,.2f}"])
    ws.append(["Total Revenue (Payments)", f"{currency_symbol}{data['total_payments']:,.2f}"])
    ws.append(["Total Expenses", f"{currency_symbol}{data['total_expenses']:,.2f}"])
    ws.append(["Balance to be Collected", f"{currency_symbol}{data['balance_to_collect']:,.2f}"])
    
    net_income_row = ws.max_row + 1
    ws[f'A{net_income_row}'] = "Net Income / (Loss)"
    ws[f'B{net_income_row}'] = f"{currency_symbol}{data['net_income']:,.2f}"
    
    for row in range(1, ws.max_row + 1):
        for cell in ws[row]:
            if cell.font.bold:
                continue
            cell.font = data_font
    
    ws[f'A{net_income_row}'].font = bold_font
    ws[f'B{net_income_row}'].font = bold_font
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Financial_Report_{data["from_date"]}_{data["to_date"]}.xlsx"'
    wb.save(response)
    return response


def share_report(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    method = request.GET.get('method', '')
    phone = request.GET.get('phone', '')
    email = request.GET.get('email', '')
    
    if from_date:
        from_date = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
    if to_date:
        to_date = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
    
    data = get_report_data(from_date, to_date)
    user_profile = request.user.userprofile
    currency_symbol = user_profile.get_currency_symbol()
    company_name = user_profile.company_name or "Nkulawua"
    
    report_message = f"""
{company_name}
Financial Report
Period: {from_date} to {to_date}

SUMMARY:
---------
Total Invoiced: {currency_symbol}{data['total_invoiced']:,.2f}
Total Revenue: {currency_symbol}{data['total_payments']:,.2f}
Total Expenses: {currency_symbol}{data['total_expenses']:,.2f}
Balance to be Collected: {currency_symbol}{data['balance_to_collect']:,.2f}
Net Income: {currency_symbol}{data['net_income']:,.2f}

Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}
"""
    
    if method == 'whatsapp':
        from urllib.parse import quote
        phone = phone.strip().replace(' ', '').replace('+', '')
        if not phone.startswith('265') and not phone.startswith('0'):
            phone = '265' + phone.lstrip('0')
        encoded_message = quote(report_message)
        whatsapp_url = f"https://wa.me/{phone}?text={encoded_message}"
        return HttpResponseRedirect(whatsapp_url)
    
    elif method == 'sms':
        profile = user_profile
        if profile.sms_provider == "none":
            messages.info(request, f"SMS to {phone}: {report_message[:160]}... (Demo Mode)")
        else:
            from .sms_service import send_sms_to_client
            result = send_sms_to_client(profile, phone, report_message[:160])
            if result["status"] == "success":
                messages.success(request, f"SMS sent successfully to {phone}")
            else:
                messages.error(request, f"Failed to send SMS: {result['message']}")
        return redirect("reports")
    
    return redirect("reports")


def send_report_email(request):
    if request.method == "POST":
        recipient_email = request.POST.get('email')
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')
        
        if not recipient_email:
            messages.error(request, "Please enter an email address")
            return redirect("reports")
        
        if from_date:
            from_date = timezone.datetime.strptime(from_date, '%Y-%m-%d').date()
        if to_date:
            to_date = timezone.datetime.strptime(to_date, '%Y-%m-%d').date()
        
        data = get_report_data(from_date, to_date)
        user_profile = request.user.userprofile
        currency_symbol = user_profile.get_currency_symbol()
        company_name = user_profile.company_name or "Nkulawua"
        
        subject = f"{company_name} - Financial Report ({from_date} to {to_date})"
        
        message = f"""
Dear Recipient,

Please find below the financial report for {company_name}.

REPORT PERIOD: {from_date} to {to_date}

FINANCIAL SUMMARY:
==================
Total Invoiced:      {currency_symbol}{data['total_invoiced']:>15,.2f}
Total Revenue:       {currency_symbol}{data['total_payments']:>15,.2f}
Total Expenses:      {currency_symbol}{data['total_expenses']:>15,.2f}
Balance to Collect:  {currency_symbol}{data['balance_to_collect']:>15,.2f}
Net Income:         {currency_symbol}{data['net_income']:>15,.2f}

{"=" * 50}

{"PROFIT" if data['net_income'] >= 0 else "LOSS"} for the period: {currency_symbol}{abs(data['net_income']):,.2f}

{"=" * 50}

This report was automatically generated on {timezone.now().strftime('%Y-%m-%d %H:%M')}.

Best regards,
{company_name}
"""
        
        try:
            send_mail(
                subject,
                message,
                user_profile.company_name + " <noreply@" + request.get_host() + ">",
                [recipient_email],
                fail_silently=False,
            )
            messages.success(request, f"Report sent successfully to {recipient_email}")
        except Exception as e:
            messages.error(request, f"Failed to send email: {str(e)}")
        
        return redirect("reports")
    
    return redirect("reports")


@login_required
def budget_list(request):
    budgets = Budget.objects.prefetch_related("lines").all()
    return render(request, "accounting_app/budget_list.html", {"budgets": budgets})


@login_required
def budget_create(request):
    if request.method == "POST":
        form = BudgetForm(request.POST)
        formset = BudgetLineFormSet(request.POST)
        print("POST keys:", request.POST.keys())
        print("Form valid:", form.is_valid())
        print("Form errors:", form.errors)
        print("Formset valid:", formset.is_valid())
        print("Formset errors:", formset.errors)
        print("Formset non_form_errors:", formset.non_form_errors())
        if form.is_valid() and formset.is_valid():
            budget = form.save()
            formset.instance = budget
            formset.save()
            messages.success(request, "Budget created successfully")
            return redirect("budget_list")
        elif request.method == "POST":
            messages.error(request, "Please fix the errors below.")
    else:
        form = BudgetForm()
        formset = BudgetLineFormSet()
    return render(request, "accounting_app/budget_form.html", {"form": form, "formset": formset, "action": "Create"})


@login_required
def budget_edit(request, pk):
    budget = get_object_or_404(Budget, pk=pk)
    if request.method == "POST":
        form = BudgetForm(request.POST, instance=budget)
        formset = BudgetLineFormSet(request.POST, instance=budget)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, "Budget updated successfully")
            return redirect("budget_list")
    else:
        form = BudgetForm(instance=budget)
        formset = BudgetLineFormSet(instance=budget)
    return render(request, "accounting_app/budget_form.html", {"form": form, "formset": formset, "action": "Update", "budget": budget})


@login_required
def budget_delete(request, pk):
    budget = get_object_or_404(Budget, pk=pk)
    if request.method == "POST":
        save_deleted_record(budget, request.user)
        budget.delete()
        messages.success(request, "Budget deleted successfully")
        return redirect("budget_list")
    return render(request, "accounting_app/budget_confirm_delete.html", {"budget": budget})


def register(request):
    if request.method == "POST":
        form = UserForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data["password"])
            user.save()
            messages.success(request, "Account created successfully")
            return redirect("login")
    else:
        form = UserForm()
    return render(request, "accounting_app/register.html", {"form": form})


@login_required
def settings(request):
    if not hasattr(request.user, 'userprofile'):
        user_profile, _ = UserProfile.objects.get_or_create(user=request.user)
    profile = request.user.userprofile
    if request.method == "POST":
        form = UserProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "Settings saved successfully")
            return redirect("settings")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = UserProfileForm(instance=profile)
    return render(request, "accounting_app/settings.html", {"form": form})


@login_required
def theme_settings(request):
    if not hasattr(request.user, 'userprofile'):
        user_profile, _ = UserProfile.objects.get_or_create(user=request.user)
    profile = request.user.userprofile
    if request.method == "POST":
        card_color = request.POST.get('card_color', 'primary')
        text_color = request.POST.get('text_color', 'dark')
        header_color = request.POST.get('header_color', 'primary')
        sidebar_color = request.POST.get('sidebar_color', 'dark')
        ticker_color = request.POST.get('ticker_color', 'dark')
        title_animation = request.POST.get('title_animation', 'gradient')
        theme = request.POST.get('theme', 'light')
        
        profile.accent_color = card_color
        profile.sidebar_color = sidebar_color
        profile.ticker_color = ticker_color
        profile.title_animation = title_animation
        profile.theme = theme
        profile.save()
        
        messages.success(request, "Theme settings saved successfully!")
        return redirect("theme_settings")
    
    form = UserProfileForm(instance=profile)
    return render(request, "accounting_app/theme_settings.html", {"form": form})


@login_required
def landing_page_settings(request):
    if not (request.user.userprofile.role == "admin" or request.user.userprofile.has_settings_access):
        messages.error(request, "You don't have permission to access landing page settings.")
        return redirect('dashboard')
    
    settings_obj = LandingPageSettings.objects.filter(is_active=True).first()
    if not settings_obj:
        settings_obj = LandingPageSettings.objects.create()
    
    if request.method == "POST":
        if request.POST.get('action') == 'update_settings':
            settings_obj.site_name = request.POST.get('site_name', settings_obj.site_name)
            settings_obj.hero_title = request.POST.get('hero_title', settings_obj.hero_title)
            settings_obj.hero_subtitle = request.POST.get('hero_subtitle', settings_obj.hero_subtitle)
            settings_obj.hero_description = request.POST.get('hero_description', settings_obj.hero_description)
            settings_obj.about_title = request.POST.get('about_title', settings_obj.about_title)
            settings_obj.about_subtitle = request.POST.get('about_subtitle', settings_obj.about_subtitle)
            settings_obj.about_content = request.POST.get('about_content', settings_obj.about_content)
            settings_obj.vision_title = request.POST.get('vision_title', settings_obj.vision_title)
            settings_obj.vision_subtitle = request.POST.get('vision_subtitle', settings_obj.vision_subtitle)
            settings_obj.vision_text = request.POST.get('vision_text', settings_obj.vision_text)
            settings_obj.mission_text = request.POST.get('mission_text', settings_obj.mission_text)
            settings_obj.values_text = request.POST.get('values_text', settings_obj.values_text)
            settings_obj.location_title = request.POST.get('location_title', settings_obj.location_title)
            settings_obj.location_subtitle = request.POST.get('location_subtitle', settings_obj.location_subtitle)
            settings_obj.location_address = request.POST.get('location_address', settings_obj.location_address)
            settings_obj.location_description = request.POST.get('location_description', settings_obj.location_description)
            settings_obj.contact_phone = request.POST.get('contact_phone', settings_obj.contact_phone)
            settings_obj.contact_email = request.POST.get('contact_email', settings_obj.contact_email)
            settings_obj.contact_whatsapp = request.POST.get('contact_whatsapp', settings_obj.contact_whatsapp)
            settings_obj.contact_whatsapp_raw = request.POST.get('contact_whatsapp_raw', settings_obj.contact_whatsapp_raw)
            settings_obj.projects_title = request.POST.get('projects_title', settings_obj.projects_title)
            settings_obj.projects_subtitle = request.POST.get('projects_subtitle', settings_obj.projects_subtitle)
            settings_obj.projects_intro = request.POST.get('projects_intro', settings_obj.projects_intro)
            settings_obj.projects_items = request.POST.get('projects_items', settings_obj.projects_items)
            settings_obj.projects_outro = request.POST.get('projects_outro', settings_obj.projects_outro)
            settings_obj.cta_title = request.POST.get('cta_title', settings_obj.cta_title)
            settings_obj.cta_description = request.POST.get('cta_description', settings_obj.cta_description)
            settings_obj.footer_text = request.POST.get('footer_text', settings_obj.footer_text)
            settings_obj.primary_color = request.POST.get('primary_color', settings_obj.primary_color)
            settings_obj.secondary_color = request.POST.get('secondary_color', settings_obj.secondary_color)
            settings_obj.hero_gradient_start = request.POST.get('hero_gradient_start', settings_obj.hero_gradient_start)
            settings_obj.hero_gradient_mid = request.POST.get('hero_gradient_mid', settings_obj.hero_gradient_mid)
            settings_obj.hero_gradient_end = request.POST.get('hero_gradient_end', settings_obj.hero_gradient_end)
            settings_obj.section_bg_light = request.POST.get('section_bg_light', settings_obj.section_bg_light)
            settings_obj.text_primary = request.POST.get('text_primary', settings_obj.text_primary)
            settings_obj.text_secondary = request.POST.get('text_secondary', settings_obj.text_secondary)
            settings_obj.cta_gradient_start = request.POST.get('cta_gradient_start', settings_obj.cta_gradient_start)
            settings_obj.cta_gradient_end = request.POST.get('cta_gradient_end', settings_obj.cta_gradient_end)
            settings_obj.whatsapp_color = request.POST.get('whatsapp_color', settings_obj.whatsapp_color)
            settings_obj.water_sources_title = request.POST.get('water_sources_title', settings_obj.water_sources_title)
            settings_obj.water_sources_subtitle = request.POST.get('water_sources_subtitle', settings_obj.water_sources_subtitle)
            settings_obj.water_sources_images = request.POST.get('water_sources_images', settings_obj.water_sources_images)
            settings_obj.news_title = request.POST.get('news_title', settings_obj.news_title)
            settings_obj.news_subtitle = request.POST.get('news_subtitle', settings_obj.news_subtitle)
            settings_obj.news_content = request.POST.get('news_content', settings_obj.news_content)
            settings_obj.news_intro = request.POST.get('news_intro', settings_obj.news_intro)
            settings_obj.meeting_objectives_title = request.POST.get('meeting_objectives_title', settings_obj.meeting_objectives_title)
            settings_obj.meeting_objectives_subtitle = request.POST.get('meeting_objectives_subtitle', settings_obj.meeting_objectives_subtitle)
            settings_obj.meeting_objectives_intro = request.POST.get('meeting_objectives_intro', settings_obj.meeting_objectives_intro)
            settings_obj.meeting_objectives_items = request.POST.get('meeting_objectives_items', settings_obj.meeting_objectives_items)
            settings_obj.schemes_title = request.POST.get('schemes_title', settings_obj.schemes_title)
            settings_obj.schemes_subtitle = request.POST.get('schemes_subtitle', settings_obj.schemes_subtitle)
            settings_obj.schemes_list = request.POST.get('schemes_list', settings_obj.schemes_list)
            settings_obj.villages_title = request.POST.get('villages_title', settings_obj.villages_title)
            settings_obj.villages_subtitle = request.POST.get('villages_subtitle', settings_obj.villages_subtitle)
            settings_obj.villages_list = request.POST.get('villages_list', settings_obj.villages_list)
            settings_obj.gallery_title = request.POST.get('gallery_title', settings_obj.gallery_title)
            settings_obj.gallery_subtitle = request.POST.get('gallery_subtitle', settings_obj.gallery_subtitle)
            settings_obj.save()
            messages.success(request, "Landing page settings saved successfully!")
        return redirect("landing_page_settings")
    
    gallery_images = GalleryImage.objects.all().order_by('-uploaded_at')
    services = Service.objects.all().order_by('order')
    
    return render(request, "accounting_app/landing_page_settings.html", {
        "settings": settings_obj,
        "gallery_images": gallery_images,
        "services": services
    })


@login_required
def gallery_image_create(request):
    if not (request.user.userprofile.role == "admin" or request.user.userprofile.has_settings_access):
        messages.error(request, "You don't have permission.")
        return redirect('landing_page_settings')
    
    if request.method == "POST":
        title = request.POST.get("title", "")
        description = request.POST.get("description", "")
        image = request.FILES.get("image")
        
        if image:
            GalleryImage.objects.create(
                title=title,
                description=description,
                image=image
            )
            messages.success(request, "Gallery image added successfully!")
        else:
            messages.error(request, "Image file is required.")
    
    return redirect("landing_page_settings")


@login_required
def gallery_image_edit(request, pk):
    if not (request.user.userprofile.role == "admin" or request.user.userprofile.has_settings_access):
        messages.error(request, "You don't have permission.")
        return redirect('landing_page_settings')
    
    image = get_object_or_404(GalleryImage, pk=pk)
    
    if request.method == "POST":
        image.title = request.POST.get("title", "")
        image.description = request.POST.get("description", "")
        image.is_active = request.POST.get("is_active") == "on"
        new_image = request.FILES.get("image")
        if new_image:
            image.image = new_image
        image.save()
        messages.success(request, "Gallery image updated successfully!")
    
    return redirect("landing_page_settings")


@login_required
def gallery_image_delete(request, pk):
    if not (request.user.userprofile.role == "admin" or request.user.userprofile.has_settings_access):
        messages.error(request, "You don't have permission.")
        return redirect('landing_page_settings')
    
    image = get_object_or_404(GalleryImage, pk=pk)
    save_deleted_record(image, request.user)
    image.delete()
    messages.success(request, "Gallery image deleted successfully!")
    return redirect("landing_page_settings")


@login_required
def service_create(request):
    if not (request.user.userprofile.role == "admin" or request.user.userprofile.has_settings_access):
        messages.error(request, "You don't have permission.")
        return redirect('landing_page_settings')
    
    if request.method == "POST":
        title = request.POST.get("title", "")
        description = request.POST.get("description", "")
        caption = request.POST.get("caption", "")
        image = request.FILES.get("image")
        order = request.POST.get("order", 0)
        
        if image and title:
            Service.objects.create(
                title=title,
                description=description,
                caption=caption,
                image=image,
                order=int(order)
            )
            messages.success(request, "Service added successfully!")
        else:
            messages.error(request, "Title and image are required.")
    
    return redirect("landing_page_settings")


@login_required
def service_edit(request, pk):
    if not (request.user.userprofile.role == "admin" or request.user.userprofile.has_settings_access):
        messages.error(request, "You don't have permission.")
        return redirect('landing_page_settings')
    
    service = get_object_or_404(Service, pk=pk)
    
    if request.method == "POST":
        service.title = request.POST.get("title", service.title)
        service.description = request.POST.get("description", service.description)
        service.caption = request.POST.get("caption", service.caption)
        service.order = int(request.POST.get("order", service.order))
        service.is_active = request.POST.get("is_active") == "on"
        new_image = request.FILES.get("image")
        if new_image:
            service.image = new_image
        service.save()
        messages.success(request, "Service updated successfully!")
    
    return redirect("landing_page_settings")


@login_required
def service_delete(request, pk):
    if not (request.user.userprofile.role == "admin" or request.user.userprofile.has_settings_access):
        messages.error(request, "You don't have permission.")
        return redirect('landing_page_settings')
    
    service = get_object_or_404(Service, pk=pk)
    save_deleted_record(service, request.user)
    service.delete()
    messages.success(request, "Service deleted successfully!")
    return redirect("landing_page_settings")


@login_required
def send_sms(request, beneficiary_id):
    beneficiary = get_object_or_404(Beneficiary, pk=beneficiary_id)
    
    if request.method == "POST":
        phone = request.POST.get("phone", beneficiary.phone)
        message = request.POST.get("message", "")
        
        if not phone:
            messages.error(request, "Phone number is required")
            return redirect("beneficiary_detail", pk=beneficiary_id)
        
        if not message:
            messages.error(request, "Message cannot be empty")
            return redirect("beneficiary_detail", pk=beneficiary_id)
        
        profile = request.user.userprofile
        
        if profile.sms_provider == "none":
            messages.success(request, f"SMS to {phone}: {message} (Demo Mode)")
            return redirect("beneficiary_detail", pk=beneficiary_id)
        
        from .sms_service import send_sms_to_client
        result = send_sms_to_client(profile, phone, message)
        
        if result["status"] == "success":
            messages.success(request, f"SMS sent successfully to {phone}")
        else:
            messages.error(request, f"Failed to send SMS: {result['message']}")
        
        return redirect("beneficiary_detail", pk=beneficiary_id)
    
    return redirect("beneficiary_detail", pk=beneficiary_id)


@login_required
def user_list(request):
    if not hasattr(request.user, 'userprofile'):
        user_profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not (request.user.userprofile.role == "admin" or request.user.userprofile.can_manage_users):
        messages.error(request, "You don't have permission to view users.")
        return redirect('dashboard')
    
    users = User.objects.all().select_related('userprofile')
    return render(request, "accounting_app/user_list.html", {"users": users})


@login_required
def login_sessions(request):
    if not hasattr(request.user, 'userprofile'):
        user_profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not (request.user.userprofile.role == "admin" or request.user.userprofile.can_manage_users):
        messages.error(request, "You don't have permission to view login sessions.")
        return redirect('dashboard')
    
    sessions = LoginSession.objects.select_related('user').all()
    active_sessions = sessions.filter(is_active=True)
    
    return render(request, "accounting_app/login_sessions.html", {
        "sessions": sessions,
        "active_sessions": active_sessions
    })


@login_required
def data_recovery(request):
    if not hasattr(request.user, 'userprofile'):
        user_profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not (request.user.userprofile.role == "admin" or request.user.userprofile.can_manage_users):
        messages.error(request, "You don't have permission to access data recovery.")
        return redirect('dashboard')
    
    deleted_records = DeletedRecord.objects.all()
    recovered_records = deleted_records.filter(recovered=True)
    pending_records = deleted_records.filter(recovered=False)

    import io, zipfile, json, tempfile, os
    from django.core.files.base import ContentFile
    from decimal import Decimal

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "export_all":
            try:
                buffer = io.BytesIO()
                with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                    for name, model_class in EXPORT_MODELS.items():
                        data = _serialize_model(model_class)
                        zf.writestr(f"{name}.json", json.dumps(data, default=str, indent=2))
                    zf.writestr("export_info.json", json.dumps({
                        "exported_at": timezone.now().isoformat(),
                        "exported_by": request.user.username,
                        "description": "Full database backup from Data Recovery",
                        "model_count": len(EXPORT_MODELS),
                    }, indent=2))

                response = HttpResponse(buffer.getvalue(), content_type="application/zip")
                response["Content-Disposition"] = f'attachment; filename="full_backup_{timezone.now().strftime("%Y%m%d_%H%M%S")}.zip"'
                return response

            except Exception as e:
                messages.error(request, f"Export failed: {e}")

        elif action == "import_preview":
            uploaded_file = request.FILES.get("file")
            mode = request.POST.get("mode", "add_new")

            if not uploaded_file or not uploaded_file.name.endswith(".zip"):
                messages.error(request, "Please upload a valid ZIP backup file.")
                return redirect("data_recovery")

            tmp_path = tempfile.mktemp(suffix=".zip")
            with open(tmp_path, "wb") as f:
                for chunk in uploaded_file.chunks():
                    f.write(chunk)

            try:
                preview_data = {}
                with zipfile.ZipFile(tmp_path, "r") as zf:
                    for name in zf.namelist():
                        if name.endswith(".json") and name != "export_info.json":
                            model_key = name.replace(".json", "")
                            content = json.loads(zf.read(name))
                            preview_data[model_key] = {
                                "total": len(content),
                                "sample": content[:2] if content else [],
                            }

                request.session["import_data"] = {
                    "mode": mode,
                    "file_path": tmp_path,
                }

                return render(request, "accounting_app/data_recovery.html", {
                    "deleted_records": deleted_records,
                    "recovered_records": recovered_records,
                    "pending_records": pending_records,
                    "pending_count": pending_records.count(),
                    "recovered_count": recovered_records.count(),
                    "preview_data": preview_data,
                    "import_mode": mode,
                    "show_preview": True,
                })

            except Exception as e:
                try: os.unlink(tmp_path)
                except: pass
                messages.error(request, f"Preview failed: {e}")

        elif action == "execute_import":
            import_data = request.session.get("import_data")
            if not import_data:
                messages.error(request, "Import session expired. Please upload the file again.")
                return redirect("data_recovery")

            mode = import_data["mode"]
            file_path = import_data["file_path"]

            if not os.path.exists(file_path):
                messages.error(request, "Backup file not found. Please upload again.")
                return redirect("data_recovery")

            from django.db import transaction

            try:
                summary = {}
                with transaction.atomic():
                    with zipfile.ZipFile(file_path, "r") as zf:
                        for name in zf.namelist():
                            if name.endswith(".json") and name != "export_info.json":
                                model_key = name.replace(".json", "")
                                model_class = EXPORT_MODELS.get(model_key)
                                if not model_class:
                                    continue

                                records = json.loads(zf.read(name))
                                created = 0
                                updated = 0
                                skipped = 0
                                pk_field = model_class._meta.pk.name

                                if mode == "replace":
                                    model_class.objects.all().delete()
                                    for record in records:
                                        record.pop(pk_field, None)
                                        try:
                                            obj = model_class(**record)
                                            obj.save()
                                            created += 1
                                        except Exception:
                                            skipped += 1

                                elif mode == "merge":
                                    for record in records:
                                        existing = _identify_record(model_class, record)
                                        if existing:
                                            changed = False
                                            for field in model_class._meta.fields:
                                                if field.name in record and field.name != pk_field:
                                                    setattr(existing, field.name, record[field.name])
                                                    changed = True
                                            if changed:
                                                existing.save()
                                                updated += 1
                                            else:
                                                skipped += 1
                                        else:
                                            record.pop(pk_field, None)
                                            try:
                                                obj = model_class(**record)
                                                obj.save()
                                                created += 1
                                            except Exception:
                                                skipped += 1

                                else:
                                    for record in records:
                                        existing = _identify_record(model_class, record)
                                        if existing:
                                            skipped += 1
                                        else:
                                            record.pop(pk_field, None)
                                            try:
                                                obj = model_class(**record)
                                                obj.save()
                                                created += 1
                                            except Exception:
                                                skipped += 1

                                summary[model_key] = {
                                    "total": len(records),
                                    "created": created,
                                    "updated": updated,
                                    "skipped": skipped,
                                }

                if "import_data" in request.session:
                    del request.session["import_data"]

                try: os.unlink(file_path)
                except: pass

                total_created = sum(s["created"] for s in summary.values())
                total_updated = sum(s["updated"] for s in summary.values())
                messages.success(request, f"Import completed. {total_created} records created, {total_updated} updated.")

                ActivityLog.objects.create(
                    user=request.user,
                    action="Import",
                    model_name="DataRecovery",
                    description=f"Database {dict(DataMigrationLog.IMPORT_MODES).get(mode, mode)} import completed: {total_created} created, {total_updated} updated",
                    ip_address=request.META.get('REMOTE_ADDR'),
                )

            except Exception as e:
                messages.error(request, f"Import failed: {e}")

            return redirect("data_recovery")

        record_id = request.POST.get("record_id")
        if record_id and action == "recover":
            record = get_object_or_404(DeletedRecord, pk=record_id)
            if not record.recovered:
                obj = record.recover()
                if obj:
                    messages.success(request, f"Successfully recovered {record.model_name} #{record.object_id}")
                else:
                    messages.error(request, f"Failed to recover {record.model_name} #{record.object_id}. The data may be corrupted or the model no longer exists.")
            else:
                messages.warning(request, "This record has already been recovered.")
        elif record_id and action == "permanent_delete":
            record = get_object_or_404(DeletedRecord, pk=record_id)
            record.delete()
            messages.success(request, "Record permanently deleted.")
    
    return render(request, "accounting_app/data_recovery.html", {
        "deleted_records": deleted_records,
        "recovered_records": recovered_records,
        "pending_records": pending_records,
        "pending_count": pending_records.count(),
        "recovered_count": recovered_records.count(),
    })


@login_required
def user_toggle_status(request, pk):
    if not hasattr(request.user, 'userprofile'):
        user_profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not (request.user.userprofile.role == "admin" or request.user.userprofile.can_manage_users):
        messages.error(request, "You don't have permission to manage users.")
        return redirect('dashboard')
    
    user = get_object_or_404(User, pk=pk)
    
    if request.user.pk == user.pk:
        messages.error(request, "You cannot block/unblock your own account!")
        return redirect("user_list")
    
    if user.is_active:
        user.is_active = False
        user.save()
        messages.success(request, f"User '{user.username}' has been blocked successfully.")
    else:
        user.is_active = True
        user.save()
        messages.success(request, f"User '{user.username}' has been unblocked successfully.")
    
    return redirect("user_list")


@login_required
def user_create(request):
    if not hasattr(request.user, 'userprofile'):
        user_profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not (request.user.userprofile.role == "admin" or request.user.userprofile.can_manage_users):
        messages.error(request, "You don't have permission to create users.")
        return redirect('dashboard')
    
    if request.method == "POST":
        form = UserForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data["username"]
            email = form.cleaned_data["email"]
            password = form.cleaned_data["password"]
            role = request.POST.get("role", "viewer")
            
            perm_fields = [
                'can_access_dashboard', 'can_view_beneficiaries', 'can_edit_beneficiaries',
                'can_delete_beneficiaries', 'can_view_invoices', 'can_edit_invoices',
                'can_delete_invoices', 'can_view_payments', 'can_edit_payments',
                'can_delete_payments', 'can_view_expenses', 'can_edit_expenses',
                'can_delete_expenses', 'can_view_reports', 'can_export_reports',
                'can_view_journal', 'can_edit_journal', 'can_view_budgets',
                'can_edit_budgets', 'can_delete_budgets', 'can_access_management',
                'can_view_employees', 'can_edit_employees', 'can_communicate',
                'can_manage_users', 'can_access_settings'
            ]
            
            permissions = {}
            for field in perm_fields:
                permissions[field] = request.POST.get(field) == 'on'
            
            try:
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password
                )
                
                UserProfile.objects.create(
                    user=user,
                    role=role,
                    company_name=request.user.userprofile.company_name or "My Company",
                    **permissions
                )
                
                try:
                    send_mail(
                        subject=f"Account Created - {request.user.userprofile.company_name or 'Accounting System'}",
                        message=f"""
Dear {username},

Your account has been created successfully.

Login Details:
Username: {username}
Password: {password}
Role: {role.title()}

Please login at: {request.build_absolute_uri('/login/')}

Best regards,
{request.user.userprofile.company_name or 'Accounting System'}
                        """,
                        from_email=request.user.userprofile.company_name + " <noreply@" + request.get_host() + ">",
                        recipient_list=[email],
                        fail_silently=True,
                    )
                    messages.success(request, f"User '{username}' created successfully. Login credentials sent to {email}")
                except:
                    messages.success(request, f"User '{username}' created successfully")
                
                return redirect("user_list")
            except Exception as e:
                messages.error(request, f"Error creating user: {str(e)}")
                return redirect("user_create")
    else:
        form = UserForm()
    return render(request, "accounting_app/user_form.html", {"form": form, "action": "Create"})


@login_required
def user_edit(request, pk):
    if not hasattr(request.user, 'userprofile'):
        user_profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not (request.user.userprofile.role == "admin" or request.user.userprofile.can_manage_users):
        messages.error(request, "You don't have permission to edit users.")
        return redirect('dashboard')
    
    user = get_object_or_404(User, pk=pk)
    profile = user.userprofile
    
    if request.method == "POST":
        user.username = request.POST.get("username")
        user.email = request.POST.get("email")
        user.is_active = request.POST.get("is_active") == "on"
        new_password = request.POST.get("password")
        
        if new_password:
            user.set_password(new_password)
        
        user.save()
        profile.role = request.POST.get("role", "viewer")
        
        perm_fields = [
            'can_access_dashboard', 'can_view_beneficiaries', 'can_edit_beneficiaries',
            'can_delete_beneficiaries', 'can_view_invoices', 'can_edit_invoices',
            'can_delete_invoices', 'can_view_payments', 'can_edit_payments',
            'can_delete_payments', 'can_view_expenses', 'can_edit_expenses',
            'can_delete_expenses', 'can_view_reports', 'can_export_reports',
            'can_view_journal', 'can_edit_journal', 'can_view_budgets',
            'can_edit_budgets', 'can_delete_budgets', 'can_access_management',
            'can_view_employees', 'can_edit_employees', 'can_communicate',
            'can_manage_users', 'can_access_settings'
        ]
        
        for field in perm_fields:
            setattr(profile, field, request.POST.get(field) == 'on')
        
        profile.save()
        
        if new_password and user.email:
            try:
                send_mail(
                    subject=f"Account Updated - {request.user.userprofile.company_name or 'Accounting System'}",
                    message=f"""
Dear {user.username},

Your account has been updated.

Updated Login Details:
Username: {user.username}
New Password: {new_password}
Role: {profile.role.title()}

Please login at: {request.build_absolute_uri('/login/')}

Best regards,
{request.user.userprofile.company_name or 'Accounting System'}
                    """,
                    from_email=request.user.userprofile.company_name + " <noreply@" + request.get_host() + ">",
                    recipient_list=[user.email],
                    fail_silently=True,
                )
                messages.success(request, f"User '{user.username}' updated successfully. New credentials sent to {user.email}")
            except:
                messages.success(request, f"User '{user.username}' updated successfully")
        else:
            messages.success(request, f"User '{user.username}' updated successfully")
        
        return redirect("user_list")
    
    return render(request, "accounting_app/user_edit.html", {"edit_user": user, "profile": profile})


@login_required
def user_delete(request, pk):
    if not hasattr(request.user, 'userprofile'):
        user_profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not (request.user.userprofile.role == "admin" or request.user.userprofile.can_manage_users):
        messages.error(request, "You don't have permission to delete users.")
        return redirect('dashboard')
    
    user = get_object_or_404(User, pk=pk)
    if request.user.pk == user.pk:
        messages.error(request, "You cannot delete your own account!")
        return redirect("user_list")
    
    if request.method == "POST":
        username = user.username
        save_deleted_record(user, request.user)
        user.delete()
        messages.success(request, f"User '{username}' deleted successfully")
        return redirect("user_list")
    
    return render(request, "accounting_app/user_delete.html", {"delete_user": user})


@login_required
def user_reset_password(request, pk):
    if not hasattr(request.user, 'userprofile'):
        user_profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not (request.user.userprofile.role == "admin" or request.user.userprofile.can_manage_users):
        messages.error(request, "You don't have permission to reset user passwords.")
        return redirect('dashboard')
    
    user = get_object_or_404(User, pk=pk)
    if request.user.pk == user.pk:
        messages.error(request, "You cannot reset your own password from here. Use profile settings.")
        return redirect("user_list")
    
    if request.method == "POST":
        new_password = request.POST.get("new_password", "")
        if not new_password or len(new_password) < 4:
            messages.error(request, "Password must be at least 4 characters long.")
            return redirect("user_list")
        
        user.set_password(new_password)
        user.save()
        
        messages.success(request, f"Password reset for user '{user.username}'. New password set.")
        
        if user.email:
            try:
                send_mail(
                    subject=f"Password Reset - {request.user.userprofile.company_name or 'Accounting System'}",
                    message=f"""
Dear {user.username},

Your password has been reset by an administrator.

Your new login credentials are:
Username: {user.username}
New Password: {new_password}

Please login at: {request.build_absolute_uri('/login/')}
and change your password immediately after logging in.

Best regards,
{request.user.userprofile.company_name or 'Accounting System'}
                    """,
                    from_email=request.user.userprofile.company_name + " <noreply@" + request.get_host() + ">",
                    recipient_list=[user.email],
                    fail_silently=True,
                )
            except:
                pass
        
        return redirect("user_list")
    
    return render(request, "accounting_app/user_reset_password.html", {"target_user": user})


@login_required
def user_profile(request):
    if not hasattr(request.user, 'userprofile'):
        user_profile, _ = UserProfile.objects.get_or_create(user=request.user)
    return render(request, "accounting_app/user_profile.html")


@login_required
def import_excel(request):
    if request.method == "POST" and request.FILES.get('excel_file'):
        model_type = request.POST.get('model_type', 'client')
        excel_file = request.FILES['excel_file']
        
        try:
            from openpyxl import load_workbook
            from datetime import datetime
            
            wb = load_workbook(excel_file)
            ws = wb.active
            
            raw_headers = [cell.value for cell in ws[1]]
            headers = [str(h).strip().lower() if h else '' for h in raw_headers]
            
            created_count = 0
            updated_count = 0
            error_count = 0
            errors = []
            total_rows = 0
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                total_rows += 1
                if not any(cell for cell in row):
                    continue
                
                row_data = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        row_data[header] = row[i]
                
                for raw_header, value in zip(raw_headers, row):
                    if raw_header:
                        rh = str(raw_header).strip().lower()
                        if rh not in row_data and value is not None:
                            row_data[rh] = value
                
                try:
                    if model_type == 'beneficiary':
                        result = import_client_data(row_data, request.user)
                    elif model_type == 'vendor':
                        result = import_vendor_data(row_data)
                    elif model_type == 'account':
                        result = import_account_data(row_data)
                    elif model_type == 'invoice':
                        result = import_invoice_data(row_data, request.user)
                    elif model_type == 'expense':
                        result = import_expense_data(row_data, request.user)
                    elif model_type == 'payment':
                        result = import_payment_data(row_data, request.user)
                    elif model_type == 'budget':
                        result = import_budget_data(row_data)
                    else:
                        errors.append(f"Row {row_idx}: Unknown model type")
                        error_count += 1
                        continue
                    
                    if result and result.get('status') == 'created':
                        created_count += 1
                    elif result and result.get('status') == 'updated':
                        updated_count += 1
                    else:
                        error_msg = result.get('message', 'Unknown error') if result else 'No result returned'
                        errors.append(f"Row {row_idx}: {error_msg}")
                        error_count += 1
                        
                except Exception as e:
                    import traceback
                    errors.append(f"Row {row_idx}: {str(e)}")
                    error_count += 1
            
            wb.close()
            
            if total_rows == 0:
                messages.warning(request, "No data found in the Excel file. Make sure the file has headers and at least one data row.")
            else:
                messages.success(request, f"Import completed: {created_count} created, {updated_count} updated, {error_count} errors out of {total_rows} rows")
            
            if errors:
                for error in errors[:20]:
                    messages.warning(request, error)
                if len(errors) > 20:
                    messages.warning(request, f"... and {len(errors) - 20} more errors")
            
        except Exception as e:
            messages.error(request, f"Error processing Excel file: {str(e)}")
        
        return redirect('import_excel')
    
    return render(request, "accounting_app/import_excel.html")


def import_client_data(row_data, user):
    name = row_data.get('name', '')
    if not name:
        return {'status': 'error', 'message': 'Name is required'}
    
    name = str(name).strip()
    beneficiary_type = str(row_data.get('beneficiary_type', 'private')).strip() or 'private'
    email = str(row_data.get('email', '')).strip()
    phone = str(row_data.get('phone', '')).strip()
    village = str(row_data.get('village', '')).strip()
    scheme = str(row_data.get('scheme', '')).strip()
    country = str(row_data.get('country', '')).strip()
    tax_id = str(row_data.get('tax_id', '')).strip()
    
    household_count = row_data.get('household_count', 0) or 0
    credit_limit = row_data.get('credit_limit', 0) or 0
    payment_terms = row_data.get('payment_terms', 30) or 30
    is_active = row_data.get('is_active', True)
    
    try:
        household_count = int(float(household_count))
    except (ValueError, TypeError):
        household_count = 0
    try:
        credit_limit = float(credit_limit)
    except (ValueError, TypeError):
        credit_limit = 0
    try:
        payment_terms = int(float(payment_terms))
    except (ValueError, TypeError):
        payment_terms = 30
    
    if isinstance(is_active, str):
        is_active = is_active.lower().strip() in ['true', 'yes', '1', 'active', 'true']
    else:
        is_active = bool(is_active)
    
    existing_client = Beneficiary.objects.filter(name=name).first()
    
    if existing_client:
        existing_client.beneficiary_type = beneficiary_type
        existing_client.email = email
        existing_client.phone = phone
        existing_client.village = village
        existing_client.scheme = scheme
        existing_client.country = country
        existing_client.tax_id = tax_id
        existing_client.household_count = household_count
        existing_client.credit_limit = credit_limit
        existing_client.payment_terms = payment_terms
        existing_client.is_active = is_active
        existing_client.created_by = user
        existing_client.save()
        return {'status': 'updated', 'beneficiary': existing_client}
    else:
        client = Beneficiary(
            name=name,
            beneficiary_type=beneficiary_type,
            email=email,
            phone=phone,
            village=village,
            scheme=scheme,
            country=country,
            tax_id=tax_id,
            household_count=household_count,
            credit_limit=credit_limit,
            payment_terms=payment_terms,
            is_active=is_active
        )
        client.created_by = user
        client.save()
        return {'status': 'created', 'beneficiary': client}


def import_vendor_data(row_data):
    name = row_data.get('name', '')
    if not name:
        return {'status': 'error', 'message': 'Name is required'}
    
    name = str(name).strip()
    email = str(row_data.get('email', '')).strip()
    phone = str(row_data.get('phone', '')).strip()
    address = str(row_data.get('address', '')).strip()
    city = str(row_data.get('city', '')).strip()
    country = str(row_data.get('country', '')).strip()
    tax_id = str(row_data.get('tax_id', '')).strip()
    
    payment_terms = row_data.get('payment_terms', 30) or 30
    is_active = row_data.get('is_active', True)
    
    try:
        payment_terms = int(float(payment_terms))
    except (ValueError, TypeError):
        payment_terms = 30
    
    if isinstance(is_active, str):
        is_active = is_active.lower().strip() in ['true', 'yes', '1', 'active']
    else:
        is_active = bool(is_active)
    
    existing_vendor = Vendor.objects.filter(name=name).first()
    
    if existing_vendor:
        existing_vendor.email = email
        existing_vendor.phone = phone
        existing_vendor.address = address
        existing_vendor.city = city
        existing_vendor.country = country
        existing_vendor.tax_id = tax_id
        existing_vendor.payment_terms = payment_terms
        existing_vendor.is_active = is_active
        existing_vendor.save()
        return {'status': 'updated', 'vendor': existing_vendor}
    else:
        vendor = Vendor.objects.create(
            name=name,
            email=email,
            phone=phone,
            address=address,
            city=city,
            country=country,
            tax_id=tax_id,
            payment_terms=payment_terms,
            is_active=is_active
        )
        return {'status': 'created', 'vendor': vendor}


def import_account_data(row_data):
    name = row_data.get('name', '')
    code = row_data.get('code', '')
    
    if not name or not code:
        return {'status': 'error', 'message': 'Name and Code are required'}
    
    name = str(name).strip()
    code = str(code).strip()
    account_type = str(row_data.get('account_type', 'asset')).strip() or 'asset'
    description = str(row_data.get('description', '')).strip()
    is_active = row_data.get('is_active', True)
    
    if isinstance(is_active, str):
        is_active = is_active.lower().strip() in ['true', 'yes', '1', 'active']
    else:
        is_active = bool(is_active)
    
    existing_account = Account.objects.filter(code=code).first()
    
    if existing_account:
        existing_account.name = name
        existing_account.account_type = account_type
        existing_account.description = description
        existing_account.is_active = is_active
        existing_account.save()
        return {'status': 'updated', 'account': existing_account}
    else:
        account = Account.objects.create(
            name=name,
            code=code,
            account_type=account_type,
            description=description,
            is_active=is_active
        )
        return {'status': 'created', 'account': account}


def import_invoice_data(row_data, user):
    invoice_number = row_data.get('invoice_number', '')
    client_name = row_data.get('beneficiary', '')
    
    if not invoice_number:
        return {'status': 'error', 'message': 'Invoice number is required'}
    if not client_name:
        return {'status': 'error', 'message': 'Beneficiary name is required'}
    
    invoice_number = str(invoice_number).strip()
    client_name = str(client_name).strip()
    
    try:
        client = Beneficiary.objects.get(name=client_name)
    except Beneficiary.DoesNotExist:
        return {'status': 'error', 'message': f'Beneficiary "{beneficiary_name}" not found. Please import client first.'}
    except Beneficiary.MultipleObjectsReturned:
        client = Beneficiary.objects.filter(name=client_name).first()
    
    issue_date = row_data.get('issue_date')
    due_date = row_data.get('due_date')
    status = str(row_data.get('status', 'draft')).strip() or 'draft'
    household_count = row_data.get('household_count', 0) or 0
    cost_per_unit = row_data.get('cost_per_unit', 0) or 0
    tax_rate = row_data.get('tax_rate', 0) or 0
    discount = row_data.get('discount', 0) or 0
    notes = str(row_data.get('notes', '')).strip()
    
    try:
        household_count = int(float(household_count))
    except (ValueError, TypeError):
        household_count = 0
    try:
        cost_per_unit = float(cost_per_unit)
    except (ValueError, TypeError):
        cost_per_unit = 0
    try:
        tax_rate = float(tax_rate)
    except (ValueError, TypeError):
        tax_rate = 0
    try:
        discount = float(discount)
    except (ValueError, TypeError):
        discount = 0
    
    if isinstance(issue_date, str):
        try:
            issue_date = datetime.strptime(issue_date, '%Y-%m-%d').date()
        except:
            try:
                issue_date = datetime.strptime(issue_date, '%Y/%m/%d').date()
            except:
                issue_date = timezone.now().date()
    elif hasattr(issue_date, 'date') and not isinstance(issue_date, datetime):
        issue_date = issue_date.date() if hasattr(issue_date, 'date') else timezone.now().date()
    elif isinstance(issue_date, datetime):
        issue_date = issue_date.date()
    else:
        issue_date = timezone.now().date()
    
    if isinstance(due_date, str):
        try:
            due_date = datetime.strptime(due_date, '%Y-%m-%d').date()
        except:
            try:
                due_date = datetime.strptime(due_date, '%Y/%m/%d').date()
            except:
                due_date = timezone.now().date() + timedelta(days=30)
    elif hasattr(due_date, 'date') and not isinstance(due_date, datetime):
        due_date = due_date.date() if hasattr(due_date, 'date') else timezone.now().date() + timedelta(days=30)
    elif isinstance(due_date, datetime):
        due_date = due_date.date()
    else:
        due_date = timezone.now().date() + timedelta(days=30)
    
    subtotal = household_count * cost_per_unit
    tax_amount = subtotal * (tax_rate / 100)
    total_amount = subtotal + tax_amount - discount
    
    existing_invoice = Invoice.objects.filter(invoice_number=invoice_number).first()
    
    if existing_invoice:
        existing_invoice.beneficiary_id = client.pk
        existing_invoice.issue_date = issue_date
        existing_invoice.due_date = due_date
        existing_invoice.status = status
        existing_invoice.household_count = household_count
        existing_invoice.cost_per_unit = cost_per_unit
        existing_invoice.tax_rate = tax_rate
        existing_invoice.tax_amount = tax_amount
        existing_invoice.discount = discount
        existing_invoice.total_amount = total_amount
        existing_invoice.notes = notes
        existing_invoice.save()
        return {'status': 'updated', 'invoice': existing_invoice}
    else:
        invoice = Invoice.objects.create(
            invoice_number=invoice_number,
            beneficiary_id=client.pk,
            issue_date=issue_date,
            due_date=due_date,
            status=status,
            household_count=household_count,
            cost_per_unit=cost_per_unit,
            tax_rate=tax_rate,
            tax_amount=tax_amount,
            discount=discount,
            total_amount=total_amount,
            notes=notes,
            created_by=user
        )
        return {'status': 'created', 'invoice': invoice}


def import_expense_data(row_data, user):
    expense_number = row_data.get('expense_number', '')
    
    if not expense_number:
        return {'status': 'error', 'message': 'Expense number is required'}
    
    expense_number = str(expense_number).strip()
    description = str(row_data.get('description', '')).strip()
    amount = row_data.get('amount', 0) or 0
    expense_date = row_data.get('expense_date')
    vendor_name = row_data.get('vendor', '')
    account_code = row_data.get('account_code', '')
    is_paid = row_data.get('is_paid', False)
    
    try:
        amount = float(amount)
    except (ValueError, TypeError):
        amount = 0
    
    if isinstance(expense_date, str):
        try:
            expense_date = datetime.strptime(expense_date, '%Y-%m-%d').date()
        except:
            try:
                expense_date = datetime.strptime(expense_date, '%Y/%m/%d').date()
            except:
                expense_date = timezone.now().date()
    elif isinstance(expense_date, datetime):
        expense_date = expense_date.date()
    elif hasattr(expense_date, 'date'):
        expense_date = expense_date.date()
    else:
        expense_date = timezone.now().date()
    
    if isinstance(is_paid, str):
        is_paid = is_paid.lower().strip() in ['true', 'yes', '1', 'paid']
    else:
        is_paid = bool(is_paid)
    
    vendor_id = None
    if vendor_name:
        vendor_name = str(vendor_name).strip()
        if vendor_name:
            try:
                vendor = Vendor.objects.get(name=vendor_name)
                vendor_id = vendor.pk
            except Vendor.DoesNotExist:
                pass
    
    account_id = None
    if account_code:
        account_code = str(account_code).strip()
        if account_code:
            try:
                account = Account.objects.get(code=account_code)
                account_id = account.pk
            except Account.DoesNotExist:
                pass
    
    existing_expense = Expense.objects.filter(expense_number=expense_number).first()
    
    if existing_expense:
        existing_expense.vendor_id = vendor_id
        existing_expense.description = description
        existing_expense.amount = amount
        existing_expense.expense_date = expense_date
        existing_expense.account_id = account_id
        existing_expense.is_paid = is_paid
        existing_expense.save()
        return {'status': 'updated', 'expense': existing_expense}
    else:
        expense = Expense.objects.create(
            expense_number=expense_number,
            vendor_id=vendor_id,
            description=description,
            amount=amount,
            expense_date=expense_date,
            account_id=account_id,
            is_paid=is_paid,
            created_by=user
        )
        return {'status': 'created', 'expense': expense}


def import_payment_data(row_data, user):
    client_name = row_data.get('beneficiary', '')
    amount = row_data.get('amount', 0) or 0
    payment_date = row_data.get('payment_date')
    payment_method = str(row_data.get('payment_method', 'bank_transfer')).strip() or 'bank_transfer'
    reference = str(row_data.get('reference', '')).strip()
    invoice_number = row_data.get('invoice_number', '')
    
    if not client_name:
        return {'status': 'error', 'message': 'Beneficiary name is required'}
    
    client_name = str(client_name).strip()
    
    try:
        client = Beneficiary.objects.get(name=client_name)
    except Beneficiary.DoesNotExist:
        return {'status': 'error', 'message': f'Beneficiary "{beneficiary_name}" not found. Please import client first.'}
    except Beneficiary.MultipleObjectsReturned:
        client = Beneficiary.objects.filter(name=client_name).first()
    
    try:
        amount = float(amount)
    except (ValueError, TypeError):
        amount = 0
    
    if isinstance(payment_date, str):
        try:
            payment_date = datetime.strptime(payment_date, '%Y-%m-%d').date()
        except:
            try:
                payment_date = datetime.strptime(payment_date, '%Y/%m/%d').date()
            except:
                payment_date = timezone.now().date()
    elif isinstance(payment_date, datetime):
        payment_date = payment_date.date()
    elif hasattr(payment_date, 'date'):
        payment_date = payment_date.date()
    else:
        payment_date = timezone.now().date()
    
    invoice_id = None
    if invoice_number:
        invoice_number = str(invoice_number).strip()
        if invoice_number:
            invoice = Invoice.objects.filter(invoice_number=invoice_number).first()
            if invoice:
                invoice_id = invoice.pk
    
    payment = Payment.objects.create(
        beneficiary_id=client.pk,
        invoice_id=invoice_id,
        amount=amount,
        payment_date=payment_date,
        payment_method=payment_method,
        reference=reference,
        created_by=user
    )
    
    if invoice_id:
        invoice = Invoice.objects.get(pk=invoice_id)
        if invoice.amount_due() <= 0:
            invoice.status = 'paid'
        else:
            invoice.status = 'partial'
        invoice.save()
    
    return {'status': 'created', 'payment': payment}


def import_budget_data(row_data):
    account_code = row_data.get('account_code', '')
    fiscal_year = row_data.get('fiscal_year', '')
    amount = row_data.get('amount', 0) or 0
    notes = str(row_data.get('notes', '')).strip()
    
    if not account_code:
        return {'status': 'error', 'message': 'Account code is required'}
    if not fiscal_year:
        return {'status': 'error', 'message': 'Fiscal year is required'}
    
    account_code = str(account_code).strip()
    
    try:
        account = Account.objects.get(code=account_code)
    except Account.DoesNotExist:
        return {'status': 'error', 'message': f'Account code "{account_code}" not found. Please import account first.'}
    except Account.MultipleObjectsReturned:
        account = Account.objects.filter(code=account_code).first()
    
    try:
        fiscal_year = int(float(fiscal_year))
    except (ValueError, TypeError):
        fiscal_year = timezone.now().year
    
    try:
        amount = float(amount)
    except (ValueError, TypeError):
        amount = 0
    
    existing_budget = Budget.objects.filter(account=account, fiscal_year=fiscal_year).first()
    
    if existing_budget:
        existing_budget.amount = amount
        existing_budget.notes = notes
        existing_budget.save()
        return {'status': 'updated', 'budget': existing_budget}
    else:
        budget = Budget.objects.create(
            account_id=account.pk,
            fiscal_year=fiscal_year,
            amount=amount,
            notes=notes
        )
        return {'status': 'created', 'budget': budget}


@login_required
def download_import_template(request):
    model_type = request.GET.get('type', 'beneficiary')
    
    wb = Workbook()
    ws = wb.active
    
    if model_type == 'beneficiary':
        ws.title = "Clients"
        headers = ['name', 'beneficiary_type', 'email', 'phone', 'village', 'scheme', 'country', 'tax_id', 'household_count', 'credit_limit', 'payment_terms', 'is_active']
        ws.append(headers)
        ws.append(['Acme Corp', 'private', 'contact@acme.com', '+265991234567', 'Lilongwe', 'Mangale', 'Malawi', 'TAX123', '50', '100000', '30', 'True'])
        ws.append(['Village Water Committee', 'communal', '', '+265991234568', 'Mzuzu', 'Nkala', 'Malawi', '', '200', '0', '30', 'True'])
    elif model_type == 'vendor':
        ws.title = "Vendors"
        headers = ['name', 'email', 'phone', 'address', 'city', 'country', 'tax_id', 'payment_terms', 'is_active']
        ws.append(headers)
        ws.append(['Office Supplies Ltd', 'orders@officesupplies.com', '+265991234569', '123 Main St', 'Lilongwe', 'Malawi', 'TAX456', '30', 'True'])
        ws.append(['Tech Solutions', 'sales@techsolutions.com', '+265991234570', '456 Business Ave', 'Blantyre', 'Malawi', 'TAX789', '45', 'True'])
    elif model_type == 'account':
        ws.title = "Accounts"
        headers = ['name', 'code', 'account_type', 'description', 'is_active']
        ws.append(headers)
        ws.append(['Cash', '1000', 'asset', 'Cash on hand', 'True'])
        ws.append(['Accounts Receivable', '1100', 'asset', 'Money owed by clients', 'True'])
        ws.append(['Sales Revenue', '4000', 'revenue', 'Revenue from water services', 'True'])
    elif model_type == 'invoice':
        ws.title = "Invoices"
        headers = ['invoice_number', 'beneficiary', 'issue_date', 'due_date', 'status', 'household_count', 'cost_per_unit', 'tax_rate', 'discount', 'notes']
        ws.append(headers)
        ws.append(['INV-20260101-0001', 'Acme Corp', '2026-01-01', '2026-01-31', 'draft', '50', '100', '16', '0', 'Monthly water service'])
        ws.append(['INV-20260101-0002', 'Village Water Committee', '2026-01-01', '2026-02-01', 'sent', '200', '50', '0', '0', 'Quarterly billing'])
    elif model_type == 'expense':
        ws.title = "Expenses"
        headers = ['expense_number', 'description', 'amount', 'expense_date', 'vendor', 'account_code', 'is_paid']
        ws.append(headers)
        ws.append(['EXP-20260101-0001', 'Office supplies purchase', '5000', '2026-01-15', 'Office Supplies Ltd', '5000', 'True'])
        ws.append(['EXP-20260101-0002', 'Internet service', '2500', '2026-01-20', '', '5200', 'True'])
    elif model_type == 'payment':
        ws.title = "Payments"
        headers = ['beneficiary', 'amount', 'payment_date', 'payment_method', 'reference', 'invoice_number']
        ws.append(headers)
        ws.append(['Acme Corp', '5000', '2026-01-20', 'bank_transfer', 'REF-001', 'INV-20260101-0001'])
        ws.append(['Village Water Committee', '10000', '2026-01-25', 'cash', 'REF-002', 'INV-20260101-0002'])
    elif model_type == 'budget':
        ws.title = "Budgets"
        headers = ['account_code', 'fiscal_year', 'amount', 'notes']
        ws.append(headers)
        ws.append(['1000', '2026', '1000000', 'Cash budget for operations'])
        ws.append(['5000', '2026', '500000', 'Operating expenses budget'])
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="import_template_{model_type}.xlsx"'
    wb.save(response)
    return response


@login_required
@login_required
def opening_balance_list(request):
    if not hasattr(request.user, 'userprofile'):
        user_profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not request.user.userprofile.can_edit():
        messages.error(request, "You don't have permission to manage opening balances.")
        return redirect('dashboard')
    
    current_year = timezone.now().year
    fiscal_year = request.GET.get('fiscal_year', current_year)
    try:
        fiscal_year = int(fiscal_year)
    except ValueError:
        fiscal_year = current_year
    
    clients = Beneficiary.objects.filter(is_active=True).order_by('name')
    
    client_balances = []
    for beneficiary in clients:
        opening = beneficiary.opening_balances.filter(fiscal_year=fiscal_year).first()
        client_balances.append({
            'beneficiary': beneficiary,
            'opening_balance': opening.amount if opening else Decimal('0.00'),
            'total_outstanding': beneficiary.total_outstanding,
            'total_with_opening': beneficiary.get_total_balance_with_opening(fiscal_year),
        })
    
    total_opening = sum(b['opening_balance'] for b in client_balances)
    total_outstanding = sum(b['total_outstanding'] for b in client_balances)
    total_with_opening = sum(b['total_with_opening'] for b in client_balances)
    
    rollover_history = YearEndRollover.objects.all()[:10]
    
    context = {
        'fiscal_year': fiscal_year,
        'client_balances': client_balances,
        'total_opening': total_opening,
        'total_outstanding': total_outstanding,
        'total_with_opening': total_with_opening,
        'rollover_history': rollover_history,
        'user_profile': request.user.userprofile,
    }
    return render(request, "accounting_app/opening_balance_list.html", context)


@login_required
def opening_balance_edit(request, beneficiary_id):
    if not hasattr(request.user, 'userprofile'):
        user_profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not request.user.userprofile.can_edit():
        messages.error(request, "You don't have permission to edit opening balances.")
        return redirect('dashboard')
    
    beneficiary = get_object_or_404(Beneficiary, pk=beneficiary_id)
    current_year = timezone.now().year
    fiscal_year = request.GET.get('fiscal_year', current_year)
    try:
        fiscal_year = int(fiscal_year)
    except ValueError:
        fiscal_year = current_year
    
    if request.method == 'POST':
        amount = request.POST.get('amount', 0)
        notes = request.POST.get('notes', '')
        
        try:
            amount = Decimal(amount)
        except:
            amount = Decimal('0.00')
        
        opening, created = OpeningBalance.objects.update_or_create(
            beneficiary=beneficiary,
            fiscal_year=fiscal_year,
            defaults={
                'amount': amount,
                'notes': notes,
                'created_by': request.user,
            }
        )
        
        if created:
            messages.success(request, f"Opening balance of {request.user.userprofile.get_currency_symbol()}{amount:,.2f} created for {beneficiary.name} (FY {fiscal_year})")
        else:
            messages.success(request, f"Opening balance updated to {request.user.userprofile.get_currency_symbol()}{amount:,.2f} for {beneficiary.name} (FY {fiscal_year})")
        
        return redirect('opening_balance_list')
    
    opening = beneficiary.opening_balances.filter(fiscal_year=fiscal_year).first()
    
    context = {
        'beneficiary': beneficiary,
        'fiscal_year': fiscal_year,
        'opening_balance': opening,
        'currency_symbol': request.user.userprofile.get_currency_symbol(),
    }
    return render(request, "accounting_app/opening_balance_edit.html", context)


@login_required
def year_end_rollover(request):
    if not hasattr(request.user, 'userprofile'):
        user_profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not request.user.userprofile.can_edit():
        messages.error(request, "You don't have permission to perform year-end rollover.")
        return redirect('dashboard')
    
    current_year = timezone.now().year
    
    if request.method == 'POST':
        from_year = request.POST.get('from_year')
        to_year = request.POST.get('to_year')
        action = request.POST.get('action', 'preview')
        
        try:
            from_year = int(from_year)
            to_year = int(to_year)
        except ValueError:
            messages.error(request, "Invalid year values")
            return redirect('year_end_rollover')
        
        if to_year != from_year + 1:
            messages.error(request, "Target year must be exactly one year after the source year")
            return redirect('year_end_rollover')
        
        start_day = request.POST.get('start_day')
        start_month = request.POST.get('start_month')
        start_year = request.POST.get('start_year', str(to_year))
        
        from datetime import date
        try:
            rollover_date = date(int(start_year), int(start_month), int(start_day))
        except (ValueError, TypeError):
            rollover_date = date(to_year, 1, 1)
            messages.warning(request, "Invalid date provided, defaulting to January 1st")
        
        clients = Beneficiary.objects.filter(is_active=True)
        
        if action == 'execute':
            total_balance = Decimal('0.00')
            rollover_list = []
            
            for beneficiary in clients:
                outstanding = beneficiary.total_outstanding
                
                opening, created = OpeningBalance.objects.update_or_create(
                    beneficiary=beneficiary,
                    fiscal_year=to_year,
                    defaults={
                        'amount': outstanding,
                        'notes': f"Carried forward from FY {from_year} (Outstanding: {outstanding})",
                        'created_by': request.user,
                    }
                )
                
                total_balance += outstanding
                rollover_list.append({
                    'beneficiary': beneficiary,
                    'amount': outstanding,
                })
            
            YearEndRollover.objects.create(
                fiscal_year=from_year,
                rollover_date=rollover_date,
                total_clients=clients.count(),
                total_opening_balance=total_balance,
                notes=f"Rollover to FY {to_year}",
                created_by=request.user,
            )
            
            messages.success(request, f"Year-end rollover completed! {clients.count()} clients' balances totaling {request.user.userprofile.get_currency_symbol()}{total_balance:,.2f} carried forward to FY {to_year}")
            return redirect('opening_balance_list')
        
        else:
            rollover_data = []
            total_balance = Decimal('0.00')
            
            for beneficiary in clients:
                outstanding = beneficiary.total_outstanding
                existing = beneficiary.opening_balances.filter(fiscal_year=to_year).first()
                
                rollover_data.append({
                    'beneficiary': beneficiary,
                    'outstanding': outstanding,
                    'existing_opening': existing.amount if existing else Decimal('0.00'),
                    'will_be_set': outstanding,
                })
                total_balance += outstanding
            
            context = {
                'from_year': from_year,
                'to_year': to_year,
                'rollover_data': rollover_data,
                'total_balance': total_balance,
                'client_count': clients.count(),
                'currency_symbol': request.user.userprofile.get_currency_symbol(),
            }
            return render(request, "accounting_app/year_end_rollover_preview.html", context)
    
    last_rollover = YearEndRollover.objects.first()
    next_year = current_year + 1
    
    context = {
        'current_year': current_year,
        'next_year': next_year,
        'last_rollover': last_rollover,
    }
    return render(request, "accounting_app/year_end_rollover.html", context)


@login_required
def beneficiary_detail(request, pk):
    beneficiary = get_object_or_404(Beneficiary, pk=pk)
    invoices = beneficiary.invoices.all().order_by('-issue_date')
    payments = beneficiary.payments.all().order_by('-payment_date')
    opening_balances = beneficiary.opening_balances.all().order_by('-fiscal_year')
    
    current_year = timezone.now().year
    current_opening = beneficiary.get_opening_balance(current_year)
    balance_history = beneficiary.balance_history.all()
    status_logs = beneficiary.status_logs.all()
    latest_activated = status_logs.filter(status='activated').order_by('-changed_at').first()
    latest_deactivated = status_logs.filter(status='deactivated').order_by('-changed_at').first()
    
    change_log = []
    for log in status_logs:
        change_log.append({
            'timestamp': log.changed_at,
            'category': 'status',
            'type': log.status,
            'description': f"Beneficiary {log.status}",
            'user': log.user,
        })
    for h in beneficiary.history.all():
        change_log.append({
            'timestamp': h.timestamp,
            'category': 'field',
            'type': h.action,
            'description': h.description,
            'user': h.user,
            'field': h.field_name,
            'old_value': h.old_value,
            'new_value': h.new_value,
        })
    change_log.sort(key=lambda x: x['timestamp'], reverse=True)
    
    return render(request, "accounting_app/beneficiary_detail.html", {
        "beneficiary": beneficiary,
        "invoices": invoices,
        "payments": payments,
        "opening_balances": opening_balances,
        "balance_history": balance_history,
        "change_log": change_log,
        "latest_activated": latest_activated,
        "latest_deactivated": latest_deactivated,
        "current_opening": current_opening,
        "current_fiscal_year": current_year,
    })


@login_required
def beneficiary_invoices_json(request, pk):
    from django.http import JsonResponse
    beneficiary = get_object_or_404(Beneficiary, pk=pk)
    invoices = beneficiary.invoices.all().order_by('-issue_date')
    data = []
    for inv in invoices:
        data.append({
            'id': inv.id,
            'invoice_number': inv.invoice_number,
            'total_amount': float(inv.total_amount),
            'amount_due': float(inv.amount_due()),
            'amount_paid': float(inv.amount_paid()),
            'status': inv.status,
            'issue_date': inv.issue_date.strftime('%Y-%m-%d'),
            'due_date': inv.due_date.strftime('%Y-%m-%d') if inv.due_date else None,
        })
    return JsonResponse(data, safe=False)


@login_required
def total_population(request):
    schemes = Scheme.objects.filter(is_active=True).prefetch_related('villages__population_records')
    
    scheme_data = []
    grand_total = 0
    grand_households = 0
    
    for scheme in schemes:
        villages = scheme.villages.filter(is_active=True)
        village_data = []
        scheme_total = 0
        scheme_households = 0
        
        for village in villages:
            latest_pop = village.population_records.order_by('-recorded_date').first()
            population = latest_pop.population if latest_pop else 0
            scheme_total += population
            scheme_households += village.household_count or 0
            village_data.append({
                'id': village.id,
                'name': village.name,
                'household_count': village.household_count or 0,
                'population': population,
                'last_updated': latest_pop.recorded_date if latest_pop else None,
            })
        
        grand_total += scheme_total
        grand_households += scheme_households
        scheme_data.append({
            'id': scheme.id,
            'name': scheme.name,
            'code': scheme.code,
            'villages': village_data,
            'total_population': scheme_total,
            'total_households': scheme_households,
        })
    
    return render(request, "accounting_app/total_population.html", {
        'scheme_data': scheme_data,
        'grand_total': grand_total,
        'grand_households': grand_households,
    })


@login_required
def update_village_population(request, village_id):
    village = get_object_or_404(Village, pk=village_id)
    
    if request.method == 'POST':
        population = request.POST.get('population', 0)
        household_count = request.POST.get('household_count', 0)
        recorded_date = request.POST.get('recorded_date', timezone.now().date())
        notes = request.POST.get('notes', '')
        
        try:
            population = int(population)
        except (ValueError, TypeError):
            population = 0
        
        try:
            household_count = int(household_count)
        except (ValueError, TypeError):
            household_count = 0
        
        if isinstance(recorded_date, str):
            try:
                recorded_date = datetime.strptime(recorded_date, '%Y-%m-%d').date()
            except:
                recorded_date = timezone.now().date()
        
        village.household_count = household_count
        village.save()
        
        VillagePopulation.objects.create(
            village=village,
            population=population,
            recorded_date=recorded_date,
            recorded_by=request.user,
            notes=notes
        )
        messages.success(request, f"{village.name}: {household_count} households, {population} population - saved successfully")
        return redirect('total_population')
    
    return redirect('total_population')


@login_required
def bulk_update_population(request):
    if request.method == 'POST':
        recorded_date = request.POST.get('recorded_date', timezone.now().date())
        notes = request.POST.get('notes', '')
        
        if isinstance(recorded_date, str):
            try:
                recorded_date = datetime.strptime(recorded_date, '%Y-%m-%d').date()
            except:
                recorded_date = timezone.now().date()
        
        updated_count = 0
        household_updated_count = 0
        for key, value in request.POST.items():
            if key.startswith('population_'):
                try:
                    village_id = int(key.replace('population_', ''))
                    population = int(value) if value else 0
                    village = Village.objects.get(pk=village_id)
                    VillagePopulation.objects.create(
                        village=village,
                        population=population,
                        recorded_date=recorded_date,
                        recorded_by=request.user,
                        notes=notes
                    )
                    updated_count += 1
                except (ValueError, Village.DoesNotExist):
                    pass
            elif key.startswith('household_'):
                try:
                    village_id = int(key.replace('household_', ''))
                    household_count = int(value) if value else 0
                    village = Village.objects.get(pk=village_id)
                    village.household_count = household_count
                    village.save()
                    household_updated_count += 1
                except (ValueError, Village.DoesNotExist):
                    pass
        
        if updated_count > 0 or household_updated_count > 0:
            messages.success(request, f"Successfully updated {updated_count} village populations and {household_updated_count} household counts")
        else:
            messages.warning(request, "No values were updated")
        
        return redirect('total_population')
    
    return redirect('total_population')


@login_required
def manage_schemes(request):
    if not request.user.userprofile.can_edit():
        messages.error(request, "You don't have permission to manage schemes.")
        return redirect('dashboard')
    
    schemes = Scheme.objects.all().prefetch_related('villages')
    return render(request, "accounting_app/manage_schemes.html", {'schemes': schemes})


@login_required
def scheme_create(request):
    if not request.user.userprofile.can_edit():
        messages.error(request, "You don't have permission to create schemes.")
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = SchemeForm(request.POST)
        if form.is_valid():
            scheme = form.save(commit=False)
            name = scheme.name.strip().upper()
            code = ''.join([w[0] for w in name.split()])[:4]
            base_code = code
            counter = 1
            while Scheme.objects.filter(code=code).exists():
                code = f"{base_code}{counter}"
                counter += 1
            scheme.code = code
            scheme.save()
            messages.success(request, "Scheme created successfully")
            return redirect('manage_schemes')
    else:
        form = SchemeForm()
    return render(request, "accounting_app/scheme_form.html", {'form': form, 'action': 'Create'})


@login_required
def scheme_edit(request, pk):
    if not request.user.userprofile.can_edit():
        messages.error(request, "You don't have permission to edit schemes.")
        return redirect('dashboard')
    
    scheme = get_object_or_404(Scheme, pk=pk)
    if request.method == 'POST':
        form = SchemeForm(request.POST, instance=scheme)
        if form.is_valid():
            form.save()
            messages.success(request, "Scheme updated successfully")
            return redirect('manage_schemes')
    else:
        form = SchemeForm(instance=scheme)
    return render(request, "accounting_app/scheme_form.html", {'form': form, 'action': 'Update', 'scheme': scheme})


@login_required
def scheme_delete(request, pk):
    if not request.user.userprofile.can_edit():
        messages.error(request, "You don't have permission to delete schemes.")
        return redirect('dashboard')
    
    scheme = get_object_or_404(Scheme, pk=pk)
    if request.method == 'POST':
        save_deleted_record(scheme, request.user)
        scheme.delete()
        messages.success(request, f"Scheme '{scheme.name}' deleted successfully")
        return redirect('manage_schemes')
    return render(request, "accounting_app/scheme_delete.html", {'scheme': scheme})


@login_required
def manage_villages(request):
    if not request.user.userprofile.can_edit():
        messages.error(request, "You don't have permission to manage villages.")
        return redirect('dashboard')
    
    villages = Village.objects.select_related('scheme').all()
    schemes = Scheme.objects.filter(is_active=True)
    return render(request, "accounting_app/manage_villages.html", {'villages': villages, 'schemes': schemes})


@login_required
def village_create(request):
    if not request.user.userprofile.can_edit():
        messages.error(request, "You don't have permission to create villages.")
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = VillageForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Village created successfully")
            return redirect('manage_villages')
    else:
        form = VillageForm()
    return render(request, "accounting_app/village_form.html", {'form': form, 'action': 'Create'})


@login_required
def village_edit(request, pk):
    if not request.user.userprofile.can_edit():
        messages.error(request, "You don't have permission to edit villages.")
        return redirect('dashboard')
    
    village = get_object_or_404(Village, pk=pk)
    if request.method == 'POST':
        form = VillageForm(request.POST, instance=village)
        if form.is_valid():
            form.save()
            messages.success(request, "Village updated successfully")
            return redirect('manage_villages')
    else:
        form = VillageForm(instance=village)
    return render(request, "accounting_app/village_form.html", {'form': form, 'action': 'Update', 'village': village})


@login_required
def village_delete(request, pk):
    if not request.user.userprofile.can_edit():
        messages.error(request, "You don't have permission to delete villages.")
        return redirect('dashboard')
    
    village = get_object_or_404(Village, pk=pk)
    if request.method == 'POST':
        save_deleted_record(village, request.user)
        village.delete()
        messages.success(request, f"Village '{village.name}' deleted successfully")
        return redirect('manage_villages')
    return render(request, "accounting_app/village_delete.html", {'village': village})


@login_required
def initialize_default_schemes(request):
    if not request.user.userprofile.can_edit():
        messages.error(request, "You don't have permission.")
        return redirect('dashboard')
    
    default_schemes = [
        {'name': 'Mangale', 'code': 'MGL', 'description': 'Mangale Scheme'},
        {'name': 'Nkala', 'code': 'NKL', 'description': 'Nkala Scheme'},
        {'name': 'Dodza', 'code': 'DDZ', 'description': 'Dodza Scheme'},
        {'name': 'Milala', 'code': 'MLL', 'description': 'Milala Scheme'},
    ]
    
    for scheme_data in default_schemes:
        scheme, created = Scheme.objects.get_or_create(
            code=scheme_data['code'],
            defaults={
                'name': scheme_data['name'],
                'description': scheme_data['description'],
            }
        )
        if created:
            default_villages = {
                'Mangale': ['Kika', 'Maoni', 'Machemba'],
                'Nkala': ['Chikondi', 'Mchenga', 'Namitete'],
                'Dodza': ['Chimwemwe', 'Dziko', 'Phwetekela'],
                'Milala': ['Chimbalanga', 'Kachulu', 'Mpenda'],
            }
            village_names = default_villages.get(scheme.name, [])
            for vname in village_names:
                Village.objects.get_or_create(
                    scheme=scheme,
                    name=vname,
                    defaults={'description': f'{vname} Village under {scheme.name}'}
                )
    
    messages.success(request, "Default schemes and villages initialized successfully")
    return redirect('manage_schemes')


@login_required
def board_of_trustees_list(request):
    members = BoardOfTrustees.objects.all().order_by('name')
    return render(request, "accounting_app/board_of_trustees_list.html", {"members": members})


@login_required
def board_of_trustees_create(request):
    if request.method == "POST":
        form = BoardOfTrusteesForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Board of Trustees member added successfully")
            return redirect("board_of_trustees_list")
    else:
        form = BoardOfTrusteesForm()
    return render(request, "accounting_app/board_of_trustees_form.html", {"form": form, "action": "Create"})


@login_required
def board_of_trustees_edit(request, pk):
    member = get_object_or_404(BoardOfTrustees, pk=pk)
    if request.method == "POST":
        form = BoardOfTrusteesForm(request.POST, instance=member)
        if form.is_valid():
            form.save()
            messages.success(request, "Board of Trustees member updated successfully")
            return redirect("board_of_trustees_list")
    else:
        form = BoardOfTrusteesForm(instance=member)
    return render(request, "accounting_app/board_of_trustees_form.html", {"form": form, "action": "Update", "member": member})


@login_required
def board_of_trustees_delete(request, pk):
    member = get_object_or_404(BoardOfTrustees, pk=pk)
    if request.method == "POST":
        save_deleted_record(member, request.user)
        member.delete()
        messages.success(request, "Board of Trustees member deleted successfully")
        return redirect("board_of_trustees_list")
    return render(request, "accounting_app/board_of_trustees_delete.html", {"member": member, "title": "Board of Trustees Member"})


@login_required
def general_assembly_list(request):
    members = GeneralAssemblyMember.objects.all().order_by('name')
    return render(request, "accounting_app/general_assembly_list.html", {"members": members})


@login_required
def general_assembly_create(request):
    if request.method == "POST":
        form = GeneralAssemblyMemberForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "General Assembly member added successfully")
            return redirect("general_assembly_list")
    else:
        form = GeneralAssemblyMemberForm()
    return render(request, "accounting_app/general_assembly_form.html", {"form": form, "action": "Create"})


@login_required
def general_assembly_edit(request, pk):
    member = get_object_or_404(GeneralAssemblyMember, pk=pk)
    if request.method == "POST":
        form = GeneralAssemblyMemberForm(request.POST, instance=member)
        if form.is_valid():
            form.save()
            messages.success(request, "General Assembly member updated successfully")
            return redirect("general_assembly_list")
    else:
        form = GeneralAssemblyMemberForm(instance=member)
    return render(request, "accounting_app/general_assembly_form.html", {"form": form, "action": "Update", "member": member})


@login_required
def general_assembly_delete(request, pk):
    member = get_object_or_404(GeneralAssemblyMember, pk=pk)
    if request.method == "POST":
        save_deleted_record(member, request.user)
        member.delete()
        messages.success(request, "General Assembly member deleted successfully")
        return redirect("general_assembly_list")
    return render(request, "accounting_app/board_of_trustees_delete.html", {"member": member, "title": "General Assembly Member"})


@login_required
def employee_list(request):
    employees = Employee.objects.all().order_by('name')
    return render(request, "accounting_app/employee_list.html", {"employees": employees})


@login_required
def employee_create(request):
    if request.method == "POST":
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Employee added successfully")
            return redirect("employee_list")
    else:
        form = EmployeeForm()
    return render(request, "accounting_app/employee_form.html", {"form": form, "action": "Create"})


@login_required
def employee_edit(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == "POST":
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, "Employee updated successfully")
            return redirect("employee_list")
    else:
        form = EmployeeForm(instance=employee)
    return render(request, "accounting_app/employee_form.html", {"form": form, "action": "Update", "employee": employee})


@login_required
def employee_delete(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == "POST":
        save_deleted_record(employee, request.user)
        employee.delete()
        messages.success(request, "Employee deleted successfully")
        return redirect("employee_list")
    return render(request, "accounting_app/board_of_trustees_delete.html", {"member": employee, "title": "Employee"})


@login_required
def pension_calculator(request):
    employees = Employee.objects.all().order_by('name')
    selected_employee = None
    pension_data = None
    
    if request.GET.get('employee_id'):
        selected_employee = get_object_or_404(Employee, pk=request.GET.get('employee_id'))
        
    pension_rate = Decimal('0.10')  # 10% pension rate
    
    context = {
        'employees': employees,
        'selected_employee': selected_employee,
        'pension_rate': pension_rate,
    }
    
    if selected_employee:
        salaries = selected_employee.salaries.all().order_by('-start_date')
        total_pension = Decimal('0.00')
        total_salary = Decimal('0.00')
        salary_breakdown = []
        
        for salary in salaries:
            days = salary.get_days_in_period()
            daily = salary.get_daily_salary()
            pension = salary.calculate_pension()
            total_pension += pension
            total_salary += daily * days
            
            salary_breakdown.append({
                'salary': salary,
                'days': days,
                'daily': daily,
                'total_earned': daily * days,
                'pension': pension,
            })
        
        context['salary_breakdown'] = salary_breakdown
        context['total_pension'] = total_pension
        context['total_salary'] = total_salary
    
    return render(request, "accounting_app/pension_calculator.html", context)


@login_required
def pension_report(request):
    employees = Employee.objects.all().order_by('name')
    pension_rate = Decimal('0.10')  # 10% pension rate
    
    employee_pensions = []
    total_pension = Decimal('0.00')
    total_salary = Decimal('0.00')
    
    for employee in employees:
        salaries = employee.salaries.all()
        
        employee_total_pension = Decimal('0.00')
        employee_total_salary = Decimal('0.00')
        periods_count = 0
        
        for salary in salaries:
            days = salary.get_days_in_period()
            daily = salary.get_daily_salary()
            pension = salary.calculate_pension()
            employee_total_pension += pension
            employee_total_salary += daily * days
            periods_count += 1
        
        if periods_count > 0 or employee.salary > 0:
            total_pension += employee_total_pension
            total_salary += employee_total_salary
            
            employee_pensions.append({
                'employee': employee,
                'total_pension': employee_total_pension,
                'total_salary': employee_total_salary,
                'periods_count': periods_count,
            })
    
    context = {
        'employee_pensions': employee_pensions,
        'total_pension': total_pension,
        'total_salary': total_salary,
        'pension_rate': pension_rate,
    }
    return render(request, "accounting_app/pension_report.html", context)


@login_required
def employee_salary_history(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    salaries = employee.salaries.all().order_by('-start_date')
    
    total_pension = Decimal('0.00')
    total_salary = Decimal('0.00')
    salary_breakdown = []
    
    for salary in salaries:
        days = salary.get_days_in_period()
        daily = salary.get_daily_salary()
        pension = salary.calculate_pension()
        total_pension += pension
        total_salary += daily * days
        
        salary_breakdown.append({
            'salary': salary,
            'days': days,
            'daily': daily,
            'total_earned': daily * days,
            'pension': pension,
        })
    
    context = {
        'employee': employee,
        'salary_breakdown': salary_breakdown,
        'total_pension': total_pension,
        'total_salary': total_salary,
    }
    return render(request, "accounting_app/employee_salary_history.html", context)


@login_required
def add_employee_salary(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    
    if request.method == 'POST':
        salary = request.POST.get('salary')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date') or None
        pension_rate = request.POST.get('pension_rate') or 10
        
        if salary and start_date:
            EmployeeSalary.objects.create(
                employee=employee,
                salary=Decimal(salary),
                start_date=start_date,
                end_date=end_date,
                pension_rate=Decimal(pension_rate),
                created_by=request.user,
            )
            messages.success(request, f"Salary record added for {employee.name}")
            return redirect('employee_salary_history', pk=employee.pk)
    
    return render(request, "accounting_app/add_employee_salary.html", {
        'employee': employee,
    })


@login_required
def delete_employee_salary(request, pk, salary_pk):
    salary = get_object_or_404(EmployeeSalary, pk=salary_pk, employee_id=pk)
    employee = salary.employee
    
    if request.method == 'POST':
        salary.delete()
        messages.success(request, "Salary record deleted")
        return redirect('employee_salary_history', pk=employee.pk)
    
    return render(request, "accounting_app/confirm_delete.html", {
        'object': salary,
        'back_url': reverse('employee_salary_history', args=[employee.pk]),
    })


@login_required
def activity_log(request):
    activities = ActivityLog.objects.all().order_by('-timestamp')[:100]
    return render(request, "accounting_app/activity_log.html", {"activities": activities})


@login_required
def audit_log(request):
    if not _is_super_admin(request.user):
        messages.error(request, "Access denied.")
        return redirect("dashboard")

    from django.utils.timesince import timesince

    all_activities = ActivityLog.objects.all().order_by('-timestamp')[:200]
    active_sessions = LoginSession.objects.select_related('user').filter(is_active=True).order_by('-login_time')
    all_sessions = LoginSession.objects.select_related('user').all().order_by('-login_time')[:50]

    module_filter = request.GET.get('module', '')
    if module_filter:
        all_activities = all_activities.filter(model_name__icontains=module_filter)

    sales_count = Invoice.objects.exclude(status='cancelled').count()
    refund_count = Payment.objects.filter(amount__lt=0).count()
    voided_count = Invoice.objects.filter(status='cancelled').count()
    adjustment_count = JournalEntry.objects.count()
    transfer_count = Payment.objects.filter(account__isnull=False).count()
    login_count = LoginSession.objects.count()

    for a in all_activities:
        a.time_ago = timesince(a.timestamp) + " ago" if a.timestamp else ""

    for s in all_sessions:
        s.login_ago = timesince(s.login_time) + " ago" if s.login_time else ""
        s.logout_ago = timesince(s.logout_time) + " ago" if s.logout_time else ""

    modules = ActivityLog.objects.values_list('model_name', flat=True).distinct().order_by('model_name')

    return render(request, "accounting_app/audit_log.html", {
        "activities": all_activities,
        "sessions": all_sessions,
        "active_sessions": active_sessions,
        "sales_count": sales_count,
        "refund_count": refund_count,
        "voided_count": voided_count,
        "adjustment_count": adjustment_count,
        "transfer_count": transfer_count,
        "login_count": login_count,
        "modules": modules,
        "current_module": module_filter,
    })


@login_required
def communication_list(request):
    communications = CommunicationLog.objects.all().order_by('-sent_at')
    comm_type = request.GET.get('type')
    if comm_type:
        communications = communications.filter(communication_type=comm_type)
    return render(request, "accounting_app/communication_list.html", {
        "communications": communications,
        "comm_type": comm_type
    })


@login_required
def communication_create(request):
    comm_type = request.GET.get('type', '')
    
    if request.method == 'POST':
        comm_type = request.POST.get('communication_type')
        recipient = request.POST.get('recipient')
        message = request.POST.get('message')
        beneficiary_id = request.POST.get('client')
        
        client = None
        if beneficiary_id:
            client = Beneficiary.objects.get(id=beneficiary_id) if beneficiary_id else None
        
        comm = CommunicationLog.objects.create(
            communication_type=comm_type,
            recipient=recipient,
            recipient_beneficiary=client,
            message=message,
            status='sent',
            sent_by=request.user
        )
        
        if comm_type == 'sms':
            from accounting_app.sms_service import send_sms_to_client
            result = send_sms_to_client(request.user.userprofile, recipient, message)
            if result.get('status') == 'success':
                comm.status = 'delivered'
            else:
                comm.status = 'failed'
            comm.notes = result.get('message', '')
            comm.save()
        
        elif comm_type == 'whatsapp':
            # For WhatsApp, we'll open WhatsApp web with pre-filled message
            # The actual sending would require WhatsApp Business API
            comm.status = 'sent'
            comm.notes = 'Opened WhatsApp Web for sending'
            comm.save()
            
            # Redirect to WhatsApp with pre-filled message
            import urllib.parse
            phone = recipient.replace('+', '').replace(' ', '').replace('-', '')
            wa_url = f"https://wa.me/{phone}?text={urllib.parse.quote(message)}"
            return redirect(wa_url)
        
        messages.success(request, f"Communication sent successfully")
        return redirect("communication_list")
    
    clients = Beneficiary.objects.filter(is_active=True)
    return render(request, "accounting_app/communication_form.html", {"clients": clients, "default_type": comm_type})


@login_required
def communication_send(request):
    if request.method == 'POST':
        comm_type = request.POST.get('communication_type')
        recipient = request.POST.get('recipient')
        message = request.POST.get('message')
        beneficiary_id = request.POST.get('client')
        
        client = None
        if beneficiary_id:
            client = Beneficiary.objects.get(id=beneficiary_id) if beneficiary_id else None
        
        comm = CommunicationLog.objects.create(
            communication_type=comm_type,
            recipient=recipient,
            recipient_beneficiary=client,
            message=message,
            status='sent',
            sent_by=request.user
        )
        
        if comm_type == 'sms':
            from accounting_app.sms_service import send_sms_to_client
            result = send_sms_to_client(request.user.userprofile, recipient, message)
            if result.get('status') == 'success':
                comm.status = 'delivered'
            else:
                comm.status = 'failed'
            comm.notes = result.get('message', '')
            comm.save()
        
        messages.success(request, f"Communication sent successfully")
        
    return redirect("communication_list")


@login_required
def whatsapp_messages(request):
    whatsapp_logs = CommunicationLog.objects.filter(communication_type='whatsapp').order_by('-sent_at')
    return render(request, "accounting_app/whatsapp_messages.html", {
        "whatsapp_logs": whatsapp_logs
    })


@login_required
def whatsapp_web_setup(request):
    if request.method == 'POST':
        whatsapp_number = request.POST.get('whatsapp_number')
        action = request.POST.get('action')
        
        user_profile = request.user.userprofile
        user_profile.whatsapp_number = whatsapp_number
        
        if action == 'generate_qr':
            # Generate QR code for WhatsApp Web
            # This would integrate with WhatsApp Web protocol
            # For now, we'll provide instructions
            user_profile.whatsapp_status = 'pending'
            user_profile.save()
            messages.info(request, f"QR code generated for {whatsapp_number}. Scan with WhatsApp to link.")
            return redirect('whatsapp_web_setup')
        
        elif action == 'logout':
            user_profile.whatsapp_status = 'pending'
            user_profile.whatsapp_session_token = ''
            user_profile.whatsapp_qr_code = ''
            user_profile.save()
            messages.success(request, "WhatsApp disconnected successfully.")
            return redirect('whatsapp_web_setup')
        
        user_profile.save()
        messages.success(request, "WhatsApp number saved.")
        return redirect('whatsapp_web_setup')
    
    return render(request, "accounting_app/whatsapp_web_setup.html", {
        "user_profile": request.user.userprofile
    })


@login_required
def chat_list(request):
    from django.db.models import Max, Case, When, IntegerField
    conversations = []
    other_users = User.objects.exclude(id=request.user.id).order_by("username")
    for user in other_users:
        last_msg = UserMessage.objects.filter(
            (Q(sender=request.user) & Q(recipient=user)) |
            (Q(sender=user) & Q(recipient=request.user))
        ).order_by("-sent_at").first()
        unread = UserMessage.objects.filter(sender=user, recipient=request.user, is_read=False).count()
        if last_msg:
            conversations.append({
                "user": user,
                "last_message": last_msg.message[:100],
                "last_time": last_msg.sent_at,
                "unread": unread,
            })
        else:
            conversations.append({
                "user": user,
                "last_message": "No messages yet",
                "last_time": None,
                "unread": 0,
            })
    conversations.sort(key=lambda x: x["last_time"] or timezone.datetime.min.replace(tzinfo=timezone.get_current_timezone()), reverse=True)
    all_users = User.objects.filter(is_active=True).exclude(id=request.user.id).order_by("username")
    return render(request, "accounting_app/chat_list.html", {
        "conversations": conversations,
        "all_users": all_users,
        "total_unread": UserMessage.get_unread_count(request.user),
    })


@login_required
def chat_conversation(request, user_id):
    other_user = get_object_or_404(User, id=user_id)
    messages = UserMessage.get_conversation(request.user, other_user)
    UserMessage.objects.filter(sender=other_user, recipient=request.user, is_read=False).update(is_read=True)
    return render(request, "accounting_app/chat_conversation.html", {
        "other_user": other_user,
        "messages": messages,
    })


@login_required
def chat_api_messages(request, user_id):
    other_user = get_object_or_404(User, id=user_id)
    last_id = request.GET.get("last_id", 0)
    msgs = UserMessage.get_conversation(request.user, other_user).filter(id__gt=last_id)
    data = []
    for m in msgs:
        data.append({
            "id": m.id,
            "sender": m.sender.username,
            "message": m.message,
            "sent_at": m.sent_at.strftime("%H:%M"),
            "is_mine": m.sender == request.user,
        })
    if data:
        UserMessage.objects.filter(sender=other_user, recipient=request.user, is_read=False).update(is_read=True)
    return JsonResponse({"messages": data})


@login_required
def chat_api_send(request, user_id):
    if request.method == "POST":
        import json
        body = json.loads(request.body)
        message_text = body.get("message", "").strip()
        if message_text:
            other_user = get_object_or_404(User, id=user_id)
            msg = UserMessage.objects.create(
                sender=request.user,
                recipient=other_user,
                message=message_text,
            )
            return JsonResponse({
                "id": msg.id,
                "sender": msg.sender.username,
                "message": msg.message,
                "sent_at": msg.sent_at.strftime("%H:%M"),
                "is_mine": True,
            })
    return JsonResponse({"error": "Invalid request"}, status=400)


@login_required
def chat_api_unread_count(request):
    return JsonResponse({"unread": UserMessage.get_unread_count(request.user)})


@login_required
def call_initiate(request, user_id):
    if request.method == "POST":
        import json
        body = json.loads(request.body)
        call_type = body.get("call_type", "voice")
        caller_offer = body.get("offer", "")
        callee = get_object_or_404(User, id=user_id)
        if callee.id == request.user.id:
            return JsonResponse({"error": "Cannot call yourself"}, status=400)
        UserCall.objects.filter(caller=request.user, callee=callee, status="ringing").update(status="missed")
        call = UserCall.objects.create(
            caller=request.user,
            callee=callee,
            call_type=call_type,
            caller_offer=caller_offer,
        )
        return JsonResponse({
            "call_id": call.id,
            "status": "ringing",
        })
    return JsonResponse({"error": "Invalid"}, status=400)


@login_required
def call_check_incoming(request):
    incoming = UserCall.objects.filter(callee=request.user, status="ringing").order_by("-started_at").first()
    if not incoming:
        return JsonResponse({"has_call": False})
    return JsonResponse({
        "has_call": True,
        "call_id": incoming.id,
        "caller_id": incoming.caller.id,
        "caller_name": incoming.caller.username,
        "caller_full_name": incoming.caller.get_full_name() or incoming.caller.username,
        "call_type": incoming.call_type,
        "offer": incoming.caller_offer,
    })


@login_required
def call_answer(request, call_id):
    if request.method == "POST":
        import json
        body = json.loads(request.body)
        call = get_object_or_404(UserCall, id=call_id)
        if call.callee != request.user or call.status != "ringing":
            return JsonResponse({"error": "Invalid"}, status=400)
        call.status = "accepted"
        call.callee_answer = body.get("answer", "")
        call.callee_ice = body.get("ice", "")
        call.started_at = timezone.now()
        call.save()
        return JsonResponse({"status": "accepted", "offer": call.caller_offer, "caller_ice": call.caller_ice})
    return JsonResponse({"error": "Invalid"}, status=400)


@login_required
def call_reject(request, call_id):
    call = get_object_or_404(UserCall, id=call_id)
    if call.callee != request.user:
        return JsonResponse({"error": "Invalid"}, status=400)
    call.status = "rejected"
    call.save()
    return JsonResponse({"status": "rejected"})


@login_required
def call_signal(request, call_id):
    import json
    if request.method == "POST":
        body = json.loads(request.body)
        call = get_object_or_404(UserCall, id=call_id)
        action = body.get("action")
        if action == "ice":
            if call.caller == request.user:
                call.caller_ice = body.get("candidate", "")
            else:
                call.callee_ice = body.get("candidate", "")
            call.save()
            return JsonResponse({"ok": True})
        elif action == "offer":
            call.caller_offer = body.get("offer", "")
            call.save()
            return JsonResponse({"ok": True})
        elif action == "answer":
            call.callee_answer = body.get("answer", "")
            call.save()
            return JsonResponse({"ok": True})
    call = get_object_or_404(UserCall, id=call_id)
    if call.caller == request.user:
        other = call.callee
    else:
        other = call.caller
    return JsonResponse({
        "status": call.status,
        "offer": call.caller_offer if other == request.user else "",
        "answer": call.callee_answer if other == request.user else "",
        "caller_ice": call.caller_ice if other == request.user else "",
        "callee_ice": call.callee_ice if other == request.user else "",
    })


@login_required
def call_end(request, call_id):
    call = get_object_or_404(UserCall, id=call_id)
    if call.caller != request.user and call.callee != request.user:
        return JsonResponse({"error": "Invalid"}, status=400)
    call.status = "ended"
    call.ended_at = timezone.now()
    if call.started_at:
        call.duration = int((call.ended_at - call.started_at).total_seconds())
    call.save()
    return JsonResponse({"status": "ended", "duration": call.duration})


@login_required
def call_get_caller_info(request, call_id):
    call = get_object_or_404(UserCall, id=call_id)
    return JsonResponse({
        "call_id": call.id,
        "caller_id": call.caller.id,
        "caller_name": call.caller.username,
        "caller_full_name": call.caller.get_full_name() or call.caller.username,
        "call_type": call.call_type,
        "status": call.status,
        "offer": call.caller_offer,
    })


# ======== MODULE 1: DATA BACKUP, EXPORT & IMPORT ========

def _is_super_admin(user):
    return user.is_superuser or (
        hasattr(user, 'userprofile') and user.userprofile.role == "admin"
    )


EXPORT_MODELS = {
    "accounts": Account,
    "beneficiaries": Beneficiary,
    "vendors": Vendor,
    "invoices": Invoice,
    "invoice_items": InvoiceItem,
    "payments": Payment,
    "expenses": Expense,
    "expense_items": ExpenseItem,
    "journal_entries": JournalEntry,
    "journal_entry_lines": JournalEntryLine,
    "budgets": Budget,
    "budget_lines": BudgetLine,
    "schemes": Scheme,
    "villages": Village,
    "village_populations": VillagePopulation,
    "board_of_trustees": BoardOfTrustees,
    "general_assembly_members": GeneralAssemblyMember,
    "employees": Employee,
    "employee_salaries": EmployeeSalary,
    "users": User,
    "user_profiles": UserProfile,
    "communication_logs": CommunicationLog,
    "user_messages": UserMessage,
    "activity_logs": ActivityLog,
    "reports": Report,
    "opening_balances": OpeningBalance,
    "balance_history": BalanceHistory,
    "gallery_images": GalleryImage,
    "services": Service,
    "landing_page_settings": LandingPageSettings,
    "tax_rates": TaxRate,
}

EXPORT_GROUPS = [
    {
        "name": "Products",
        "keys": ["accounts", "services"],
        "icon": "bi-box-seam",
    },
    {
        "name": "Categories",
        "keys": [],
        "icon": "bi-tags",
    },
    {
        "name": "Users & Roles",
        "keys": ["users", "user_profiles"],
        "icon": "bi-people",
    },
    {
        "name": "Sales",
        "keys": ["invoices", "invoice_items", "payments"],
        "icon": "bi-cart",
    },
    {
        "name": "Stock Transactions",
        "keys": [],
        "icon": "bi-arrow-left-right",
    },
    {
        "name": "Chat Messages",
        "keys": ["user_messages", "communication_logs"],
        "icon": "bi-chat-dots",
    },
    {
        "name": "Notifications",
        "keys": [],
        "icon": "bi-bell",
    },
    {
        "name": "Settings & Shops",
        "keys": ["landing_page_settings", "tax_rates", "gallery_images", "schemes", "services"],
        "icon": "bi-gear",
    },
]


def _serialize_model(model_class):
    data = []
    for obj in model_class.objects.all():
        record = {}
        for field in obj._meta.fields:
            value = getattr(obj, field.name, None)
            if hasattr(value, 'isoformat'):
                value = value.isoformat()
            elif hasattr(value, 'pk'):
                value = value.pk
            elif isinstance(value, Decimal):
                value = str(value)
            record[field.name] = value
        data.append(record)
    return data


def _identify_record(model_class, record):
    pk_field = model_class._meta.pk.name
    pk_value = record.get(pk_field)
    unique_fields = []

    for field in model_class._meta.fields:
        if field.unique and field.name != pk_field:
            unique_fields.append(field.name)

    for uf in unique_fields:
        val = record.get(uf)
        if val:
            try:
                obj = model_class.objects.filter(**{uf: val}).first()
                if obj:
                    return obj
            except Exception:
                pass
    return None


@login_required
def data_migration(request):
    if not _is_super_admin(request.user):
        messages.error(request, "Access denied. Super Administrator access required.")
        return redirect("dashboard")

    export_history = DataMigrationLog.objects.filter(direction="export")[:20]
    import_history = DataMigrationLog.objects.filter(direction="import")[:20]

    selected_groups = request.POST.getlist("groups")

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "export":
            log = DataMigrationLog.objects.create(
                direction="export",
                status="in_progress",
                performed_by=request.user,
            )
            try:
                import io, zipfile

                selected_model_keys = set()
                for group in EXPORT_GROUPS:
                    if group["name"] in selected_groups or not selected_groups:
                        selected_model_keys.update(group["keys"])

                models_to_export = {k: v for k, v in EXPORT_MODELS.items() if k in selected_model_keys}
                if not models_to_export:
                    models_to_export = dict(EXPORT_MODELS)

                buffer = io.BytesIO()
                with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                    for name, model_class in models_to_export.items():
                        data = _serialize_model(model_class)
                        zf.writestr(f"{name}.json", json.dumps(data, default=str, indent=2))
                    zf.writestr("export_info.json", json.dumps({
                        "exported_at": timezone.now().isoformat(),
                        "exported_by": request.user.username,
                        "model_count": len(models_to_export),
                    }, indent=2))

                from django.core.files.base import ContentFile
                filename = f"backup_{timezone.now().strftime('%Y%m%d_%H%M%S')}.zip"
                log.file.save(filename, ContentFile(buffer.getvalue()))
                log.status = "completed"
                log.completed_at = timezone.now()
                summary = {}
                for name, model_class in models_to_export.items():
                    summary[name] = model_class.objects.count()
                log.summary = summary
                log.save()

                messages.success(request, f"Data exported successfully. {len(models_to_export)} tables exported.")
            except Exception as e:
                log.status = "failed"
                log.notes = str(e)
                log.save()
                messages.error(request, f"Export failed: {e}")

            return redirect("data_migration")

        elif action == "import":
            mode = request.POST.get("mode", "add_new")
            uploaded_file = request.FILES.get("file")

            if not uploaded_file:
                messages.error(request, "Please select a backup file to import.")
                return redirect("data_migration")

            if not uploaded_file.name.endswith(".zip"):
                messages.error(request, "Only ZIP files are supported.")
                return redirect("data_migration")

            import io, zipfile, tempfile, shutil

            tmp_path = tempfile.mktemp(suffix=".zip")
            with open(tmp_path, "wb") as f:
                for chunk in uploaded_file.chunks():
                    f.write(chunk)

            log = DataMigrationLog.objects.create(
                direction="import",
                mode=mode,
                status="in_progress",
                performed_by=request.user,
                notes=f"Import mode: {dict(DataMigrationLog.IMPORT_MODES).get(mode, mode)}",
            )
            from django.core.files.base import ContentFile
            with open(tmp_path, "rb") as f:
                log.file.save(uploaded_file.name, ContentFile(f.read()))
            log.save()

            try:
                preview_data = {}
                with zipfile.ZipFile(tmp_path, "r") as zf:
                    for name in zf.namelist():
                        if name.endswith(".json") and name != "export_info.json":
                            model_key = name.replace(".json", "")
                            content = json.loads(zf.read(name))
                            preview_data[model_key] = {
                                "total": len(content),
                                "sample": content[:3] if content else [],
                            }

                request.session["import_preview"] = {
                    "log_id": log.id,
                    "mode": mode,
                    "file_path": tmp_path,
                }

                return render(request, "accounting_app/data_migration.html", {
                    "export_history": export_history,
                    "import_history": import_history,
                    "export_groups": EXPORT_GROUPS,
                    "selected_groups": selected_groups,
                    "preview_data": preview_data,
                    "mode": mode,
                    "show_preview": True,
                    "log_id": log.id,
                })

            except Exception as e:
                log.status = "failed"
                log.notes = str(e)
                log.save()
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass
                messages.error(request, f"Import preview failed: {e}")
                return redirect("data_migration")

        elif action == "execute_import":
            log_id = request.POST.get("log_id")
            mode = request.POST.get("mode", "add_new")
            file_path = request.POST.get("file_path", "")

            if not log_id:
                messages.error(request, "Import session not found.")
                return redirect("data_migration")

            log = get_object_or_404(DataMigrationLog, id=log_id)

            if not os.path.exists(file_path):
                messages.error(request, "Backup file not found. Please upload again.")
                return redirect("data_migration")

            from django.db import transaction

            try:
                import io, zipfile
                summary = {}
                with transaction.atomic():
                    with zipfile.ZipFile(file_path, "r") as zf:
                        for name in zf.namelist():
                            if name.endswith(".json") and name != "export_info.json":
                                model_key = name.replace(".json", "")
                                model_class = EXPORT_MODELS.get(model_key)
                                if not model_class:
                                    continue

                                records = json.loads(zf.read(name))
                                created = 0
                                updated = 0
                                skipped = 0

                                for record in records:
                                    pk_field = model_class._meta.pk.name
                                    pk_value = record.get(pk_field)

                                    if mode == "replace":
                                        model_class.objects.all().delete()
                                        pk_fields = [f for f in model_class._meta.fields if f.primary_key]
                                        if pk_fields:
                                            record.pop(pk_field, None)
                                        obj = model_class(**record)
                                        obj.save()
                                        created += 1

                                    elif mode == "merge":
                                        existing = _identify_record(model_class, record)
                                        if existing:
                                            changed = False
                                            for field in model_class._meta.fields:
                                                if field.name in record and field.name != pk_field:
                                                    setattr(existing, field.name, record[field.name])
                                                    changed = True
                                            if changed:
                                                existing.save()
                                                updated += 1
                                            else:
                                                skipped += 1
                                        else:
                                            record.pop(pk_field, None)
                                            obj = model_class(**record)
                                            obj.save()
                                            created += 1

                                    else:
                                        existing = _identify_record(model_class, record)
                                        if existing:
                                            skipped += 1
                                        else:
                                            record.pop(pk_field, None)
                                            obj = model_class(**record)
                                            obj.save()
                                            created += 1

                                summary[model_key] = {
                                    "total": len(records),
                                    "created": created,
                                    "updated": updated,
                                    "skipped": skipped,
                                }

                log.status = "completed"
                log.completed_at = timezone.now()
                log.summary = summary
                log.save()

                total_created = sum(s["created"] for s in summary.values())
                total_updated = sum(s["updated"] for s in summary.values())
                messages.success(request, f"Import completed. {total_created} created, {total_updated} updated.")

            except Exception as e:
                log.status = "failed"
                log.notes = str(e)
                log.save()
                messages.error(request, f"Import failed: {e}")

            try:
                os.unlink(file_path)
            except Exception:
                pass

            return redirect("data_migration")

    return render(request, "accounting_app/data_migration.html", {
        "export_history": export_history,
        "import_history": import_history,
        "export_groups": EXPORT_GROUPS,
        "selected_groups": selected_groups,
    })


# ======== MODULE 2: SYSTEM UPDATE MANAGER ========

@login_required
def ai_assistant(request):
    ai = GroqAIService()
    return render(request, "accounting_app/ai_assistant.html", {
        "configured": ai.is_available()
    })

@login_required
def ai_chat_api(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    import json as json_module

    try:
        body = json_module.loads(request.body)
    except (ValueError, TypeError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    message = body.get("message", "").strip()
    history = body.get("history", [])

    if not message:
        return JsonResponse({"error": "Message is required"}, status=400)

    ai = GroqAIService(user=request.user)
    if not ai.is_available():
        return JsonResponse({
            "type": "error",
            "message": "Groq AI is not configured. Ask the administrator to set the GROQ_API_KEY in the .env file. Get a free key at https://console.groq.com/keys"
        })

    result = ai.chat(message, history)
    
    action_result = None
    if result.get("action"):
        action_result = ai.execute_action(result["action"])

    response_data = {
        "type": result.get("type", "response"),
        "message": result.get("message", ""),
    }
    if action_result:
        response_data["action_result"] = action_result

    return JsonResponse(response_data)


def system_updates(request):
    if not _is_super_admin(request.user):
        messages.error(request, "Access denied. Super Administrator access required.")
        return redirect("dashboard")

    current_version = SystemVersion.objects.filter(is_current=True).first()
    available_versions = SystemVersion.objects.filter(is_current=False)
    update_history = SystemUpdateLog.objects.all()[:20]

    if not current_version:
        from accounting_app import __version__
        current_version = SystemVersion.objects.create(
            version=__version__,
            release_date=timezone.now().date(),
            changelog="Initial system version.",
            is_current=True,
        )

    context = {
        "current_version": current_version,
        "available_versions": available_versions,
        "update_history": update_history,
    }

    if request.method == "POST":
        action = request.POST.get("action")
        version_id = request.POST.get("version_id")

        if action == "install" and version_id:
            version = get_object_or_404(SystemVersion, id=version_id)
            upload_log = SystemUpdateLog.objects.create(
                version=version.version,
                previous_version=current_version.version,
                status="in_progress",
                installed_by=request.user,
            )

            try:
                from django.db import transaction

                with transaction.atomic():
                    backup_path = None
                    try:
                        backup_path = _create_system_backup()
                        if backup_path:
                            from django.core.files.base import ContentFile
                            with open(backup_path, "rb") as f:
                                upload_log.backup_file.save(
                                    f"pre_update_backup_{timezone.now().strftime('%Y%m%d_%H%M%S')}.zip",
                                    ContentFile(f.read())
                                )
                    except Exception as e:
                        pass

                    current_version.is_current = False
                    current_version.save()

                    version.is_current = True
                    version.save()

                    from django.db.migrations.executor import MigrationExecutor
                    from django.db import connections
                    connection = connections['default']
                    executor = MigrationExecutor(connection)
                    plan = executor.migration_plan(executor.loader.graph.leaf_nodes())
                    if plan:
                        executor.migrate(executor.loader.graph.leaf_nodes())

                upload_log.status = "completed"
                upload_log.completed_at = timezone.now()
                upload_log.summary = {
                    "previous_version": current_version.version,
                    "new_version": version.version,
                    "backup_created": backup_path is not None,
                }
                upload_log.save()

                try:
                    if backup_path and os.path.exists(backup_path):
                        os.unlink(backup_path)
                except Exception:
                    pass

                messages.success(request, f"System updated to version {version.version} successfully!")

            except Exception as e:
                upload_log.status = "failed"
                upload_log.error_message = str(e)
                upload_log.save()
                messages.error(request, f"Update failed: {e}")

            return redirect("system_updates")

        elif action == "rollback":
            update_log_id = request.POST.get("update_log_id")
            if not update_log_id:
                messages.error(request, "No update log specified.")
                return redirect("system_updates")

            update_log = get_object_or_404(SystemUpdateLog, id=update_log_id)
            if not update_log.backup_file:
                messages.error(request, "No backup file available for rollback.")
                return redirect("system_updates")

            try:
                update_log.status = "rolled_back"
                update_log.save()

                prev_version = SystemVersion.objects.filter(
                    version=update_log.previous_version
                ).first()
                if prev_version:
                    current_version.is_current = False
                    current_version.save()
                    prev_version.is_current = True
                    prev_version.save()

                messages.success(request, f"Rolled back to version {update_log.previous_version}.")
            except Exception as e:
                messages.error(request, f"Rollback failed: {e}")

            return redirect("system_updates")

        elif action == "upload_package":
            version = request.POST.get("version")
            release_date = request.POST.get("release_date")
            changelog = request.POST.get("changelog")
            package_file = request.FILES.get("package_file")

            if not version:
                messages.error(request, "Version number is required.")
                return redirect("system_updates")

            SystemVersion.objects.create(
                version=version,
                release_date=release_date or timezone.now().date(),
                changelog=changelog or "",
                package_file=package_file,
            )
            messages.success(request, f"Update package for version {version} uploaded.")

            return redirect("system_updates")

    return render(request, "accounting_app/system_updates.html", context)


def _create_system_backup():
    import io, zipfile, tempfile
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, model_class in EXPORT_MODELS.items():
            data = _serialize_model(model_class)
            zf.writestr(f"data/{name}.json", json.dumps(data, default=str, indent=2))
        zf.writestr("backup_info.json", json.dumps({
            "backup_at": timezone.now().isoformat(),
            "version": SystemVersion.objects.filter(is_current=True).first().version if SystemVersion.objects.filter(is_current=True).exists() else "unknown",
        }, indent=2))

    tmp_path = tempfile.mktemp(suffix=".zip")
    with open(tmp_path, "wb") as f:
        f.write(buffer.getvalue())
    return tmp_path
